"""
Profile Automation Tool
A professional desktop application for generating elevation profiles between two geographic points.

Developed by Omid Zanganeh
Version: 1.5.0 - Enhanced with File Import (CSV, Shapefiles)
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import tkintermapview
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D
import pandas as pd
import numpy as np
import requests
import json
import logging
import threading
import time
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import math
from pathlib import Path
from PIL import Image, ImageTk
import os


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%H:%M:%S'
)


class TextHandler(logging.Handler):
    """Custom logging handler to redirect logs to text widget"""
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        
    def emit(self, record):
        msg = self.format(record)
        def append():
            self.text_widget.configure(state='normal')
            self.text_widget.insert(tk.END, msg + '\n')
            self.text_widget.see(tk.END)
            self.text_widget.configure(state='disabled')
        self.text_widget.after(0, append)


class CollapsibleFrame(ctk.CTkFrame):
    """A collapsible frame widget with expand/collapse functionality"""
    
    def __init__(self, parent, title="", start_collapsed=False, **kwargs):
        super().__init__(parent, **kwargs)
        
        self.is_collapsed = start_collapsed
        
        # Header button
        self.header_btn = ctk.CTkButton(
            self,
            text=f"{'▶' if start_collapsed else '▼'} {title}",
            command=self.toggle,
            fg_color="transparent",
            hover_color="gray25" if ctk.get_appearance_mode() == "Dark" else "gray80",
            text_color=("gray10", "gray90"),
            anchor="w",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.header_btn.pack(fill=tk.X, pady=(5, 0))
        
        # Content frame
        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        if not start_collapsed:
            self.content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    def toggle(self):
        """Toggle between expanded and collapsed states"""
        if self.is_collapsed:
            self.expand()
        else:
            self.collapse()
    
    def collapse(self):
        """Collapse the content frame"""
        self.content_frame.pack_forget()
        self.is_collapsed = True
        # Update arrow
        current_text = self.header_btn.cget("text")
        self.header_btn.configure(text=current_text.replace("▼", "▶"))
    
    def expand(self):
        """Expand the content frame"""
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.is_collapsed = False
        # Update arrow
        current_text = self.header_btn.cget("text")
        self.header_btn.configure(text=current_text.replace("▶", "▼"))
    
    def get_content_frame(self):
        """Return the content frame for adding widgets"""
        return self.content_frame


class RunningLineConfigWindow(ctk.CTkToplevel):
    """Dedicated window for running line configuration with live 2D chart"""
    
    def __init__(self, parent, profile_data, break_points, callback):
        super().__init__(parent)
        
        self.parent = parent
        self.profile_data = profile_data
        self.break_points = list(break_points)  # Make a copy
        self.callback = callback
        
        # Window setup
        self.title("Running Line Configuration - 2D Chart Editor")
        self.geometry("1400x900")
        
        # Make it modal
        self.transient(parent)
        self.grab_set()
        
        # Center the window
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 700
        y = (self.winfo_screenheight() // 2) - 450
        self.geometry(f"1400x900+{x}+{y}")
        
        # Configuration values
        self.break_point_mode = False
        self.depth_edit_mode = False
        self.selected_bp_index = None
        self.sensitivity = 5.0
        self.max_segments = 5
        self.running_line_depth_inches = parent.running_line_depth_inches
        self.depth_zones = list(parent.depth_zones) if hasattr(parent, 'depth_zones') else []
        self.depth_zone_start = None
        
        # Drag state
        self.dragging_bp_index = None
        self.drag_start_x = None
        
        # Curve smoothing state
        self.curve_enabled = False
        
        self._setup_ui()
        self._draw_chart()
        
    def _setup_ui(self):
        """Setup the UI layout"""
        # Main container
        main_container = ctk.CTkFrame(self)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left side - Chart
        chart_frame = ctk.CTkFrame(main_container)
        chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        chart_header = ctk.CTkLabel(
            chart_frame,
            text="📊 Live Profile Chart",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        chart_header.pack(pady=(10, 5))
        
        # Chart canvas placeholder
        self.chart_container = ctk.CTkFrame(chart_frame)
        self.chart_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Right side - Controls
        controls_frame = ctk.CTkFrame(main_container, width=400)
        controls_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(5, 0))
        controls_frame.pack_propagate(False)
        
        controls_header = ctk.CTkLabel(
            controls_frame,
            text="⚙️ Running Line Controls",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        controls_header.pack(pady=(10, 10))
        
        # Scrollable frame for controls
        controls_scroll = ctk.CTkScrollableFrame(controls_frame)
        controls_scroll.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 10))
        
        # === Depth Configuration ===
        depth_section = ctk.CTkFrame(controls_scroll, fg_color="gray20", corner_radius=8)
        depth_section.pack(fill=tk.X, pady=5)
        
        depth_header = ctk.CTkLabel(
            depth_section,
            text="🔧 Depth Configuration",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        depth_header.pack(pady=(10, 5))
        
        # Default depth
        depth_frame = ctk.CTkFrame(depth_section, fg_color="transparent")
        depth_frame.pack(fill=tk.X, padx=10, pady=5)
        
        depth_label = ctk.CTkLabel(
            depth_frame,
            text="Default Depth (inches):",
            font=ctk.CTkFont(size=11)
        )
        depth_label.pack(anchor="w")
        
        depth_input_frame = ctk.CTkFrame(depth_frame, fg_color="transparent")
        depth_input_frame.pack(fill=tk.X, pady=2)
        
        self.depth_entry = ctk.CTkEntry(depth_input_frame, width=100)
        self.depth_entry.insert(0, str(self.running_line_depth_inches))
        self.depth_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        update_depth_btn = ctk.CTkButton(
            depth_input_frame,
            text="Apply",
            command=self._update_depth,
            width=60,
            height=28
        )
        update_depth_btn.pack(side=tk.LEFT)
        
        # Quick presets
        preset_frame = ctk.CTkFrame(depth_section, fg_color="transparent")
        preset_frame.pack(fill=tk.X, padx=10, pady=5)
        
        preset_label = ctk.CTkLabel(
            preset_frame,
            text="Quick Presets:",
            font=ctk.CTkFont(size=10),
            text_color="gray70"
        )
        preset_label.pack(anchor="w")
        
        preset_btns = ctk.CTkFrame(preset_frame, fg_color="transparent")
        preset_btns.pack(fill=tk.X, pady=2)
        
        for depth_val in [36, 42, 48, 60, 72]:
            btn = ctk.CTkButton(
                preset_btns,
                text=f'{depth_val}"',
                width=50,
                height=25,
                command=lambda d=depth_val: self._set_preset_depth(d)
            )
            btn.pack(side=tk.LEFT, padx=2)
        
        # Variable depth zones
        var_depth_frame = ctk.CTkFrame(depth_section, fg_color="transparent")
        var_depth_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        var_label = ctk.CTkLabel(
            var_depth_frame,
            text="Variable Depth Zones:",
            font=ctk.CTkFont(size=10, weight="bold")
        )
        var_label.pack(anchor="w")
        
        var_btn_frame = ctk.CTkFrame(var_depth_frame, fg_color="transparent")
        var_btn_frame.pack(fill=tk.X, pady=2)
        
        self.depth_zone_btn = ctk.CTkButton(
            var_btn_frame,
            text="🖱️ Click Chart to Add Zone",
            command=self._toggle_depth_zone_mode,
            fg_color="#8E44AD",
            hover_color="#7D3C98",
            height=30
        )
        self.depth_zone_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        clear_zones_btn = ctk.CTkButton(
            var_btn_frame,
            text="Clear",
            command=self._clear_depth_zones,
            fg_color="darkred",
            hover_color="red",
            width=60,
            height=30
        )
        clear_zones_btn.pack(side=tk.LEFT)
        
        # === Break Point Controls ===
        bp_section = ctk.CTkFrame(controls_scroll, fg_color="gray20", corner_radius=8)
        bp_section.pack(fill=tk.X, pady=5)
        
        bp_header = ctk.CTkLabel(
            bp_section,
            text="📍 Break Points (Multi-Segment)",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        bp_header.pack(pady=(10, 5))
        
        # Break point mode button
        self.bp_mode_btn = ctk.CTkButton(
            bp_section,
            text="🖱️ Click Chart to Add Break Point",
            command=self._toggle_break_point_mode,
            fg_color="gray40",
            hover_color="gray30",
            height=35
        )
        self.bp_mode_btn.pack(fill=tk.X, padx=10, pady=5)
        
        # Curve smoothing toggle
        self.curve_toggle_btn = ctk.CTkButton(
            bp_section,
            text="🔄 Path: Linear (Sharp Angles)",
            command=self._toggle_curve_smoothing,
            fg_color="#FF6B35",
            hover_color="#E85D2A",
            height=35
        )
        self.curve_toggle_btn.pack(fill=tk.X, padx=10, pady=5)
        
        # Action buttons
        bp_btn_frame = ctk.CTkFrame(bp_section, fg_color="transparent")
        bp_btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.bp_clear_btn = ctk.CTkButton(
            bp_btn_frame,
            text="Clear All",
            command=self._clear_break_points,
            fg_color="darkred",
            hover_color="red",
            width=110
        )
        self.bp_clear_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.bp_auto_btn = ctk.CTkButton(
            bp_btn_frame,
            text="Auto-Optimize",
            command=self._auto_optimize,
            fg_color="blue",
            hover_color="darkblue",
            width=110
        )
        self.bp_auto_btn.pack(side=tk.LEFT)
        
        # === Auto-Optimize Settings ===
        auto_section = ctk.CTkFrame(controls_scroll, fg_color="gray20", corner_radius=8)
        auto_section.pack(fill=tk.X, pady=5)
        
        auto_header = ctk.CTkLabel(
            auto_section,
            text="🔧 Auto-Optimize Settings",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        auto_header.pack(pady=(10, 5))
        
        # Sensitivity slider
        sens_frame = ctk.CTkFrame(auto_section, fg_color="transparent")
        sens_frame.pack(fill=tk.X, padx=10, pady=5)
        
        sens_label = ctk.CTkLabel(
            sens_frame,
            text="Sensitivity (slope change):",
            font=ctk.CTkFont(size=11)
        )
        sens_label.pack(anchor="w")
        
        sens_slider_frame = ctk.CTkFrame(sens_frame, fg_color="transparent")
        sens_slider_frame.pack(fill=tk.X, pady=2)
        
        self.sensitivity_slider = ctk.CTkSlider(
            sens_slider_frame,
            from_=0.5,
            to=15,
            number_of_steps=29,
            command=self._on_sensitivity_change
        )
        self.sensitivity_slider.set(5)
        self.sensitivity_slider.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.sensitivity_label = ctk.CTkLabel(
            sens_slider_frame,
            text="5.0%",
            font=ctk.CTkFont(size=11),
            width=50
        )
        self.sensitivity_label.pack(side=tk.LEFT)
        
        # Max segments slider
        max_seg_frame = ctk.CTkFrame(auto_section, fg_color="transparent")
        max_seg_frame.pack(fill=tk.X, padx=10, pady=(5, 10))
        
        max_seg_label = ctk.CTkLabel(
            max_seg_frame,
            text="Max break points:",
            font=ctk.CTkFont(size=11)
        )
        max_seg_label.pack(anchor="w")
        
        max_slider_frame = ctk.CTkFrame(max_seg_frame, fg_color="transparent")
        max_slider_frame.pack(fill=tk.X, pady=2)
        
        self.max_segments_slider = ctk.CTkSlider(
            max_slider_frame,
            from_=2,
            to=10,
            number_of_steps=8,
            command=self._on_max_segments_change
        )
        self.max_segments_slider.set(5)
        self.max_segments_slider.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.max_segments_label = ctk.CTkLabel(
            max_slider_frame,
            text="5",
            font=ctk.CTkFont(size=11),
            width=50
        )
        self.max_segments_label.pack(side=tk.LEFT)
        
        # === Break Points List ===
        list_section = ctk.CTkFrame(controls_scroll, fg_color="gray20", corner_radius=8)
        list_section.pack(fill=tk.BOTH, expand=True, pady=5)
        
        list_header = ctk.CTkLabel(
            list_section,
            text="📋 Break Points List",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        list_header.pack(pady=(10, 5))
        
        self.bp_listbox = tk.Listbox(
            list_section,
            height=10,
            bg="gray25",
            fg="white",
            selectbackground="blue",
            selectforeground="white",
            font=("Courier", 9)
        )
        self.bp_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 5))
        self.bp_listbox.bind('<<ListboxSelect>>', self._on_bp_select)
        
        # Edit buttons
        edit_btn_frame = ctk.CTkFrame(list_section, fg_color="transparent")
        edit_btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        self.bp_delete_btn = ctk.CTkButton(
            edit_btn_frame,
            text="Delete",
            command=self._delete_selected_break_point,
            fg_color="darkred",
            hover_color="red",
            width=90
        )
        self.bp_delete_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.bp_edit_btn = ctk.CTkButton(
            edit_btn_frame,
            text="Move",
            command=self._edit_selected_break_point,
            fg_color="darkorange",
            hover_color="orange",
            width=90
        )
        self.bp_edit_btn.pack(side=tk.LEFT)
        
        # === Action Buttons ===
        action_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        action_frame.pack(fill=tk.X, padx=5, pady=(5, 10))
        
        self.apply_btn = ctk.CTkButton(
            action_frame,
            text="✅ Apply Changes",
            command=self._apply_changes,
            fg_color="green",
            hover_color="darkgreen",
            height=40,
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.apply_btn.pack(fill=tk.X, pady=(0, 5))
        
        self.cancel_btn = ctk.CTkButton(
            action_frame,
            text="Cancel",
            command=self.destroy,
            fg_color="gray40",
            hover_color="gray30",
            height=35
        )
        self.cancel_btn.pack(fill=tk.X)
        
        self._update_break_points_display()
    
    def _catmull_rom_spline(self, points, num_samples=50):
        """Generate smooth curve through points using Catmull-Rom spline
        
        Args:
            points: List of (x, y) tuples
            num_samples: Number of samples per segment
            
        Returns:
            Two lists: smoothed_x, smoothed_y
        """
        if len(points) < 2:
            return [], []
        
        if len(points) == 2:
            # Just linear for 2 points
            return [p[0] for p in points], [p[1] for p in points]
        
        smoothed_x = []
        smoothed_y = []
        
        # Add first point
        smoothed_x.append(points[0][0])
        smoothed_y.append(points[0][1])
        
        # Process each segment
        for i in range(len(points) - 1):
            # Get control points (with virtual endpoints)
            p0 = points[max(0, i - 1)]
            p1 = points[i]
            p2 = points[i + 1]
            p3 = points[min(len(points) - 1, i + 2)]
            
            # If at boundaries, use end points twice for better behavior
            if i == 0:
                p0 = p1
            if i == len(points) - 2:
                p3 = p2
            
            # Generate samples for this segment
            for t_step in range(1, num_samples + 1):
                t = t_step / num_samples
                t2 = t * t
                t3 = t2 * t
                
                # Catmull-Rom basis functions
                x = 0.5 * (
                    (2 * p1[0]) +
                    (-p0[0] + p2[0]) * t +
                    (2*p0[0] - 5*p1[0] + 4*p2[0] - p3[0]) * t2 +
                    (-p0[0] + 3*p1[0] - 3*p2[0] + p3[0]) * t3
                )
                
                y = 0.5 * (
                    (2 * p1[1]) +
                    (-p0[1] + p2[1]) * t +
                    (2*p0[1] - 5*p1[1] + 4*p2[1] - p3[1]) * t2 +
                    (-p0[1] + 3*p1[1] - 3*p2[1] + p3[1]) * t3
                )
                
                smoothed_x.append(x)
                smoothed_y.append(y)
        
        return smoothed_x, smoothed_y
    
    def _draw_chart(self):
        """Draw the profile chart with running line - matches main 2D view style"""
        # Clear existing chart
        for widget in self.chart_container.winfo_children():
            widget.destroy()
        
        if not self.profile_data:
            return
        
        # Create figure with white background to match main view
        fig = Figure(figsize=(10, 6), dpi=100, facecolor='white')
        ax = fig.add_subplot(111)
        
        # Extract data
        distances = [p['distance'] for p in self.profile_data]
        elevations = [p['elevation'] for p in self.profile_data]
        running_line = [p.get('running_line', p['elevation'] - 3.5) for p in self.profile_data]
        
        # Plot Ground Elevation (Green like main view)
        ax.plot(distances, elevations, color='#2d7a2d', linewidth=2, 
               label='Ground Elevation (feet)', zorder=2)
        
        # Plot Running Line (Red) - with optional curve smoothing
        if self.break_points:
            # Multi-segment: color each segment differently
            segment_ids = [p.get('segment_id', 0) for p in self.profile_data]
            unique_segments = sorted(set(segment_ids))
            segment_colors = ['#cc0000', '#0066cc', '#ff6600', '#9900cc', '#00cc66']
            
            if self.curve_enabled and len(self.break_points) > 0:
                # Apply smooth curves across all segments
                # Collect key points (start, break points, end)
                key_points = [(distances[0], running_line[0])]
                
                for bp_dist, bp_elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
                    # Find running line elevation at break point
                    for p in self.profile_data:
                        if abs(p['distance'] - bp_dist) < 1.0:
                            key_points.append((bp_dist, p['running_line']))
                            break
                
                key_points.append((distances[-1], running_line[-1]))
                
                # Generate smooth curve through key points
                smooth_x, smooth_y = self._catmull_rom_spline(key_points, num_samples=30)
                
                # Plot the smooth curve
                ax.plot(smooth_x, smooth_y, color='#cc0000', linewidth=2.5,
                       label=f'Running Line (Curved, {self.running_line_depth_inches}")', zorder=3)
            else:
                # Linear segments (original behavior)
                for seg_id in unique_segments:
                    seg_distances = [p['distance'] for p in self.profile_data if p.get('segment_id', 0) == seg_id]
                    seg_running = [p['running_line'] for p in self.profile_data if p.get('segment_id', 0) == seg_id]
                    
                    color = segment_colors[seg_id % len(segment_colors)]
                    label = f'Segment {seg_id + 1}' if seg_id == unique_segments[0] else None
                    ax.plot(seg_distances, seg_running, color=color, linewidth=2.5,
                           label=label, zorder=3)
        else:
            # Single segment: red line (no smoothing needed)
            ax.plot(distances, running_line, color='#cc0000', linewidth=2.5,
                   label=f'Running Line ({self.running_line_depth_inches}")', zorder=3)
        
        # Plot break point markers
        if self.break_points:
            for bp_dist, bp_elev, grade, locked, note in self.break_points:
                # Find running line elevation at break point
                bp_running = None
                for p in self.profile_data:
                    if abs(p['distance'] - bp_dist) < 1.0:
                        bp_running = p['running_line']
                        break
                
                if bp_running is not None:
                    # Draw vertical line at break point
                    ax.axvline(x=bp_dist, color='#9900cc', linestyle='--', 
                             linewidth=1.5, alpha=0.7, zorder=5)
                    
                    # Draw marker on running line
                    ax.scatter([bp_dist], [bp_running], color='#9900cc', 
                             s=150, zorder=11, marker='D', edgecolors='white', 
                             linewidths=2, label='Break Point' if bp_dist == self.break_points[0][0] else None)
        
        # Styling to match main view
        ax.set_xlabel('Distance (feet)', fontsize=12, fontweight='bold')
        ax.set_ylabel('Elevation (feet)', fontsize=12, fontweight='bold')
        ax.set_title('Elevation Profile - Running Line Configuration', fontsize=14, fontweight='bold', pad=15)
        ax.legend(loc='upper right', fontsize=10, framealpha=0.9)
        ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        ax.set_facecolor('#f0f0f0')
        
        # Set proper margins
        fig.tight_layout(pad=2.0)
        
        # Embed in tkinter
        canvas = FigureCanvasTkAgg(fig, master=self.chart_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Connect mouse events
        canvas.mpl_connect('button_press_event', self._on_chart_click)
        canvas.mpl_connect('button_release_event', self._on_chart_release)
        canvas.mpl_connect('motion_notify_event', self._on_chart_motion)
        
        self.chart_canvas = canvas
        self.chart_figure = fig
        self.chart_ax = ax
    
    def _on_chart_click(self, event):
        """Handle chart click for adding break points, depth zones, or starting drag"""
        if event.inaxes != self.chart_ax:
            return
        
        clicked_distance = event.xdata
        if clicked_distance is None:
            return
        
        # Right-click: Remove break point
        if event.button == 3:  # Right mouse button
            self._remove_break_point_at(clicked_distance)
            return
        
        # Left-click: Check if clicking near existing break point to drag
        if event.button == 1 and not self.break_point_mode and not self.depth_edit_mode:
            for i, (bp_dist, bp_elev, grade, locked, note) in enumerate(self.break_points):
                if abs(bp_dist - clicked_distance) < 10:  # Within 10 ft
                    self.dragging_bp_index = i
                    self.drag_start_x = clicked_distance
                    return
        
        # Handle depth zone mode
        if self.depth_edit_mode:
            if self.depth_zone_start is None:
                # First click - set start
                self.depth_zone_start = clicked_distance
                messagebox.showinfo("Zone Start Set", f"Zone start set at {clicked_distance:.1f} ft.\nNow click to set the END of the zone.")
                logging.info(f"Depth zone start: {clicked_distance:.1f} ft")
            else:
                # Second click - set end and ask for depth
                zone_end = clicked_distance
                
                if abs(zone_end - self.depth_zone_start) < 5:
                    messagebox.showwarning("Too Short", "Zone must be at least 5 feet long.")
                    self.depth_zone_start = None
                    return
                
                # Ensure start < end
                if zone_end < self.depth_zone_start:
                    self.depth_zone_start, zone_end = zone_end, self.depth_zone_start
                
                # Create a custom dialog that stays on top
                depth_dialog = ctk.CTkInputDialog(
                    text=f"Enter depth (inches):",
                    title="Zone Depth"
                )
                depth_dialog.lift()
                depth_dialog.attributes('-topmost', True)
                depth_dialog.after(100, lambda: depth_dialog.focus_force())
                depth_str = depth_dialog.get_input()
                
                if depth_str:
                    try:
                        zone_depth = float(depth_str)
                        
                        # Add zone
                        self.depth_zones.append((self.depth_zone_start, zone_end, zone_depth))
                        self.depth_zones.sort(key=lambda x: x[0])
                        
                        # Recalculate with zones
                        self._recalculate_running_line()
                        self._draw_chart()
                        
                        logging.info(f"Added depth zone: {self.depth_zone_start:.1f}-{zone_end:.1f} ft @ {zone_depth}\"")
                        messagebox.showinfo("Success", f"Zone added: {self.depth_zone_start:.1f}-{zone_end:.1f} ft @ {zone_depth}\"")
                        
                    except ValueError:
                        messagebox.showerror("Invalid", "Please enter a valid depth.")
                
                self.depth_zone_start = None
            return
        
        # Handle break point mode
        if self.break_point_mode:
            # Find closest profile point
            closest_point = min(self.profile_data, key=lambda p: abs(p['distance'] - clicked_distance))
            
            # Check if too close to existing break points
            for bp in self.break_points:
                if abs(bp[0] - closest_point['distance']) < 5:
                    messagebox.showwarning("Too Close", "Break point too close to existing one.")
                    return
            
            # Ask user for depth below ground
            from tkinter import simpledialog
            
            # Use self as parent (this window is the CTkToplevel)
            self.lift()  # Bring this window to front
            self.focus_force()
            
            depth_inches = simpledialog.askfloat(
                "Break Point Depth",
                f"Enter depth below ground elevation (inches):\nGround elevation: {closest_point['elevation']:.2f} ft",
                parent=self,
                initialvalue=42,
                minvalue=0,
                maxvalue=120
            )
            
            if depth_inches is None:  # User canceled
                return
            
            # Add break point at specified depth below ground
            elev_at_bp = closest_point['elevation'] - (depth_inches / 12.0)
            self.break_points.append((closest_point['distance'], elev_at_bp, 0.0, False, "Manual"))
            self.break_points.sort(key=lambda x: x[0])
            
            self._recalculate_running_line()
            self._update_break_points_display()
            self._draw_chart()
            
            logging.info(f"Added break point at {closest_point['distance']:.1f} ft")
    
    def _on_chart_motion(self, event):
        """Handle mouse motion for dragging break points"""
        if self.dragging_bp_index is None or event.inaxes != self.chart_ax:
            return
        
        new_distance = event.xdata
        if new_distance is None:
            return
        
        # Get the break point being dragged
        bp_dist, bp_elev, grade, locked, note = self.break_points[self.dragging_bp_index]
        
        # Find closest profile point at new distance
        closest_point = min(self.profile_data, key=lambda p: abs(p['distance'] - new_distance))
        new_dist = closest_point['distance']
        
        # Check boundaries (don't drag past adjacent break points)
        min_dist = self.profile_data[0]['distance'] + 5
        max_dist = self.profile_data[-1]['distance'] - 5
        
        if self.dragging_bp_index > 0:
            min_dist = self.break_points[self.dragging_bp_index - 1][0] + 5
        if self.dragging_bp_index < len(self.break_points) - 1:
            max_dist = self.break_points[self.dragging_bp_index + 1][0] - 5
        
        if new_dist < min_dist or new_dist > max_dist:
            return
        
        # Update break point position - maintain the same depth below ground as before
        original_ground = None
        for p in self.profile_data:
            if abs(p['distance'] - bp_dist) < 1.0:
                original_ground = p['elevation']
                break
        
        if original_ground is not None:
            # Calculate the original depth
            original_depth_ft = original_ground - bp_elev
            # Apply the same depth at new location
            new_elev = closest_point['elevation'] - original_depth_ft
        else:
            # Fallback if can't find original ground
            new_elev = closest_point['elevation'] - (self.running_line_depth_inches / 12.0)
        
        self.break_points[self.dragging_bp_index] = (new_dist, new_elev, grade, locked, note)
        
        # Quick visual update only - draw on existing canvas without full redraw
        if hasattr(self, '_drag_line_artist'):
            self._drag_line_artist.remove()
        
        # Draw temporary line at new position
        self._drag_line_artist = self.chart_ax.axvline(
            x=new_dist, 
            color='purple', 
            linestyle='--', 
            linewidth=2, 
            alpha=0.7
        )
        self.chart_canvas.draw_idle()
    
    def _on_chart_release(self, event):
        """Handle mouse release to end dragging"""
        if self.dragging_bp_index is not None:
            # Clean up temporary drag line
            if hasattr(self, '_drag_line_artist'):
                self._drag_line_artist.remove()
                delattr(self, '_drag_line_artist')
            
            # Now do full recalculation and redraw
            self._recalculate_running_line()
            self._update_break_points_display()
            self._draw_chart()
            
            logging.info(f"Break point {self.dragging_bp_index + 1} moved to {self.break_points[self.dragging_bp_index][0]:.1f} ft")
            self.dragging_bp_index = None
            self.drag_start_x = None
    
    def _remove_break_point_at(self, distance):
        """Remove break point near the clicked distance"""
        if not self.break_points:
            return
        
        # Find closest break point
        closest_idx = None
        closest_dist = float('inf')
        
        for i, (bp_dist, bp_elev, grade, locked, note) in enumerate(self.break_points):
            dist_diff = abs(bp_dist - distance)
            if dist_diff < closest_dist and dist_diff < 10:  # Within 10 ft
                closest_idx = i
                closest_dist = dist_diff
        
        if closest_idx is not None:
            removed_bp = self.break_points.pop(closest_idx)
            self._recalculate_running_line()
            self._update_break_points_display()
            self._draw_chart()
            logging.info(f"Removed break point at {removed_bp[0]:.1f} ft")
            messagebox.showinfo("Break Point Removed", f"Removed break point at {removed_bp[0]:.1f} ft")
    
    def _toggle_curve_smoothing(self):
        """Toggle curve smoothing on/off"""
        self.curve_enabled = not self.curve_enabled
        
        if self.curve_enabled:
            self.curve_toggle_btn.configure(
                text="🔄 Path: Curved (Smooth)",
                fg_color="#4CAF50",
                hover_color="#45A049"
            )
            logging.info("Curve smoothing enabled - running line will be smoothed")
        else:
            self.curve_toggle_btn.configure(
                text="🔄 Path: Linear (Sharp Angles)",
                fg_color="#FF6B35",
                hover_color="#E85D2A"
            )
            logging.info("Curve smoothing disabled - running line will be linear")
        
        self._draw_chart()
    
    def _toggle_break_point_mode(self):
        """Toggle break point adding mode"""
        self.break_point_mode = not self.break_point_mode
        
        if self.break_point_mode:
            self.bp_mode_btn.configure(
                text="✅ Break Point Mode Active - Click Chart",
                fg_color="green",
                hover_color="darkgreen"
            )
        else:
            self.bp_mode_btn.configure(
                text="🖱️ Click Chart to Add Break Point",
                fg_color="gray40",
                hover_color="gray30"
            )
    
    def _clear_break_points(self):
        """Clear all break points"""
        if not self.break_points:
            return
        
        if messagebox.askyesno("Confirm", "Clear all break points?"):
            self.break_points.clear()
            self._recalculate_running_line()
            self._update_break_points_display()
            self._draw_chart()
            logging.info("Cleared all break points")
    
    def _auto_optimize(self):
        """Auto-optimize break points based on terrain"""
        if not self.profile_data or len(self.profile_data) < 10:
            messagebox.showinfo("Not Enough Data", "Need at least 10 profile points for auto-optimization.")
            return
        
        threshold = self.sensitivity / 100.0
        max_breakpoints = int(self.max_segments)
        
        # Calculate slopes
        slope_changes = []
        for i in range(1, len(self.profile_data) - 1):
            p1, p2, p3 = self.profile_data[i-1], self.profile_data[i], self.profile_data[i+1]
            
            slope1 = (p2['elevation'] - p1['elevation']) / (p2['distance'] - p1['distance']) if p2['distance'] != p1['distance'] else 0
            slope2 = (p3['elevation'] - p2['elevation']) / (p3['distance'] - p2['distance']) if p3['distance'] != p2['distance'] else 0
            
            slope_change = abs(slope2 - slope1)
            
            if slope_change > threshold:
                slope_changes.append((i, slope_change, p2))
        
        if not slope_changes:
            messagebox.showinfo("No Changes", f"No significant slope changes found at {self.sensitivity:.1f}% threshold.")
            return
        
        # Sort by magnitude and take top N
        slope_changes.sort(key=lambda x: x[1], reverse=True)
        selected = slope_changes[:max_breakpoints]
        
        # Clear existing and add new break points
        self.break_points.clear()
        for _, _, point in selected:
            elev_at_bp = point['elevation'] - (42 / 12.0)
            self.break_points.append((point['distance'], elev_at_bp, 0.0, False, "Auto"))
        
        self.break_points.sort(key=lambda x: x[0])
        
        self._recalculate_running_line()
        self._update_break_points_display()
        self._draw_chart()
        
        messagebox.showinfo("Success", f"Added {len(self.break_points)} optimized break points.")
        logging.info(f"Auto-optimized: {len(self.break_points)} break points")
    
    def _on_sensitivity_change(self, value):
        """Update sensitivity value"""
        self.sensitivity = value
        self.sensitivity_label.configure(text=f"{value:.1f}%")
    
    def _on_max_segments_change(self, value):
        """Update max segments value"""
        self.max_segments = int(value)
        self.max_segments_label.configure(text=f"{int(value)}")
    
    def _update_break_points_display(self):
        """Update the break points listbox"""
        self.bp_listbox.delete(0, tk.END)
        self.bp_listbox.insert(tk.END, "  Distance  | Elevation | Grade  | Note")
        self.bp_listbox.insert(tk.END, "─" * 45)
        
        for i, (dist, elev, grade, locked, note) in enumerate(self.break_points):
            line = f"{i+1}. {dist:7.1f} ft | {elev:7.1f} ft | {grade:5.2f}% | {note}"
            self.bp_listbox.insert(tk.END, line)
    
    def _on_bp_select(self, event):
        """Handle break point selection"""
        selection = self.bp_listbox.curselection()
        if selection:
            idx = selection[0]
            if idx >= 2:  # Skip header rows
                self.selected_bp_index = idx - 2
            else:
                self.selected_bp_index = None
    
    def _delete_selected_break_point(self):
        """Delete selected break point"""
        if self.selected_bp_index is not None and 0 <= self.selected_bp_index < len(self.break_points):
            self.break_points.pop(self.selected_bp_index)
            self.selected_bp_index = None
            self._recalculate_running_line()
            self._update_break_points_display()
            self._draw_chart()
            logging.info("Deleted break point")
    
    def _edit_selected_break_point(self):
        """Edit selected break point distance"""
        if self.selected_bp_index is None or not (0 <= self.selected_bp_index < len(self.break_points)):
            messagebox.showwarning("No Selection", "Please select a break point to move.")
            return
        
        old_dist, old_elev, grade, locked, note = self.break_points[self.selected_bp_index]
        
        # Simple dialog
        new_dist_str = tk.simpledialog.askstring(
            "Move Break Point",
            f"Current: {old_dist:.1f} ft\nEnter new distance:",
            initialvalue=f"{old_dist:.1f}"
        )
        
        if new_dist_str:
            try:
                new_dist = float(new_dist_str)
                max_dist = max(p['distance'] for p in self.profile_data)
                
                if new_dist < 5 or new_dist > max_dist - 5:
                    messagebox.showerror("Invalid", f"Distance must be between 5 and {max_dist-5:.1f} ft.")
                    return
                
                # Interpolate elevation
                new_elev = None
                for i, point in enumerate(self.profile_data):
                    if abs(point['distance'] - new_dist) < 0.5:
                        new_elev = point['elevation'] - (42 / 12.0)
                        break
                    elif i > 0 and self.profile_data[i-1]['distance'] <= new_dist <= point['distance']:
                        d1, e1 = self.profile_data[i-1]['distance'], self.profile_data[i-1]['elevation']
                        d2, e2 = point['distance'], point['elevation']
                        t = (new_dist - d1) / (d2 - d1) if d2 != d1 else 0
                        new_elev = (e1 + t * (e2 - e1)) - (42 / 12.0)
                        break
                
                if new_elev is not None:
                    self.break_points[self.selected_bp_index] = (new_dist, new_elev, grade, locked, note)
                    self.break_points.sort(key=lambda x: x[0])
                    self._recalculate_running_line()
                    self._update_break_points_display()
                    self._draw_chart()
                    logging.info(f"Moved break point to {new_dist:.1f} ft")
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter a valid number.")
    
    def _recalculate_running_line(self):
        """Recalculate running line with current break points and depth zones"""
        if not self.profile_data:
            return
        
        # Use parent's calculation method for break points
        self.parent._calculate_running_line_regression(self.profile_data, self.break_points)
        
        # Apply depth zones if any exist
        if self.depth_zones:
            for start_dist, end_dist, zone_depth in self.depth_zones:
                zone_depth_ft = zone_depth / 12.0
                for point in self.profile_data:
                    # Check if point is within this zone
                    if start_dist <= point['distance'] <= end_dist:
                        # Override running line elevation to match zone depth
                        point['running_line'] = point['elevation'] - zone_depth_ft
                        point['depth_inches'] = zone_depth
                        point['cover'] = zone_depth_ft
            logging.info(f"Applied {len(self.depth_zones)} depth zones to running line")
    
    def _update_depth(self):
        """Update running line depth and recalculate"""
        try:
            new_depth = float(self.depth_entry.get())
            if new_depth < 12 or new_depth > 120:
                messagebox.showerror("Invalid Depth", "Depth must be between 12 and 120 inches.")
                return
            
            self.running_line_depth_inches = new_depth
            self.parent.running_line_depth_inches = new_depth
            
            # Recalculate with new depth
            depth_ft = new_depth / 12.0
            for point in self.profile_data:
                point['running_line'] = point['elevation'] - depth_ft
                point['depth_inches'] = new_depth
            
            # If break points exist, recalculate with them
            if self.break_points:
                self._recalculate_running_line()
            
            self._draw_chart()
            logging.info(f"Updated running line depth to {new_depth:.0f} inches")
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid number.")
    
    def _set_preset_depth(self, depth):
        """Set preset depth value"""
        self.depth_entry.delete(0, tk.END)
        self.depth_entry.insert(0, str(depth))
        self._update_depth()
    
    def _toggle_depth_zone_mode(self):
        """Toggle depth zone creation mode"""
        self.depth_edit_mode = not self.depth_edit_mode
        
        if self.depth_edit_mode:
            self.depth_zone_btn.configure(
                text="✅ Zone Mode Active - Click Start",
                fg_color="green",
                hover_color="darkgreen"
            )
            messagebox.showinfo("Depth Zone Mode", "Click on the chart to set START of zone, then click END of zone.")
        else:
            self.depth_zone_btn.configure(
                text="🖱️ Click Chart to Add Zone",
                fg_color="#8E44AD",
                hover_color="#7D3C98"
            )
            self.depth_zone_start = None
    
    def _clear_depth_zones(self):
        """Clear all depth zones"""
        if not self.depth_zones:
            return
        
        if messagebox.askyesno("Clear Zones", "Clear all variable depth zones?"):
            self.depth_zones.clear()
            self.parent.depth_zones = []
            self._recalculate_running_line()
            self._draw_chart()
            logging.info("Cleared all depth zones")
    
    def _apply_changes(self):
        """Apply changes and close window"""
        # Update parent with all changes
        self.parent.break_points = self.break_points
        self.parent.depth_zones = self.depth_zones
        self.parent.running_line_depth_inches = self.running_line_depth_inches
        self.parent.curve_enabled = self.curve_enabled  # Save curve smoothing preference
        
        # Update parent's profile data with the configured running line
        self.parent.profile_data = self.profile_data
        
        # If break points exist, recalculate parent's running line using regression method
        if self.break_points:
            self.parent._calculate_running_line_regression(self.parent.profile_data, self.break_points)
        else:
            # No break points - use simple equal depth
            depth_ft = self.running_line_depth_inches / 12.0
            for point in self.parent.profile_data:
                point['running_line'] = point['elevation'] - depth_ft
                point['depth_inches'] = self.running_line_depth_inches
        
        # Apply depth zones if any exist (after break points)
        if self.depth_zones:
            for start_dist, end_dist, zone_depth in self.depth_zones:
                zone_depth_ft = zone_depth / 12.0
                for point in self.parent.profile_data:
                    # Check if point is within this zone
                    if start_dist <= point['distance'] <= end_dist:
                        # Override running line elevation to match zone depth
                        point['running_line'] = point['elevation'] - zone_depth_ft
                        point['depth_inches'] = zone_depth
                        point['cover'] = zone_depth_ft
            logging.info(f"Applied {len(self.depth_zones)} depth zones to parent running line")
        
        # Redraw parent chart to show the configured design
        if hasattr(self.parent, '_draw_profile_chart'):
            self.parent._draw_profile_chart()
        
        logging.info(f"Applied running line configuration: {len(self.break_points)} break points, {len(self.depth_zones)} depth zones, base depth={self.running_line_depth_inches}\"")
        self.destroy()


class ProfileAutomationTool(ctk.CTk):
    """Main application class for Profile Automation Tool"""
    
    def __init__(self):
        super().__init__()
        
        # Application metadata
        self.title("Profile Automation Tool")
        self.minsize(1200, 800)
        
        # Set appearance mode and color theme
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")
        
        # Application state - Waypoints system (multiple points)
        self.waypoints = []  # List of (lat, lon) tuples
        self.waypoint_selection = None  # [start_idx, end_idx] for selected range, None = all selected
        self.waypoint_markers = []  # List of marker objects
        self.path_line = None
        self.point_placement_active = False  # Toggle for adding points on map click
        self.profile_data = None  # List of dicts: {distance, elevation, x, y, slope}
        self.chart_canvas = None
        self.chart_figure = None
        self.chart_ax = None
        self.view_3d = False  # Toggle for 3D view
        self.map_type = "street"  # Track current map type: "street" or "satellite"
        self.imagery_quality = 2  # Imagery quality: 1=Low, 2=Medium, 3=High, 4=Very High
        self.cancel_generation = False  # Flag to cancel profile generation
        
        # API Service configuration
        self.api_source = "USGS"  # Options: "ArcGIS", "USGS", "Mock"
        self.service_url = "https://elevation.arcgis.com/arcgis/rest/services/Tools/ElevationSync/GPServer/Profile/execute"
        self.api_token = ""
        self.interval_ft = 10
        self.mock_mode = False  # Legacy support
        
        # Running Line configuration
        self.running_line_depth_inches = 42  # Default depth below ground
        self.enable_running_line = True  # Enable by default
        self.curve_enabled = False  # Curve smoothing for running line
        
        # Break Points for Multi-Segment Running Line Design
        # Each segment between break points has its own grade/slope
        # Format: [(distance_ft, elevation_ft, grade_percent, is_locked, note), ...]
        # grade_percent: None = auto-calculate, float = manual grade (e.g., 2.5 for 2.5%)
        self.break_points = []  # Start and end are automatic break points
        self.break_point_markers = []  # Chart markers for break points
        self.break_point_edit_mode = False  # Toggle for adding/editing break points
        
        # Depth Override Zones - list of (start_ft, end_ft, depth_inches, note)
        # Allows variable depth along route (e.g., deeper for river crossings)
        self.depth_zones = []  # [(start, end, depth, note), ...]
        self.depth_edit_mode = False  # Toggle for chart click editing
        self.pending_zone_start = None  # First click point for zone definition
        
        # Station labeling configuration
        self.station_interval_ft = 100  # Default station interval
        self.enable_stations = True  # Enable by default
        
        # Create menu bar
        self._create_menu_bar()
        
        # Setup UI
        self._setup_ui()
        self._setup_logging()
        
        # Initial log message
        logging.info("Profile Automation Tool initialized successfully")
        logging.info("Click two points on the map to create an elevation profile")
        
        # Maximize window AFTER everything is loaded (delayed)
        self.after(100, self._maximize_window)
        
    def _maximize_window(self):
        """Maximize the window after UI is fully loaded"""
        try:
            self.state('zoomed')  # Windows
            logging.info("Window maximized (Windows)")
        except:
            try:
                self.attributes('-zoomed', True)  # Linux
                logging.info("Window maximized (Linux)")
            except:
                # Mac - use full screen dimensions
                screen_width = self.winfo_screenwidth()
                screen_height = self.winfo_screenheight()
                self.geometry(f"{screen_width}x{screen_height}+0+0")
                logging.info(f"Window maximized (Mac): {screen_width}x{screen_height}")
    
    def _create_menu_bar(self):
        """Create traditional menu bar at top of window"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Save Project", command=self._save_project, accelerator="Ctrl+S")
        file_menu.add_command(label="Load Project", command=self._load_project, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit, accelerator="Alt+F4")
        
        # Configuration menu
        config_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Configuration", menu=config_menu)
        config_menu.add_command(label="Running Line Configuration", command=self._open_running_line_config_window)
        config_menu.add_command(label="Station Labeling", command=self._open_station_config_dialog)
        
        # Export menu
        export_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Export", menu=export_menu)
        export_menu.add_command(label="Export Package (Excel+PNG+KML)", command=self._export_package, accelerator="Ctrl+Shift+E")
        export_menu.add_separator()
        export_menu.add_command(label="Export Excel", command=lambda: self._export_data('excel'), accelerator="Ctrl+E")
        export_menu.add_command(label="Export CSV", command=lambda: self._export_data('csv'))
        export_menu.add_command(label="Export KML", command=self._export_kml)
        export_menu.add_separator()
        export_menu.add_command(label="Export Chart as PNG", command=lambda: self._export_chart('png'))
        export_menu.add_command(label="Export Chart as PDF", command=lambda: self._export_chart('pdf'))
        export_menu.add_command(label="Export Chart as SVG", command=lambda: self._export_chart('svg'))
        
        # Bind keyboard shortcuts
        self.bind('<Control-s>', lambda e: self._save_project())
        self.bind('<Control-o>', lambda e: self._load_project())
        self.bind('<Control-e>', lambda e: self._export_data('excel'))
        self.bind('<Control-Shift-E>', lambda e: self._export_package())
    
    def _setup_ui(self):
        """Setup the main user interface"""
        # Main container with padding
        main_container = ctk.CTkFrame(self, fg_color="transparent")
        main_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Header section
        self._create_header(main_container)
        
        # Main content area (2-column)
        content_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        content_frame.grid_columnconfigure(0, weight=1, minsize=300)  # Left column - fixed minimum width
        content_frame.grid_columnconfigure(1, weight=6)  # Right column (85%) - Controls + Chart
        content_frame.grid_rowconfigure(0, weight=1)
        
        # Left column: Map and Log
        self._create_left_column(content_frame)
        
        # Right column: Controls and Chart
        self._create_controls_frame(content_frame)
        
    def _create_header(self, parent):
        """Create header with logo, title and theme controls"""
        header_frame = ctk.CTkFrame(parent, fg_color="transparent")
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Left side: Logo + Title and credit
        left_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        left_frame.pack(side=tk.LEFT)
        
        # Logo
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "Logo.png")
            if not os.path.exists(logo_path):
                logo_path = "Logo.png"  # Try current directory
            
            if os.path.exists(logo_path):
                img = Image.open(logo_path)
                # Constrain size - max 80px for header
                max_side = 80
                w, h = img.size
                scale = min(1.0, max_side / max(w, h))
                if scale < 1.0:
                    img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
                img = img.convert("RGBA")
                self._header_logo = ImageTk.PhotoImage(img)
                logo_label = ctk.CTkLabel(left_frame, image=self._header_logo, text="")
                logo_label.pack(side=tk.LEFT, padx=(0, 15), pady=5)
        except Exception as e:
            logging.debug(f"Could not load logo: {e}")
        
        # Title container (next to logo)
        title_container = ctk.CTkFrame(left_frame, fg_color="transparent")
        title_container.pack(side=tk.LEFT)
        
        title_label = ctk.CTkLabel(
            title_container,
            text="Profile Automation Tool",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(anchor="w")
        
        credit_label = ctk.CTkLabel(
            title_container,
            text="Developed by Omid Zanganeh",
            font=ctk.CTkFont(size=12, slant="italic"),
            text_color="gray60"
        )
        credit_label.pack(anchor="w", pady=(5, 0))
        
        # Right side: Theme controls
        right_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        right_frame.pack(side=tk.RIGHT)
        
        theme_label = ctk.CTkLabel(
            right_frame,
            text="Appearance Mode:",
            font=ctk.CTkFont(size=12)
        )
        theme_label.pack(side=tk.LEFT, padx=(0, 10))
        
        self.theme_menu = ctk.CTkOptionMenu(
            right_frame,
            values=["Light", "Dark", "System"],
            command=self._change_appearance_mode,
            width=120
        )
        self.theme_menu.set("Dark")
        self.theme_menu.pack(side=tk.LEFT)
        
    def _create_left_column(self, parent):
        """Create left column with map and log"""
        left_column = ctk.CTkFrame(parent, fg_color="transparent")
        left_column.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        # Configure row weights: map gets 70%, log gets 30%
        left_column.rowconfigure(0, weight=70)  # Map section
        left_column.rowconfigure(1, weight=30)  # Log section
        left_column.columnconfigure(0, weight=1)
        
        # Map section (70% of left column)
        self._create_map_frame(left_column)
        
        # Log section (30% of left column)
        self._create_log_section(left_column)
    
    def _create_map_frame(self, parent):
        """Create interactive map frame"""
        map_frame = ctk.CTkFrame(parent)
        map_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        
        # Map title and waypoint status
        title_frame = ctk.CTkFrame(map_frame, fg_color="transparent")
        title_frame.pack(fill=tk.X, pady=(10, 5), padx=10)
        
        map_title = ctk.CTkLabel(
            title_frame,
            text="🗺️ Interactive Map",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        map_title.pack(side=tk.LEFT)
        
        # Waypoint status (moved here to prevent frame resizing)
        self.waypoint_label = ctk.CTkLabel(
            title_frame,
            text="Waypoints: 0 points",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="gray60"
        )
        self.waypoint_label.pack(side=tk.LEFT, padx=15)
        
        # Search bar
        search_frame = ctk.CTkFrame(map_frame, fg_color="transparent")
        search_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        search_label = ctk.CTkLabel(
            search_frame,
            text="Search:",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        search_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_entry = ctk.CTkEntry(
            search_frame,
            placeholder_text="Enter address or lat, lon"
        )
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.search_entry.bind('<Return>', lambda e: self._search_location())
        
        search_btn = ctk.CTkButton(
            search_frame,
            text="🔍 Search",
            width=80,
            command=self._search_location
        )
        search_btn.pack(side=tk.LEFT)
        
        # Map widget with optimized settings
        self.map_widget = tkintermapview.TkinterMapView(
            map_frame,
            corner_radius=0,
            use_database_only=False,  # Allow caching for faster loads
            database_path=os.path.join(os.path.dirname(__file__), "map_tiles.db")  # Persistent cache
        )
        self.map_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Use faster tile server (OpenStreetMap with CDN)
        self.map_widget.set_tile_server(
            "https://tile.openstreetmap.org/{z}/{x}/{y}.png",
            max_zoom=19
        )
        
        # Set initial position (Lincoln, Nebraska)
        self.map_widget.set_position(40.8108333, -96.7538888)
        self.map_widget.set_zoom(14)  # Increased from 12 to 14 (2 levels closer)
        
        # Add click handler
        self.map_widget.add_left_click_map_command(self._on_map_click)
        
        # Map controls
        controls_frame = ctk.CTkFrame(map_frame, fg_color="transparent")
        controls_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # Satellite/Street map toggle button
        self.map_toggle_btn = ctk.CTkButton(
            controls_frame,
            text="🛰️ Satellite",
            width=100,
            height=28,
            command=self._toggle_map_type,
            fg_color="#0078D4",
            hover_color="#005A9E"
        )
        self.map_toggle_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Toggle button for point placement mode
        self.placement_toggle_btn = ctk.CTkButton(
            controls_frame,
            text="📍 Add Points: OFF",
            width=140,
            height=28,
            command=self._toggle_point_placement,
            fg_color="gray40",
            hover_color="gray30"
        )
        self.placement_toggle_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Point controls
        clear_btn = ctk.CTkButton(
            controls_frame,
            text="Clear Points",
            command=self._clear_points,
            fg_color="gray40",
            hover_color="gray30"
        )
        clear_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        # Undo last point button
        self.undo_btn = ctk.CTkButton(
            controls_frame,
            text="↶ Undo Last",
            width=90,
            height=28,
            command=self._undo_last_waypoint,
            fg_color="gray40",
            hover_color="gray30"
        )
        self.undo_btn.pack(side=tk.LEFT, padx=5)
        
        # Select waypoints button
        self.select_waypoints_btn = ctk.CTkButton(
            controls_frame,
            text="✂️ Select Range",
            width=110,
            height=28,
            command=self._open_waypoint_selector,
            fg_color="#1f6aa5",
            hover_color="#1a5a8f"
        )
        self.select_waypoints_btn.pack(side=tk.LEFT, padx=5)
        
        # Import points button
        self.import_btn = ctk.CTkButton(
            controls_frame,
            text="📂 Import",
            width=90,
            height=28,
            command=self._import_points,
            fg_color="#2E7D32",
            hover_color="#1B5E20"
        )
        self.import_btn.pack(side=tk.LEFT, padx=5)
        
        # Interval setting (NEW - moved from API config)
        interval_label = ctk.CTkLabel(
            controls_frame,
            text="Interval:",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="gray60"
        )
        interval_label.pack(side=tk.LEFT, padx=(15, 5))
        
        self.interval_entry = ctk.CTkEntry(
            controls_frame,
            width=60,
            height=28
        )
        self.interval_entry.pack(side=tk.LEFT, padx=(0, 2))
        self.interval_entry.insert(0, str(self.interval_ft))
        
        interval_unit = ctk.CTkLabel(
            controls_frame,
            text="ft",
            font=ctk.CTkFont(size=11),
            text_color="gray60"
        )
        interval_unit.pack(side=tk.LEFT)
        
    def _create_controls_frame(self, parent):
        """Create controls and chart frame"""
        controls_frame = ctk.CTkFrame(parent)
        controls_frame.grid(row=0, column=1, sticky="nsew")
        
        # Make it scrollable
        scrollable = ctk.CTkScrollableFrame(controls_frame)
        scrollable.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # API Configuration Section
        self._create_api_config_section(scrollable)
        
        # Action Buttons
        self._create_action_buttons(scrollable)
        
        # Chart Display
        self._create_chart_section(scrollable)
    
    def _create_api_config_section(self, parent):
        """Create API configuration section with collapsible frames"""
        # Main container
        config_container = ctk.CTkFrame(parent, fg_color="transparent")
        config_container.pack(fill=tk.X, pady=(0, 15))
        
        # Project Name only (API settings are hardcoded)
        project_frame = ctk.CTkFrame(config_container, fg_color="transparent")
        project_frame.pack(fill=tk.X, pady=5)
        
        project_label = ctk.CTkLabel(
            project_frame,
            text="Project Name:",
            font=ctk.CTkFont(size=12, weight="bold"),
            width=100,
            anchor="w"
        )
        project_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.project_entry = ctk.CTkEntry(
            project_frame,
            placeholder_text="Enter project/sheet name for reports"
        )
        self.project_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
    def _create_action_buttons(self, parent):
        """Create action buttons with collapsible export section"""
        btn_frame = ctk.CTkFrame(parent, fg_color="transparent")
        btn_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Generate Profile button (always visible)
        generate_frame = ctk.CTkFrame(btn_frame, fg_color="transparent")
        generate_frame.pack(fill=tk.X, pady=5)
        
        self.generate_btn = ctk.CTkButton(
            generate_frame,
            text="🔄 Generate Elevation Profile",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=40,
            command=self._generate_profile,
            fg_color="#66BB6A",
            hover_color="#81C784"
        )
        self.generate_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        # Cancel button (hidden by default)
        self.cancel_generation_btn = ctk.CTkButton(
            generate_frame,
            text="⛔ Cancel",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=40,
            width=100,
            command=self._cancel_generation,
            fg_color="#E53935",
            hover_color="#C62828"
        )
        # Don't pack it yet - will show only during generation
        
        # View toggle buttons (always visible)
        view_frame = ctk.CTkFrame(btn_frame, fg_color="transparent")
        view_frame.pack(fill=tk.X, pady=5)
        
        self.view_2d_btn = ctk.CTkButton(
            view_frame,
            text="📊 2D Profile",
            command=lambda: self._toggle_view(False),
            fg_color="#42A5F5",
            hover_color="#2196F3"
        )
        self.view_2d_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 3))
        
        self.view_3d_btn = ctk.CTkButton(
            view_frame,
            text="🌄 3D View",
            command=lambda: self._toggle_view(True),
            fg_color="gray40",
            hover_color="gray30"
        )
        self.view_3d_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(3, 0))
        
    def _create_chart_section(self, parent):
        """Create chart display section"""
        chart_frame = ctk.CTkFrame(parent)
        chart_frame.pack(fill=tk.BOTH, expand=True)
        
        title = ctk.CTkLabel(
            chart_frame,
            text="📈 Elevation Profile",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        title.pack(pady=10, padx=10, anchor="w")
        
        # Chart container
        self.chart_container = ctk.CTkFrame(chart_frame, height=400)
        self.chart_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Initial placeholder
        placeholder = ctk.CTkLabel(
            self.chart_container,
            text="📊 Elevation profile will appear here\n\nClick two points on the map and generate profile",
            font=ctk.CTkFont(size=13),
            text_color="gray50"
        )
        placeholder.pack(expand=True)
        
        # Buttons below chart
        chart_buttons_frame = ctk.CTkFrame(chart_frame, fg_color="transparent")
        chart_buttons_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        configure_line_btn = ctk.CTkButton(
            chart_buttons_frame,
            text="⚙️ Configure Line",
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._open_running_line_config_window,
            fg_color="#1f6aa5",
            hover_color="#1e5a8f"
        )
        configure_line_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        download_imagery_btn = ctk.CTkButton(
            chart_buttons_frame,
            text="🛰️ Download Imagery",
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._download_imagery_dialog,
            fg_color="#9C27B0",
            hover_color="#7B1FA2"
        )
        download_imagery_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        
        export_btn = ctk.CTkButton(
            chart_buttons_frame,
            text="📤 Export",
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._open_export_dialog,
            fg_color="#2e7d32",
            hover_color="#1b5e20"
        )
        export_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
    def _create_log_section(self, parent):
        """Create log section in left column"""
        log_container = ctk.CTkFrame(parent)
        log_container.grid(row=1, column=0, sticky="nsew")
        
        # Header with title and clear button
        header_frame = ctk.CTkFrame(log_container, fg_color="transparent")
        header_frame.pack(fill=tk.X, pady=(8, 5), padx=10)
        
        title = ctk.CTkLabel(
            header_frame,
            text="📋 Progress & Log",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        title.pack(side=tk.LEFT)
        
        clear_log_btn = ctk.CTkButton(
            header_frame,
            text="Clear Log",
            width=90,
            height=28,
            command=self._clear_log,
            fg_color="gray40",
            hover_color="gray30"
        )
        clear_log_btn.pack(side=tk.RIGHT)
        
        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(log_container, height=18)
        self.progress_bar.pack(fill=tk.X, padx=10, pady=(0, 8))
        self.progress_bar.set(0)
        
        # Log text area
        log_frame = ctk.CTkFrame(log_container, fg_color="transparent")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        self.log_text = tk.Text(
            log_frame,
            wrap=tk.WORD,
            font=("Consolas", 9),
            bg="#303030" if ctk.get_appearance_mode() == "Dark" else "#F0F0F0",
            fg="#F0F0F0" if ctk.get_appearance_mode() == "Dark" else "#303030",
            relief=tk.FLAT,
            state='disabled'
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ctk.CTkScrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
    def _setup_logging(self):
        """Setup logging to redirect to text widget"""
        text_handler = TextHandler(self.log_text)
        text_handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s', '%H:%M:%S'))
        logging.getLogger().addHandler(text_handler)
        
    def _change_appearance_mode(self, mode):
        """Change application appearance mode"""
        ctk.set_appearance_mode(mode)
        # Update log text colors
        if mode == "Dark" or (mode == "System" and ctk.get_appearance_mode() == "Dark"):
            self.log_text.configure(bg="#303030", fg="#F0F0F0")
        else:
            self.log_text.configure(bg="#F0F0F0", fg="#303030")
        # Redraw chart if exists
        if self.chart_figure:
            self._draw_profile_chart()
        logging.info(f"Appearance mode changed to: {mode}")
        
    def _toggle_depth_edit_mode(self):
        """Toggle depth edit mode for chart clicking"""
        self.depth_edit_mode = not self.depth_edit_mode
        self.pending_zone_start = None  # Reset pending start point
        
        if self.depth_edit_mode:
            self.depth_edit_btn.configure(
                text="✅ EDIT MODE ON - Click Start Point",
                fg_color="#27AE60",
                hover_color="#229954"
            )
            logging.info("Depth edit mode ENABLED - Click on the chart to define depth zones")
            messagebox.showinfo(
                "Depth Zone Edit Mode",
                "Define depth zones by clicking TWO points on the chart:\n\n"
                "1️⃣ Click the START of the zone (e.g., where river begins)\n"
                "2️⃣ Click the END of the zone (e.g., where river ends)\n"
                "3️⃣ Enter the depth for that zone\n\n"
                "• Useful for river crossings, road crossings, etc.\n"
                "• Click the button again to exit edit mode"
            )
        else:
            self.depth_edit_btn.configure(
                text="🖱️ Click Chart to Add Depth Zone",
                fg_color="#8E44AD",
                hover_color="#7D3C98"
            )
            logging.info("Depth edit mode DISABLED")
    
    def _clear_depth_overrides(self):
        """Clear all depth zones"""
        if self.depth_zones:
            if messagebox.askyesno("Clear Zones", "Are you sure you want to clear all depth zones?"):
                self.depth_zones = []
                self.pending_zone_start = None
                self._update_overrides_display()
                if self.profile_data:
                    self._recalculate_running_line()  # Recalculate with default depth
                    self._draw_profile_chart()
                logging.info("All depth zones cleared")
    
    def _update_overrides_display(self):
        """Log depth zones to main log window"""
        if self.depth_zones:
            logging.info(f"Depth zones configured: {len(self.depth_zones)} zone(s)")
            for start, end, depth, note in sorted(self.depth_zones, key=lambda x: x[0]):
                note_str = f" - {note}" if note else ""
                logging.info(f"  📍 {start:.0f} - {end:.0f} ft → {depth}\"{note_str}")
    
    def _add_depth_zone(self, start_ft, end_ft):
        """Add a depth zone between two distances"""
        # Ensure start < end
        if start_ft > end_ft:
            start_ft, end_ft = end_ft, start_ft
        
        # Create dialog to get depth value
        dialog = ctk.CTkToplevel(self)
        dialog.title("Add Depth Zone")
        dialog.geometry("380x280")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (190)
        y = (dialog.winfo_screenheight() // 2) - (140)
        dialog.geometry(f"380x280+{x}+{y}")
        
        result = {'depth': None, 'note': None}
        
        # Content
        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(
            main_frame,
            text=f"📍 Depth Zone: {start_ft:.0f} ft - {end_ft:.0f} ft",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        title.pack(pady=(0, 5))
        
        zone_length = ctk.CTkLabel(
            main_frame,
            text=f"Zone Length: {end_ft - start_ft:.0f} feet",
            font=ctk.CTkFont(size=11),
            text_color="gray60"
        )
        zone_length.pack(pady=(0, 15))
        
        # Depth input
        depth_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        depth_frame.pack(fill=tk.X, pady=5)
        
        depth_label = ctk.CTkLabel(depth_frame, text="Depth (inches):", width=110, anchor="w")
        depth_label.pack(side=tk.LEFT)
        
        depth_entry = ctk.CTkEntry(depth_frame, width=100)
        depth_entry.pack(side=tk.LEFT, padx=(5, 0))
        depth_entry.insert(0, "60")  # Default to deeper for crossings
        depth_entry.focus()
        
        # Quick presets
        preset_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        preset_frame.pack(fill=tk.X, pady=5)
        
        preset_label = ctk.CTkLabel(preset_frame, text="Presets:", width=110, anchor="w")
        preset_label.pack(side=tk.LEFT)
        
        for d in [48, 60, 72, 96, 120]:
            btn = ctk.CTkButton(
                preset_frame, text=f'{d}"', width=50, height=28,
                command=lambda val=d: (depth_entry.delete(0, tk.END), depth_entry.insert(0, str(val)))
            )
            btn.pack(side=tk.LEFT, padx=2)
        
        # Note input
        note_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        note_frame.pack(fill=tk.X, pady=5)
        
        note_label = ctk.CTkLabel(note_frame, text="Note (optional):", width=110, anchor="w")
        note_label.pack(side=tk.LEFT)
        
        note_entry = ctk.CTkEntry(note_frame, width=200, placeholder_text="e.g., River crossing")
        note_entry.pack(side=tk.LEFT, padx=(5, 0))
        
        # Buttons
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        
        def on_add():
            try:
                result['depth'] = int(depth_entry.get())
                result['note'] = note_entry.get().strip()
                dialog.destroy()
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter a valid depth in inches.")
        
        def on_cancel():
            dialog.destroy()
        
        add_btn = ctk.CTkButton(
            btn_frame, text="✅ Add Zone", command=on_add,
            fg_color="#27AE60", hover_color="#229954"
        )
        add_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        cancel_btn = ctk.CTkButton(
            btn_frame, text="Cancel", command=on_cancel,
            fg_color="#95A5A6", hover_color="#7F8C8D"
        )
        cancel_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        dialog.wait_window()
        
        if result['depth'] is not None:
            # Add zone
            self.depth_zones.append((start_ft, end_ft, result['depth'], result['note']))
            self._update_overrides_display()
            if self.profile_data:
                self._recalculate_running_line()  # Recalculate with new zone
                self._draw_profile_chart()
            logging.info(f"Added depth zone: {start_ft:.0f} - {end_ft:.0f} ft → {result['depth']}\" {result['note'] or ''}")
    
    def _get_depth_at_distance(self, distance_ft):
        """Get the depth (in inches) at a given distance, considering depth zones"""
        if not self.depth_zones:
            return self.running_line_depth_inches
        
        # Check if distance falls within any zone
        for start, end, depth, _ in self.depth_zones:
            if start <= distance_ft <= end:
                return depth
        
        # Not in any zone - use default
        return self.running_line_depth_inches
    
    # ===== BREAK POINT METHODS =====
    
    def _toggle_break_point_mode(self):
        """Toggle break point edit mode for chart clicking"""
        self.break_point_edit_mode = not self.break_point_edit_mode
        
        if self.break_point_edit_mode:
            self.bp_mode_btn.configure(
                text="✅ BREAK POINT MODE ON",
                fg_color="#27AE60",
                hover_color="#229954"
            )
            logging.info("Break point mode ENABLED - Click on the chart to add break points")
            messagebox.showinfo(
                "Break Point Mode",
                "Click on the profile chart to add break points.\nEach break point creates a new segment with its own grade.\n\nClick 'Disable Break Point Mode' when done."
            )
        else:
            self.bp_mode_btn.configure(
                text="📍 Enable Break Point Mode",
                fg_color="gray40",
                hover_color="gray30"
            )
            logging.info("Break point mode DISABLED")
    
    def _open_running_line_config_window(self):
        """Open the dedicated running line configuration window"""
        if not self.profile_data:
            messagebox.showwarning("No Profile", "Please generate a profile first before configuring the running line.")
            return
        
        def on_apply(updated_break_points):
            """Callback when user applies changes from the config window"""
            self.break_points = updated_break_points
            self._update_break_points_display()
            self._recalculate_running_line()
            self._draw_profile_chart()
            logging.info("Running line configuration applied")
        
        # Open the configuration window
        config_window = RunningLineConfigWindow(
            parent=self,
            profile_data=self.profile_data,
            break_points=self.break_points,
            callback=on_apply
        )
    
    def _open_station_config_dialog(self):
        """Open dialog for station labeling configuration"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Station Labeling Configuration")
        dialog.geometry("400x250")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 200
        y = (dialog.winfo_screenheight() // 2) - 125
        dialog.geometry(f"400x250+{x}+{y}")
        
        # Content frame
        content = ctk.CTkFrame(dialog)
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(content, text="📏 Station Labeling", font=ctk.CTkFont(size=16, weight="bold"))
        title.pack(pady=(0, 15))
        
        # Enable checkbox
        enable_var = tk.BooleanVar(value=self.enable_stations)
        enable_check = ctk.CTkCheckBox(content, text="Show Station Markers on Chart", variable=enable_var, font=ctk.CTkFont(size=12))
        enable_check.pack(pady=10)
        
        # Interval input
        interval_frame = ctk.CTkFrame(content, fg_color="transparent")
        interval_frame.pack(pady=10)
        
        interval_label = ctk.CTkLabel(interval_frame, text="Station Interval (feet):", font=ctk.CTkFont(size=12))
        interval_label.pack(side=tk.LEFT, padx=(0, 10))
        
        interval_entry = ctk.CTkEntry(interval_frame, width=100)
        interval_entry.pack(side=tk.LEFT)
        interval_entry.insert(0, str(self.station_interval_ft))
        
        # Buttons
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(pady=20)
        
        def apply_settings():
            try:
                self.enable_stations = enable_var.get()
                self.station_interval_ft = float(interval_entry.get())
                if hasattr(self, 'enable_station_checkbox'):
                    if self.enable_stations:
                        self.enable_station_checkbox.select()
                    else:
                        self.enable_station_checkbox.deselect()
                if hasattr(self, 'station_entry'):
                    self.station_entry.delete(0, tk.END)
                    self.station_entry.insert(0, str(self.station_interval_ft))
                if self.profile_data:
                    self._draw_profile_chart()
                dialog.destroy()
                messagebox.showinfo("Success", "Station labeling settings updated")
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter a valid number for station interval")
        
        apply_btn = ctk.CTkButton(btn_frame, text="Apply", command=apply_settings, width=100)
        apply_btn.pack(side=tk.LEFT, padx=5)
        
        cancel_btn = ctk.CTkButton(btn_frame, text="Cancel", command=dialog.destroy, width=100, fg_color="gray40")
        cancel_btn.pack(side=tk.LEFT, padx=5)
    
    def _open_export_dialog(self):
        """Open dialog for choosing export type"""
        if not self.profile_data:
            messagebox.showwarning("No Data", "Please generate an elevation profile first before exporting.")
            return
        
        dialog = ctk.CTkToplevel(self)
        dialog.title("Export Options")
        dialog.geometry("450x650")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 225
        y = (dialog.winfo_screenheight() // 2) - 325
        dialog.geometry(f"450x650+{x}+{y}")
        
        # Content frame
        content = ctk.CTkFrame(dialog)
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(content, text="📤 Export Profile", font=ctk.CTkFont(size=18, weight="bold"))
        title.pack(pady=(0, 20))
        
        # Export options
        export_options = [
            ("📦 Package (Excel + PNG + KML)", self._export_package, "#1976D2", "Export all formats to a folder"),
            ("📊 Excel Spreadsheet", lambda: self._export_data('excel'), "#217346", "Detailed data table"),
            ("📄 CSV File", lambda: self._export_data('csv'), "#555555", "Comma-separated values"),
            ("🗺️ KML (Google Earth)", self._export_kml, "#4285F4", "View in Google Earth"),
            ("🖼️ PNG Image", lambda: self._export_chart('png'), "#E91E63", "Raster image format"),
            ("📑 PDF Document", lambda: self._export_chart('pdf'), "#D32F2F", "Print-ready format"),
            ("🎨 SVG Vector", lambda: self._export_chart('svg'), "#9C27B0", "Scalable vector graphics")
        ]
        
        for label, command, color, description in export_options:
            btn_container = ctk.CTkFrame(content, fg_color="transparent")
            btn_container.pack(fill=tk.X, pady=5)
            
            btn = ctk.CTkButton(
                btn_container,
                text=label,
                height=40,
                font=ctk.CTkFont(size=13, weight="bold"),
                command=lambda cmd=command, dlg=dialog: (cmd(), dlg.destroy()),
                fg_color=color,
                hover_color=self._darken_color(color)
            )
            btn.pack(fill=tk.X)
            
            desc_label = ctk.CTkLabel(
                btn_container,
                text=description,
                font=ctk.CTkFont(size=10),
                text_color="gray60"
            )
            desc_label.pack(anchor="w", padx=5, pady=(2, 0))
        
        # Close button
        close_btn = ctk.CTkButton(
            content,
            text="Cancel",
            command=dialog.destroy,
            width=100,
            fg_color="gray40",
            hover_color="gray30"
        )
        close_btn.pack(pady=(15, 0))
    
    def _darken_color(self, hex_color):
        """Darken a hex color by 20% for hover effect"""
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r, g, b = int(r * 0.8), int(g * 0.8), int(b * 0.8)
        return f'#{r:02x}{g:02x}{b:02x}'
    
    def _clear_break_points(self):
        """Clear all break points"""
        if self.break_points:
            if messagebox.askyesno("Clear Break Points", "Are you sure you want to clear all break points and reset to single-segment design?"):
                self.break_points = []
                self._update_break_points_display()
                if self.profile_data:
                    self._recalculate_running_line()
                    self._draw_profile_chart()
                logging.info("All break points cleared - using single-segment design")
    
    def _add_break_point(self, distance_ft):
        """Add a break point at the specified distance"""
        if not self.profile_data:
            return
        
        # Find elevation at this distance (interpolate if needed)
        elevation_ft = None
        for i, point in enumerate(self.profile_data):
            if abs(point['distance'] - distance_ft) < 0.5:  # Close enough
                elevation_ft = point['elevation']
                break
            elif i > 0 and self.profile_data[i-1]['distance'] <= distance_ft <= point['distance']:
                # Interpolate
                d1, e1 = self.profile_data[i-1]['distance'], self.profile_data[i-1]['elevation']
                d2, e2 = point['distance'], point['elevation']
                t = (distance_ft - d1) / (d2 - d1) if d2 != d1 else 0
                elevation_ft = e1 + t * (e2 - e1)
                break
        
        if elevation_ft is None:
            logging.warning(f"Could not find elevation at distance {distance_ft:.1f} ft")
            return
        
        # Check if break point already exists nearby
        for bp_dist, _, _, _, _ in self.break_points:
            if abs(bp_dist - distance_ft) < 5.0:  # Within 5 feet
                logging.warning(f"Break point already exists near {distance_ft:.1f} ft")
                return
        
        # Open dialog to get grade for this segment
        grade = self._get_break_point_grade_dialog(distance_ft, elevation_ft)
        
        if grade is not None:  # None means canceled
            # Add break point: (distance, elevation, grade, is_locked, note)
            self.break_points.append((distance_ft, elevation_ft, grade, False, ""))
            self.break_points.sort(key=lambda x: x[0])  # Sort by distance
            self._update_break_points_display()
            
            if self.profile_data:
                self._recalculate_running_line()
                self._draw_profile_chart()
            
            grade_str = f"{grade:.2f}%" if grade is not None else "Auto"
            logging.info(f"Added break point at {distance_ft:.1f} ft (elevation: {elevation_ft:.2f} ft, grade: {grade_str})")
    
    def _update_break_points_display(self):
        """Update the break points listbox (if it exists)"""
        # Check if listbox exists (it's in the dedicated window now, not main window)
        if not hasattr(self, 'bp_listbox'):
            return
        
        self.bp_listbox.delete(0, tk.END)
        
        if not self.break_points:
            self.bp_listbox.insert(tk.END, "  No break points - single segment design")
            return
        
        self.bp_listbox.insert(tk.END, f"  {'Distance':>8}  {'Elevation':>10}  {'Grade':>8}  Note")
        self.bp_listbox.insert(tk.END, "  " + "-" * 55)
        
        for dist, elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
            lock_icon = "🔒" if locked else "  "
            grade_str = f"{grade:>7.2f}%" if grade is not None else "   Auto"
            note_str = f" - {note}" if note else ""
            self.bp_listbox.insert(tk.END, f"{lock_icon} {dist:>8.1f} ft  {elev:>8.2f} ft  {grade_str}{note_str}")
    
    def _auto_optimize_segments(self):
        """Automatically find optimal break points to minimize cut/fill"""
        if not self.profile_data or len(self.profile_data) < 10:
            messagebox.showwarning("Not Enough Data", "Need at least 10 profile points to auto-optimize.")
            return
        
        # Get settings from sliders (if they exist) or use defaults
        if hasattr(self, 'sensitivity_slider'):
            threshold = self.sensitivity_slider.get() / 100.0
            max_breakpoints = int(self.max_segments_slider.get())
        else:
            threshold = 0.05  # Default 5%
            max_breakpoints = 5  # Default 5 break points
        
        logging.info(f"Auto-optimizing with sensitivity={threshold*100:.1f}%, max_breakpoints={max_breakpoints}")
        
        # Simple algorithm: Find points where slope changes significantly
        significant_changes = []
        
        for i in range(1, len(self.profile_data) - 1):
            prev_slope = (self.profile_data[i]['elevation'] - self.profile_data[i-1]['elevation']) / \
                        (self.profile_data[i]['distance'] - self.profile_data[i-1]['distance'])
            next_slope = (self.profile_data[i+1]['elevation'] - self.profile_data[i]['elevation']) / \
                        (self.profile_data[i+1]['distance'] - self.profile_data[i]['distance'])
            
            slope_change = abs(next_slope - prev_slope)
            if slope_change > threshold:
                significant_changes.append((self.profile_data[i]['distance'], 
                                          self.profile_data[i]['elevation'], 
                                          slope_change))
        
        if not significant_changes:
            messagebox.showinfo("No Significant Changes", 
                              f"No slope changes > {threshold*100:.1f}% found.\nTry lowering sensitivity or the profile is relatively uniform.")
            return
        
        # Sort by slope change magnitude and take top N
        significant_changes.sort(key=lambda x: x[2], reverse=True)
        top_changes = significant_changes[:min(max_breakpoints, len(significant_changes))]
        
        # Add as break points (with auto-calculated grades)
        self.break_points = []
        for dist, elev, change in top_changes:
            self.break_points.append((dist, elev, None, False, f"Auto: {change:.3f} slope change"))
        
        self.break_points.sort(key=lambda x: x[0])
        self._update_break_points_display()
        
        if self.profile_data:
            self._recalculate_running_line()
            self._draw_profile_chart()
        
        logging.info(f"Auto-optimized: Added {len(self.break_points)} break points")
    
    def _on_sensitivity_change(self, value):
        """Update sensitivity label when slider changes"""
        if hasattr(self, 'sensitivity_value_label'):
            self.sensitivity_value_label.configure(text=f"{value:.1f}%")
    
    def _on_max_segments_change(self, value):
        """Update max segments label when slider changes"""
        if hasattr(self, 'max_segments_value_label'):
            self.max_segments_value_label.configure(text=f"{int(value)}")
    
    def _delete_selected_break_point(self):
        """Delete the selected break point from the list"""
        if not hasattr(self, 'bp_listbox'):
            return
        
        selection = self.bp_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a break point to delete.")
            return
        
        if not self.break_points:
            return
        
        # Get selected index (skip header rows)
        selected_idx = selection[0]
        if selected_idx < 2:  # Header rows
            messagebox.showwarning("Invalid Selection", "Please select a break point, not the header.")
            return
        
        # Calculate actual break point index (subtract header rows)
        bp_idx = selected_idx - 2
        
        if 0 <= bp_idx < len(self.break_points):
            deleted_bp = self.break_points[bp_idx]
            self.break_points.pop(bp_idx)
            self._update_break_points_display()
            
            if self.profile_data:
                self._recalculate_running_line()
                self._draw_profile_chart()
            
            logging.info(f"Deleted break point at {deleted_bp[0]:.1f} ft")
    
    def _edit_selected_break_point(self):
        """Edit the distance of the selected break point"""
        if not hasattr(self, 'bp_listbox'):
            return
        
        selection = self.bp_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a break point to edit.")
            return
        
        if not self.break_points or not self.profile_data:
            return
        
        # Get selected index (skip header rows)
        selected_idx = selection[0]
        if selected_idx < 2:  # Header rows
            messagebox.showwarning("Invalid Selection", "Please select a break point, not the header.")
            return
        
        # Calculate actual break point index
        bp_idx = selected_idx - 2
        
        if 0 <= bp_idx < len(self.break_points):
            old_dist, old_elev, grade, locked, note = self.break_points[bp_idx]
            
            # Create dialog to get new distance
            dialog = ctk.CTkToplevel(self)
            dialog.title("Edit Break Point Distance")
            dialog.geometry("380x200")
            dialog.transient(self)
            dialog.grab_set()
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - 190
            y = (dialog.winfo_screenheight() // 2) - 100
            dialog.geometry(f"380x200+{x}+{y}")
            
            result = {'distance': None}
            
            # Content
            main_frame = ctk.CTkFrame(dialog)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Title
            title = ctk.CTkLabel(
                main_frame,
                text=f"Move Break Point",
                font=ctk.CTkFont(size=14, weight="bold")
            )
            title.pack(pady=(0, 5))
            
            info = ctk.CTkLabel(
                main_frame,
                text=f"Current position: {old_dist:.1f} ft\\nEnter new distance:",
                font=ctk.CTkFont(size=11)
            )
            info.pack(pady=(0, 10))
            
            # Distance entry
            dist_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            dist_frame.pack(fill=tk.X, pady=5)
            
            dist_entry = ctk.CTkEntry(dist_frame, width=100, placeholder_text=f"{old_dist:.1f}")
            dist_entry.insert(0, f"{old_dist:.1f}")
            dist_entry.pack(side=tk.LEFT, padx=(0, 5))
            
            ft_label = ctk.CTkLabel(dist_frame, text="feet")
            ft_label.pack(side=tk.LEFT)
            
            # Buttons
            btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            btn_frame.pack(fill=tk.X, pady=(15, 0))
            
            def on_save():
                try:
                    new_dist = float(dist_entry.get())
                    max_dist = max(p['distance'] for p in self.profile_data)
                    
                    if new_dist < 5 or new_dist > max_dist - 5:
                        messagebox.showerror("Invalid Distance", 
                                           f"Distance must be between 5 and {max_dist-5:.1f} feet.")
                        return
                    
                    result['distance'] = new_dist
                    dialog.destroy()
                except ValueError:
                    messagebox.showerror("Invalid Input", "Please enter a valid distance.")
            
            def on_cancel():
                dialog.destroy()
            
            save_btn = ctk.CTkButton(
                btn_frame, text="✅ Move", command=on_save,
                fg_color="#27AE60", hover_color="#229954"
            )
            save_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
            
            cancel_btn = ctk.CTkButton(
                btn_frame, text="Cancel", command=on_cancel,
                fg_color="#95A5A6", hover_color="#7F8C8D"
            )
            cancel_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
            
            dialog.wait_window()
            
            if result['distance'] is not None:
                new_dist = result['distance']
                
                # Find new elevation at this distance
                new_elev = None
                for i, point in enumerate(self.profile_data):
                    if abs(point['distance'] - new_dist) < 0.5:
                        new_elev = point['elevation']
                        break
                    elif i > 0 and self.profile_data[i-1]['distance'] <= new_dist <= point['distance']:
                        # Interpolate
                        d1, e1 = self.profile_data[i-1]['distance'], self.profile_data[i-1]['elevation']
                        d2, e2 = point['distance'], point['elevation']
                        t = (new_dist - d1) / (d2 - d1) if d2 != d1 else 0
                        new_elev = e1 + t * (e2 - e1)
                        break
                
                if new_elev is not None:
                    # Update break point
                    self.break_points[bp_idx] = (new_dist, new_elev, grade, locked, note)
                    self.break_points.sort(key=lambda x: x[0])
                    self._update_break_points_display()
                    
                    if self.profile_data:
                        self._recalculate_running_line()
                        self._draw_profile_chart()
                    
                    logging.info(f"Moved break point from {old_dist:.1f} to {new_dist:.1f} ft")
    
    def _get_break_point_grade_dialog(self, distance_ft, elevation_ft):
        """Dialog to get manual grade for a segment or use auto-calculate"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Break Point Grade")
        dialog.geometry("420x320")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 210
        y = (dialog.winfo_screenheight() // 2) - 160
        dialog.geometry(f"420x320+{x}+{y}")
        
        result = {'grade': None, 'canceled': False}
        
        # Content
        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(
            main_frame,
            text=f"📍 Break Point at {distance_ft:.0f} ft",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        title.pack(pady=(0, 5))
        
        info = ctk.CTkLabel(
            main_frame,
            text=f"Ground Elevation: {elevation_ft:.2f} ft\\n\\nDefine grade for the NEXT segment:",
            font=ctk.CTkFont(size=11)
        )
        info.pack(pady=(0, 15))
        
        # Grade options
        grade_label = ctk.CTkLabel(
            main_frame,
            text="Segment Grade:",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        grade_label.pack(anchor="w", pady=(0, 5))
        
        # Auto or Manual radio
        grade_mode = tk.StringVar(value="auto")
        
        auto_radio = ctk.CTkRadioButton(
            main_frame,
            text="Auto-calculate from ground points (linear regression)",
            variable=grade_mode,
            value="auto",
            font=ctk.CTkFont(size=11)
        )
        auto_radio.pack(anchor="w", pady=2)
        
        manual_radio = ctk.CTkRadioButton(
            main_frame,
            text="Manual grade entry:",
            variable=grade_mode,
            value="manual",
            font=ctk.CTkFont(size=11)
        )
        manual_radio.pack(anchor="w", pady=2)
        
        # Manual grade entry
        grade_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        grade_frame.pack(fill=tk.X, pady=5, padx=20)
        
        grade_entry = ctk.CTkEntry(grade_frame, width=100, placeholder_text="e.g., 2.5")
        grade_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        percent_label = ctk.CTkLabel(grade_frame, text="% (positive = uphill, negative = downhill)")
        percent_label.pack(side=tk.LEFT)
        
        # Preset buttons
        preset_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        preset_frame.pack(fill=tk.X, pady=10, padx=20)
        
        preset_label = ctk.CTkLabel(preset_frame, text="Presets:", font=ctk.CTkFont(size=10))
        preset_label.pack(side=tk.LEFT, padx=(0, 5))
        
        for g in [-5, -2, -1, 0, 1, 2, 5]:
            btn = ctk.CTkButton(
                preset_frame, text=f"{g:+.0f}%", width=50,
                command=lambda val=g: (grade_entry.delete(0, tk.END), grade_entry.insert(0, str(val)), grade_mode.set("manual"))
            )
            btn.pack(side=tk.LEFT, padx=2)
        
        # Buttons
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        
        def on_add():
            if grade_mode.get() == "auto":
                result['grade'] = None  # None means auto-calculate
                dialog.destroy()
            else:
                try:
                    result['grade'] = float(grade_entry.get())
                    dialog.destroy()
                except ValueError:
                    messagebox.showerror("Invalid Input", "Please enter a valid grade percentage (e.g., 2.5 for 2.5%)")
        
        def on_cancel():
            result['canceled'] = True
            dialog.destroy()
        
        add_btn = ctk.CTkButton(
            btn_frame, text="✅ Add Break Point", command=on_add,
            fg_color="#27AE60", hover_color="#229954"
        )
        add_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        cancel_btn = ctk.CTkButton(
            btn_frame, text="Cancel", command=on_cancel,
            fg_color="#95A5A6", hover_color="#7F8C8D"
        )
        cancel_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        dialog.wait_window()
        
        return None if result['canceled'] else result.get('grade')
    
    # ===== RUNNING LINE CALCULATION =====
    
    def _calculate_running_line_regression(self, profile_data, break_points=None):
        """Calculate multi-segment running line with break points"""
        # Use provided break_points or fall back to instance variable
        if break_points is None:
            break_points = self.break_points
        
        min_clearance_ft = self.running_line_depth_inches / 12.0
        
        if not break_points:
            # Single segment design - use linear regression through all points
            logging.info(f"Calculating SINGLE-SEGMENT running line for {len(profile_data)} points")
            logging.info(f"Target perpendicular distance: {min_clearance_ft:.2f} ft ({self.running_line_depth_inches:.0f} inches)")
            
            # Calculate linear regression through all ground points
            distances = np.array([p['distance'] for p in profile_data])
            elevations = np.array([p['elevation'] for p in profile_data])
            
            n = len(distances)
            sum_x = np.sum(distances)
            sum_y = np.sum(elevations)
            sum_xy = np.sum(distances * elevations)
            sum_x2 = np.sum(distances ** 2)
            
            # Best fit line: y = mx + b
            slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2)
            ground_intercept = (sum_y - slope * sum_x) / n
            
            # Shift down by perpendicular distance
            perpendicular_factor = np.sqrt(1 + slope**2)
            perpendicular_shift = min_clearance_ft * perpendicular_factor
            running_intercept = ground_intercept - perpendicular_shift
            
            logging.info(f"Ground regression: slope={slope:.6f} ({slope*100:.2f}%), intercept={ground_intercept:.2f}")
            logging.info(f"Running line: slope={slope:.6f}, intercept={running_intercept:.2f}")
            logging.info(f"Perpendicular shift: {min_clearance_ft:.2f} × √(1+m²) = {perpendicular_shift:.2f} ft")
            
            # Apply to all points
            for i, point in enumerate(profile_data):
                dist = point['distance']
                ground_elev = point['elevation']
                
                # Running line is straight: y = mx + b
                running_line_elev = slope * dist + running_intercept
                
                # Calculate depths
                vertical_depth_ft = ground_elev - running_line_elev
                vertical_depth_inches = vertical_depth_ft * 12.0
                
                # Perpendicular distance from point to line
                perpendicular_distance_ft = abs(ground_elev - (slope * dist + running_intercept)) / perpendicular_factor
                perpendicular_distance_inches = perpendicular_distance_ft * 12.0
                
                point['running_line'] = running_line_elev
                point['cut_fill'] = vertical_depth_ft
                point['cover'] = vertical_depth_ft
                point['depth_inches'] = vertical_depth_inches
                point['perpendicular_distance_inches'] = perpendicular_distance_inches
                point['segment_slope'] = slope
                point['segment_id'] = 0
                
                if i < 3 or i >= len(profile_data) - 3:
                    logging.info(f"Point {i}: dist={dist:.1f}ft, ground={ground_elev:.2f}ft, "
                               f"running={running_line_elev:.2f}ft, cover={vertical_depth_inches:.1f}in")
        
        else:
            # Multi-segment design with break points
            logging.info(f"Calculating MULTI-SEGMENT running line with {len(break_points)} break points")
            logging.info(f"Running line connects break points: P1→P2→P3...→Pn")
            logging.info(f"Target perpendicular distance at break points: {min_clearance_ft:.2f} ft ({self.running_line_depth_inches:.0f} inches)")
            
            # Create control points: Start + Break Points + End
            # Each break point gets running line elevation = ground - min_clearance
            control_points = []
            
            # First point (start)
            first_ground = profile_data[0]['elevation']
            first_dist = profile_data[0]['distance']
            first_running = first_ground - min_clearance_ft
            control_points.append((first_dist, first_running, "Start"))
            
            # Break points (bp_ground_elev is already at the desired depth below ground)
            for bp_dist, bp_ground_elev, bp_grade, bp_locked, bp_note in sorted(break_points, key=lambda x: x[0]):
                bp_running = bp_ground_elev  # Don't subtract again - already at target depth
                control_points.append((bp_dist, bp_running, f"BP: {bp_note or 'Auto'}"))
            
            # Last point (end)
            last_ground = profile_data[-1]['elevation']
            last_dist = profile_data[-1]['distance']
            last_running = last_ground - min_clearance_ft
            control_points.append((last_dist, last_running, "End"))
            
            logging.info(f"Control points for running line:")
            for i, (dist, elev, note) in enumerate(control_points):
                logging.info(f"  P{i+1} ({note}): {dist:.1f} ft, elevation {elev:.2f} ft")
            
            # Create segments between control points
            segments = []
            for i in range(len(control_points) - 1):
                start_dist, start_elev, start_note = control_points[i]
                end_dist, end_elev, end_note = control_points[i + 1]
                
                # Find point indices for this segment
                start_idx = next((j for j, p in enumerate(profile_data) if p['distance'] >= start_dist), 0)
                end_idx = next((j for j, p in enumerate(profile_data) if p['distance'] >= end_dist), len(profile_data) - 1)
                
                # Calculate slope between control points: (y2-y1)/(x2-x1)
                seg_slope = (end_elev - start_elev) / (end_dist - start_dist) if end_dist != start_dist else 0
                
                # Calculate y-intercept: b = y - mx
                seg_intercept = start_elev - seg_slope * start_dist
                
                segments.append({
                    'index': i + 1,
                    'start_dist': start_dist,
                    'end_dist': end_dist,
                    'start_idx': start_idx,
                    'end_idx': end_idx,
                    'slope': seg_slope,
                    'intercept': seg_intercept,
                    'start_note': start_note,
                    'end_note': end_note
                })
                
                grade_percent = seg_slope * 100
                logging.info(f"Segment {i+1} ({start_note}→{end_note}): "
                           f"{start_dist:.1f}-{end_dist:.1f} ft, "
                           f"slope={seg_slope:.6f} ({grade_percent:+.2f}%), "
                           f"intercept={seg_intercept:.2f}")
            
            # Apply running line equation to all points
            for seg in segments:
                seg_points = profile_data[seg['start_idx']:seg['end_idx'] + 1]
                
                for point in seg_points:
                    dist = point['distance']
                    ground_elev = point['elevation']
                    
                    # Running line: y = mx + b
                    running_line_elev = seg['slope'] * dist + seg['intercept']
                    
                    # Calculate depths
                    vertical_depth_ft = ground_elev - running_line_elev
                    vertical_depth_inches = vertical_depth_ft * 12.0
                    
                    # Perpendicular distance
                    seg_perpendicular_factor = np.sqrt(1 + seg['slope']**2)
                    perpendicular_distance_ft = abs(vertical_depth_ft) / seg_perpendicular_factor
                    perpendicular_distance_inches = perpendicular_distance_ft * 12.0
                    
                    point['running_line'] = running_line_elev
                    point['cut_fill'] = vertical_depth_ft
                    point['cover'] = vertical_depth_ft
                    point['depth_inches'] = vertical_depth_inches
                    point['perpendicular_distance_inches'] = perpendicular_distance_inches
                    point['segment_slope'] = seg['slope']
                    point['segment_id'] = seg['index'] - 1
        
        # Calculate statistics
        cover_depths = [p['depth_inches'] for p in profile_data]
        
        min_cover = min(cover_depths)
        max_cover = max(cover_depths)
        avg_cover = sum(cover_depths) / len(cover_depths)
        
        logging.info(f"Cover depth range: {min_cover:.1f} to {max_cover:.1f} inches (avg: {avg_cover:.1f} inches)")
        
        if not break_points:
            logging.info(f"Single-segment design maintains minimum {self.running_line_depth_inches:.0f} inches perpendicular distance")
        else:
            logging.info(f"Multi-segment design with {len(break_points)} break points")
    
    def _filter_redundant_points(self, profile_data, slope_threshold=0.01):
        """Filter out redundant points where slope doesn't change significantly
        
        Keeps points where:
        1. Slope changes significantly from previous segment
        2. Elevation change from last kept point exceeds threshold
        """
        if len(profile_data) <= 3:
            return profile_data
        
        # Always keep first point
        filtered = [profile_data[0]]
        
        # Iterate through middle points
        prev_slope = None
        
        for i in range(1, len(profile_data) - 1):
            last_kept = filtered[-1]
            curr_point = profile_data[i]
            next_point = profile_data[i + 1]
            
            # Calculate current segment slope (from last kept point to current)
            curr_slope = (curr_point['elevation'] - last_kept['elevation']) / (curr_point['distance'] - last_kept['distance']) if curr_point['distance'] != last_kept['distance'] else 0
            
            # Calculate next segment slope (from current to next)
            next_slope = (next_point['elevation'] - curr_point['elevation']) / (next_point['distance'] - curr_point['distance']) if next_point['distance'] != curr_point['distance'] else 0
            
            # Calculate slope change
            if prev_slope is not None:
                slope_change = abs(next_slope - curr_slope)
            else:
                slope_change = float('inf')  # First point after start - always keep
            
            # Calculate elevation change from last kept point
            elevation_change = abs(curr_point['elevation'] - last_kept['elevation'])
            
            # Keep point if:
            # 1. Slope changes significantly, OR
            # 2. Elevation changed by more than 2 feet from last kept point
            if slope_change > slope_threshold or elevation_change > 2.0:
                filtered.append(curr_point)
                prev_slope = curr_slope
        
        # Always keep last point
        filtered.append(profile_data[-1])
        
        return filtered
        
    def _recalculate_running_line(self):
        """Recalculate running line elevations based on current depth zones or regression"""
        if not self.profile_data:
            return
        
        # If depth zones are defined, use zone-based calculation
        if self.depth_zones:
            for point in self.profile_data:
                # Get depth at this distance (considers zones)
                point_depth_inches = self._get_depth_at_distance(point['distance'])
                depth_ft = point_depth_inches / 12.0
                
                point['running_line'] = point['elevation'] - depth_ft
                point['cut_fill'] = depth_ft
                point['cover'] = depth_ft
                point['depth_inches'] = point_depth_inches
            
            logging.info("Running line recalculated with updated depth zones")
        else:
            # Use regression method
            self._calculate_running_line_regression(self.profile_data)
            logging.info("Running line recalculated using linear regression")
    
    def _on_chart_click(self, event):
        """Handle click events on the chart for depth zone editing and break point placement"""
        
        # Check if break point mode is active
        if self.break_point_edit_mode:
            if not self.profile_data:
                return
            
            # Check if click is within the axes
            if event.inaxes != self.chart_ax:
                return
            
            # Get the x coordinate (distance)
            distance_ft = event.xdata
            
            if distance_ft is None:
                return
            
            # Validate distance is within profile range (not at endpoints)
            max_distance = max(p['distance'] for p in self.profile_data)
            min_distance = min(p['distance'] for p in self.profile_data)
            
            if distance_ft <= min_distance + 5 or distance_ft >= max_distance - 5:
                logging.warning("Cannot place break point at start or end - must be in the middle")
                return
            
            # Add break point
            self._add_break_point(distance_ft)
            return
        
        # Otherwise handle depth zone editing
        if not self.depth_edit_mode:
            return
        
        if not self.profile_data:
            return
        
        # Check if click is within the axes
        if event.inaxes != self.chart_ax:
            return
        
        # Get the x coordinate (distance)
        distance_ft = event.xdata
        
        if distance_ft is None:
            return
        
        # Validate distance is within profile range
        max_distance = max(p['distance'] for p in self.profile_data)
        if distance_ft < 0 or distance_ft > max_distance:
            return
        
        # Two-click zone definition
        if self.pending_zone_start is None:
            # First click - set start point
            self.pending_zone_start = distance_ft
            self.depth_edit_btn.configure(
                text=f"📍 Start: {distance_ft:.0f} ft - Click End Point",
                fg_color="#E67E22",
                hover_color="#D35400"
            )
            logging.info(f"Zone start set at {distance_ft:.0f} ft - Now click the END point")
            
            # Draw a temporary marker on the chart
            self._draw_profile_chart()  # Redraw to show marker
        else:
            # Second click - set end point and open dialog
            start_ft = self.pending_zone_start
            end_ft = distance_ft
            self.pending_zone_start = None
            
            # Reset button
            self.depth_edit_btn.configure(
                text="✅ EDIT MODE ON - Click Start Point",
                fg_color="#27AE60",
                hover_color="#229954"
            )
            
            # Open dialog to set depth for this zone
            self._add_depth_zone(start_ft, end_ft)
        
    def _toggle_stations(self):
        """Toggle station markers visibility"""
        self.enable_stations = self.enable_station_checkbox.get()
        if self.profile_data:
            self._draw_profile_chart()
        logging.info(f"Station markers: {'Enabled' if self.enable_stations else 'Disabled'}")
        
    def _set_station_interval(self, interval):
        """Set station interval from preset button"""
        self.station_entry.delete(0, tk.END)
        self.station_entry.insert(0, str(interval))
        self.station_interval_ft = interval
        if self.profile_data:
            self._draw_profile_chart()
        logging.info(f"Station interval set to: {interval} feet")
    
    def _toggle_view(self, is_3d):
        """Toggle between 2D and 3D views"""
        self.view_3d = is_3d
        
        # Update button states
        if is_3d:
            self.view_2d_btn.configure(fg_color="gray40", hover_color="gray30")
            self.view_3d_btn.configure(fg_color="#42A5F5", hover_color="#2196F3")
        else:
            self.view_2d_btn.configure(fg_color="#42A5F5", hover_color="#2196F3")
            self.view_3d_btn.configure(fg_color="gray40", hover_color="gray30")
        
        # Redraw chart
        if self.profile_data:
            self._draw_profile_chart()
        
        logging.info(f"View changed to: {'3D' if is_3d else '2D'}")
    
    def _download_imagery_dialog(self):
        """Show dialog to configure and download satellite imagery"""
        if not self.waypoints or len(self.waypoints) < 2:
            messagebox.showwarning("No Route", "Please add at least 2 waypoints to define a route before downloading imagery.")
            return
        
        # Create dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title("Download Satellite Imagery")
        dialog.geometry("750x600")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 375
        y = (dialog.winfo_screenheight() // 2) - 300
        dialog.geometry(f"750x600+{x}+{y}")
        
        # Main frame (NO SCROLL)
        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(
            main_frame,
            text="🛰️ NAIP Satellite Imagery Download",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title.pack(pady=(0, 3))
        
        # Description
        desc = ctk.CTkLabel(
            main_frame,
            text="Download high-resolution aerial imagery along your route",
            font=ctk.CTkFont(size=10),
            text_color="gray70"
        )
        desc.pack(pady=(0, 15))
        
        # === TWO COLUMN LAYOUT ===
        columns_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        columns_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # LEFT COLUMN
        left_column = ctk.CTkFrame(columns_frame, fg_color="transparent")
        left_column.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # === QUALITY SECTION ===
        quality_section = ctk.CTkFrame(left_column, fg_color="gray20", corner_radius=10)
        quality_section.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        quality_label = ctk.CTkLabel(
            quality_section,
            text="📐 Image Quality",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        quality_label.pack(anchor="w", padx=12, pady=(12, 8))
        
        # Quality slider
        quality_slider_frame = ctk.CTkFrame(quality_section, fg_color="transparent")
        quality_slider_frame.pack(fill=tk.X, padx=12, pady=(0, 8))
        
        quality_slider = ctk.CTkSlider(
            quality_slider_frame,
            from_=1,
            to=4,
            number_of_steps=3,
            command=lambda v: quality_value_label.configure(text=self._get_quality_text(int(v)))
        )
        quality_slider.set(self.imagery_quality)
        quality_slider.pack(fill=tk.X, pady=(0, 5))
        
        # Quality labels
        labels_frame = ctk.CTkFrame(quality_slider_frame, fg_color="transparent")
        labels_frame.pack(fill=tk.X)
        
        ctk.CTkLabel(labels_frame, text="Low", font=ctk.CTkFont(size=9), text_color="gray60").pack(side=tk.LEFT)
        ctk.CTkLabel(labels_frame, text="Very High", font=ctk.CTkFont(size=9), text_color="gray60").pack(side=tk.RIGHT)
        
        quality_value_label = ctk.CTkLabel(
            quality_section,
            text=self._get_quality_text(self.imagery_quality),
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#2196F3"
        )
        quality_value_label.pack(pady=(3, 12))
        
        # === BUFFER SECTION ===
        buffer_section = ctk.CTkFrame(left_column, fg_color="gray20", corner_radius=10)
        buffer_section.pack(fill=tk.BOTH, expand=True)
        
        buffer_label = ctk.CTkLabel(
            buffer_section,
            text="📏 Buffer Distance",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        buffer_label.pack(anchor="w", padx=12, pady=(12, 5))
        
        buffer_desc = ctk.CTkLabel(
            buffer_section,
            text="Area around route to include",
            font=ctk.CTkFont(size=9),
            text_color="gray60"
        )
        buffer_desc.pack(anchor="w", padx=12, pady=(0, 8))
        
        # Buffer slider (100 to 2000 feet)
        buffer_slider_frame = ctk.CTkFrame(buffer_section, fg_color="transparent")
        buffer_slider_frame.pack(fill=tk.X, padx=12, pady=(0, 8))
        
        buffer_slider = ctk.CTkSlider(
            buffer_slider_frame,
            from_=100,
            to=2000,
            number_of_steps=19,
            command=lambda v: buffer_value_label.configure(text=f"{int(v)} feet")
        )
        buffer_slider.set(2000)
        buffer_slider.pack(fill=tk.X, pady=(0, 5))
        
        # Buffer labels
        buffer_labels_frame = ctk.CTkFrame(buffer_slider_frame, fg_color="transparent")
        buffer_labels_frame.pack(fill=tk.X)
        
        ctk.CTkLabel(buffer_labels_frame, text="100 ft", font=ctk.CTkFont(size=9), text_color="gray60").pack(side=tk.LEFT)
        ctk.CTkLabel(buffer_labels_frame, text="2000 ft", font=ctk.CTkFont(size=9), text_color="gray60").pack(side=tk.RIGHT)
        
        buffer_value_label = ctk.CTkLabel(
            buffer_section,
            text="2000 feet",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#4CAF50"
        )
        buffer_value_label.pack(pady=(3, 12))
        
        # RIGHT COLUMN
        right_column = ctk.CTkFrame(columns_frame, fg_color="transparent")
        right_column.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # === FORMAT SECTION ===
        format_section = ctk.CTkFrame(right_column, fg_color="gray20", corner_radius=10)
        format_section.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        format_label = ctk.CTkLabel(
            format_section,
            text="💾 Output Format",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        format_label.pack(anchor="w", padx=12, pady=(12, 8))
        
        # Format radio buttons
        format_var = tk.StringVar(value="tiff")
        
        formats = [
            ("tiff", "GeoTIFF (.tif)", "For GIS software"),
            ("png", "PNG (.png)", "General purpose"),
            ("jpg", "JPEG (.jpg)", "Smallest size")
        ]
        
        for val, text, desc_text in formats:
            radio_frame = ctk.CTkFrame(format_section, fg_color="transparent")
            radio_frame.pack(fill=tk.X, padx=12, pady=2)
            
            radio = ctk.CTkRadioButton(
                radio_frame,
                text=text,
                variable=format_var,
                value=val,
                font=ctk.CTkFont(size=10, weight="bold")
            )
            radio.pack(anchor="w")
            
            desc_label = ctk.CTkLabel(
                radio_frame,
                text=f"  {desc_text}",
                font=ctk.CTkFont(size=8),
                text_color="gray60"
            )
            desc_label.pack(anchor="w", padx=(20, 0))
        
        # Spacing
        ctk.CTkLabel(format_section, text="").pack(pady=8)
        
        # === DOWNLOAD BUTTON IN RIGHT COLUMN ===
        download_btn = ctk.CTkButton(
            right_column,
            text="📥 Download Imagery",
            height=45,
            command=lambda: self._execute_imagery_download(
                int(quality_slider.get()), 
                int(buffer_slider.get()),
                format_var.get(),
                False,  # No worldfile
                False,  # No metadata
                dialog
            ),
            fg_color="#2E7D32",
            hover_color="#1B5E20",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        download_btn.pack(fill=tk.X, pady=(10, 0))
        
        # === INFO BOX ===
        info_frame = ctk.CTkFrame(main_frame, fg_color="#1a1a1a", corner_radius=8)
        info_frame.pack(fill=tk.X, pady=(5, 10))
        
        info_text = ctk.CTkLabel(
            info_frame,
            text=f"ℹ️ Route: {len(self.waypoints)} waypoints  •  View: {self.map_type.capitalize()}  •  Source: USGS NAIP",
            font=ctk.CTkFont(size=10),
            justify="left"
        )
        info_text.pack(padx=12, pady=10)
        
        # === CANCEL BUTTON ===
        cancel_btn = ctk.CTkButton(
            main_frame,
            text="Cancel",
            height=35,
            command=dialog.destroy,
            fg_color="gray40",
            hover_color="gray30"
        )
        cancel_btn.pack(fill=tk.X)
    
    def _get_quality_text(self, quality_level):
        """Get descriptive text for quality level"""
        quality_map = {
            1: "Low (512x512, ~50KB, Fast)",
            2: "Medium (1024x1024, ~200KB, Balanced)",
            3: "High (2048x2048, ~800KB, Detailed)",
            4: "Very High (4096x4096, ~3MB, Best Quality)"
        }
        return quality_map.get(quality_level, "Medium")
    
    def _execute_imagery_download(self, quality_level, buffer_ft, output_format, include_worldfile, include_metadata, dialog):
        """Execute the imagery download"""
        dialog.destroy()
        
        # Save quality preference
        self.imagery_quality = quality_level
        
        # Get file extension based on format
        ext_map = {'tiff': '.tif', 'png': '.png', 'jpg': '.jpg'}
        file_ext = ext_map.get(output_format, '.tif')
        
        # File types for dialog
        filetypes_map = {
            'tiff': [('GeoTIFF', '*.tif'), ('TIFF', '*.tiff'), ('All Files', '*.*')],
            'png': [('PNG Image', '*.png'), ('All Files', '*.*')],
            'jpg': [('JPEG Image', '*.jpg *.jpeg'), ('All Files', '*.*')]
        }
        
        # File dialog
        filename = filedialog.asksaveasfilename(
            defaultextension=file_ext,
            filetypes=filetypes_map.get(output_format, [('All Files', '*.*')]),
            initialfile=f'satellite_imagery_{datetime.now().strftime("%Y%m%d_%H%M%S")}{file_ext}'
        )
        
        if not filename:
            return
        
        # Start download in background thread
        self.progress_bar.set(0.1)
        logging.info(f"Starting imagery download: Quality={self._get_quality_text(quality_level)}, Buffer={buffer_ft}ft, Format={output_format}")
        
        def download():
            try:
                self._download_naip_imagery(filename, quality_level, buffer_ft, output_format, include_worldfile, include_metadata)
                self.after(0, lambda: self._on_imagery_download_success(filename))
            except Exception as e:
                self.after(0, lambda: self._on_imagery_download_error(str(e)))
        
        threading.Thread(target=download, daemon=True).start()
    
    def _download_naip_imagery(self, filename, quality_level, buffer_ft, output_format, include_worldfile, include_metadata):
        """Download NAIP imagery using USGS ArcGIS REST service"""
        # Calculate bounding box from waypoints
        lats = [wp[0] for wp in self.waypoints]
        lons = [wp[1] for wp in self.waypoints]
        
        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)
        
        # Convert buffer from feet to degrees (approximate: 1 degree ~= 364,000 feet at equator)
        # More accurate: 1 foot = 1/(364,000 * cos(latitude)) degrees
        avg_lat = (min_lat + max_lat) / 2
        import math
        buffer_deg = buffer_ft / (364000 * math.cos(math.radians(avg_lat)))
        
        # Map quality to image size
        size_map = {1: 512, 2: 1024, 3: 2048, 4: 4096}
        image_size = size_map.get(quality_level, 1024)
        
        # Map output format to REST API format parameter
        format_map = {
            'tiff': 'tiff',
            'png': 'png32',
            'jpg': 'jpg'
        }
        rest_format = format_map.get(output_format, 'tiff')
        
        # Use ArcGIS REST Export Image endpoint (more reliable than WMS)
        export_url = "https://basemap.nationalmap.gov/arcgis/rest/services/USGSImageryOnly/MapServer/export"
        
        # Calculate proper bbox string: xmin,ymin,xmax,ymax (west,south,east,north)
        bbox_str = f"{min_lon - buffer_deg},{min_lat - buffer_deg},{max_lon + buffer_deg},{max_lat + buffer_deg}"
        
        # Store bbox coordinates for worldfile and metadata
        xmin, ymin, xmax, ymax = min_lon - buffer_deg, min_lat - buffer_deg, max_lon + buffer_deg, max_lat + buffer_deg
        
        params = {
            'bbox': bbox_str,
            'bboxSR': 4326,  # WGS84
            'size': f"{image_size},{image_size}",
            'imageSR': 4326,
            'format': rest_format,
            'pixelType': 'U8',  # 8-bit unsigned
            'noData': '',
            'noDataInterpretation': 'esriNoDataMatchAny',
            'interpolation': 'RSP_BilinearInterpolation',
            'compression': 'LZ77' if rest_format == 'tiff' else 'None',
            'compressionQuality': 75,
            'bandIds': '',
            'mosaicRule': '',
            'renderingRule': '',
            'f': 'image'
        }
        
        logging.info(f"Downloading imagery: {image_size}x{image_size} pixels, buffer={buffer_ft}ft ({buffer_deg:.6f}°), format={output_format}")
        logging.info(f"Bbox: {bbox_str}")
        logging.info(f"URL: {export_url}")
        self.after(0, lambda: self.progress_bar.set(0.3))
        
        # Download
        response = requests.get(export_url, params=params, timeout=90)
        response.raise_for_status()
        
        # Check if we got an error message instead of image
        content_type = response.headers.get('content-type', '')
        if 'json' in content_type or 'text' in content_type:
            error_msg = response.text[:500]
            raise Exception(f"Server returned error: {error_msg}")
        
        self.after(0, lambda: self.progress_bar.set(0.7))
        
        # Save to file
        with open(filename, 'wb') as f:
            f.write(response.content)
        
        file_size_kb = len(response.content) / 1024
        logging.info(f"Imagery downloaded: {file_size_kb:.1f} KB")
        
        # Create worldfile if requested
        if include_worldfile:
            self._create_worldfile(filename, output_format, xmin, ymin, xmax, ymax, image_size)
        
        # Create metadata file if requested
        if include_metadata:
            self._create_metadata_file(filename, quality_level, buffer_ft, output_format, 
                                      xmin, ymin, xmax, ymax, image_size, file_size_kb)
        
        self.after(0, lambda: self.progress_bar.set(1.0))
    
    def _create_worldfile(self, image_filename, output_format, xmin, ymin, xmax, ymax, image_size):
        """Create a worldfile for the downloaded imagery"""
        # Calculate pixel size (resolution)
        pixel_width = (xmax - xmin) / image_size
        pixel_height = (ymax - ymin) / image_size
        
        # Worldfile format (6 lines):
        # Line 1: pixel size in x-direction
        # Line 2: rotation about y-axis (usually 0)
        # Line 3: rotation about x-axis (usually 0)
        # Line 4: pixel size in y-direction (negative)
        # Line 5: x-coordinate of center of upper left pixel
        # Line 6: y-coordinate of center of upper left pixel
        
        worldfile_content = f"{pixel_width}\n0.0\n0.0\n{-pixel_height}\n{xmin + pixel_width/2}\n{ymax - pixel_height/2}\n"
        
        # Determine worldfile extension based on image format
        ext_map = {
            'tiff': '.tfw',
            'png': '.pgw',
            'jpg': '.jgw'
        }
        worldfile_ext = ext_map.get(output_format, '.tfw')
        
        # Replace image extension with worldfile extension
        worldfile_path = os.path.splitext(image_filename)[0] + worldfile_ext
        
        with open(worldfile_path, 'w') as f:
            f.write(worldfile_content)
        
        logging.info(f"Worldfile created: {worldfile_path}")
    
    def _create_metadata_file(self, image_filename, quality_level, buffer_ft, output_format,
                             xmin, ymin, xmax, ymax, image_size, file_size_kb):
        """Create a metadata file with download information"""
        from datetime import datetime
        
        # Calculate coverage area
        import math
        # Approximate area calculation (not perfect for large areas)
        width_deg = xmax - xmin
        height_deg = ymax - ymin
        
        # Convert to miles (very approximate)
        miles_per_deg_lat = 69.0
        avg_lat = (ymin + ymax) / 2
        miles_per_deg_lon = 69.0 * math.cos(math.radians(avg_lat))
        
        width_miles = width_deg * miles_per_deg_lon
        height_miles = height_deg * miles_per_deg_lat
        area_sq_miles = width_miles * height_miles
        
        # Calculate pixel resolution in feet
        pixel_size_deg = (xmax - xmin) / image_size
        pixel_size_ft = pixel_size_deg * (364000 * math.cos(math.radians(avg_lat)))
        
        metadata_content = f"""SATELLITE IMAGERY DOWNLOAD METADATA
=====================================

Download Date/Time: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Source: USGS National Agriculture Imagery Program (NAIP)
Application: Profile Automation Tool

IMAGE PROPERTIES:
-----------------
Format: {output_format.upper()}
Dimensions: {image_size} x {image_size} pixels
File Size: {file_size_kb:.2f} KB
Quality Level: {self._get_quality_text(quality_level)}
Pixel Resolution: ~{pixel_size_ft:.2f} feet/pixel

GEOGRAPHIC EXTENT:
------------------
Coordinate System: WGS84 (EPSG:4326)
Bounding Box (Decimal Degrees):
  West:  {xmin:.6f}
  South: {ymin:.6f}
  East:  {xmax:.6f}
  North: {ymax:.6f}

Coverage Area:
  Width: {width_miles:.3f} miles ({width_deg:.6f}°)
  Height: {height_miles:.3f} miles ({height_deg:.6f}°)
  Total Area: {area_sq_miles:.4f} square miles

Buffer Distance: {buffer_ft} feet from route

ROUTE INFORMATION:
------------------
Number of Waypoints: {len(self.waypoints)}
Waypoint Coordinates:
"""
        
        # Add waypoint list
        for i, (lat, lon) in enumerate(self.waypoints, 1):
            metadata_content += f"  Point {i}: {lat:.6f}, {lon:.6f}\n"
        
        metadata_content += f"\n"
        metadata_content += f"USAGE NOTES:\n"
        metadata_content += f"------------\n"
        metadata_content += f"This imagery can be opened in GIS software such as:\n"
        metadata_content += f"• ArcGIS / ArcGIS Pro\n"
        metadata_content += f"• QGIS (free)\n"
        metadata_content += f"• Google Earth Pro\n"
        metadata_content += f"• Any software that supports georeferenced imagery\n"
        
        if output_format == 'tiff':
            metadata_content += f"\nThe GeoTIFF format includes built-in georeferencing.\n"
        else:
            metadata_content += f"\nUse the accompanying worldfile (.{['tfw', 'pgw', 'jgw'][['tiff', 'png', 'jpg'].index(output_format)]}) for georeferencing.\n"
        
        # Save metadata file
        metadata_path = os.path.splitext(image_filename)[0] + '_metadata.txt'
        with open(metadata_path, 'w') as f:
            f.write(metadata_content)
        
        logging.info(f"Metadata file created: {metadata_path}")
    
    def _on_imagery_download_success(self, filename):
        """Handle successful imagery download"""
        self.progress_bar.set(0)
        logging.info(f"Imagery saved successfully: {filename}")
        
        # Check which optional files were created
        base_name = os.path.splitext(filename)[0]
        file_ext = os.path.splitext(filename)[1].lower()
        
        additional_files = []
        
        # Check for worldfile
        worldfile_exts = {'.tif': '.tfw', '.tiff': '.tfw', '.png': '.pgw', '.jpg': '.jgw', '.jpeg': '.jgw'}
        worldfile_ext = worldfile_exts.get(file_ext, '.tfw')
        if os.path.exists(base_name + worldfile_ext):
            additional_files.append(f"• Worldfile ({worldfile_ext})")
        
        # Check for metadata
        if os.path.exists(base_name + '_metadata.txt'):
            additional_files.append("• Metadata file (_metadata.txt)")
        
        message = f"Satellite imagery downloaded successfully!\n\n"
        message += f"File: {os.path.basename(filename)}\n"
        message += f"Location: {os.path.dirname(filename)}\n"
        
        if additional_files:
            message += f"\nAdditional files created:\n" + "\n".join(additional_files)
        
        message += f"\n\nYou can open this imagery in:\n"
        message += f"• ArcGIS / QGIS\n"
        message += f"• Google Earth Pro\n"
        message += f"• Any GIS software"
        
        messagebox.showinfo("Download Complete", message)
    
    def _on_imagery_download_error(self, error):
        """Handle imagery download error"""
        self.progress_bar.set(0)
        logging.error(f"Imagery download failed: {error}")
        messagebox.showerror(
            "Download Failed",
            f"Failed to download satellite imagery:\n\n{error}\n\n"
            f"Please check your internet connection and try again."
        )
    
    def _toggle_map_type(self):
        """Toggle between street and satellite map views"""
        if self.map_type == "street":
            # Switch to NAIP satellite imagery (optimized URL)
            self.map_widget.set_tile_server(
                "https://basemap.nationalmap.gov/arcgis/rest/services/USGSImageryOnly/MapServer/tile/{z}/{y}/{x}",
                max_zoom=19  # Increased max zoom for better detail
            )
            self.map_type = "satellite"
            self.map_toggle_btn.configure(
                text="🗺️ Street",
                fg_color="#2E7D32",
                hover_color="#1B5E20"
            )
            logging.info("Map view: Switched to NAIP Satellite Imagery (USGS)")
        else:
            # Switch back to OpenStreetMap (with faster CDN)
            self.map_widget.set_tile_server(
                "https://tile.openstreetmap.org/{z}/{x}/{y}.png",
                max_zoom=19  # Increased max zoom
            )
            self.map_type = "street"
            self.map_toggle_btn.configure(
                text="🛰️ Satellite",
                fg_color="#0078D4",
                hover_color="#005A9E"
            )
            logging.info("Map view: Switched to OpenStreetMap")
    
    def _toggle_point_placement(self):
        """Toggle point placement mode on/off"""
        self.point_placement_active = not self.point_placement_active
        
        if self.point_placement_active:
            self.placement_toggle_btn.configure(
                text="📍 Add Points: ON",
                fg_color="#2E7D32",
                hover_color="#1B5E20"
            )
            logging.info("Point placement mode: ACTIVE - Click on map to add waypoints")
        else:
            self.placement_toggle_btn.configure(
                text="📍 Add Points: OFF",
                fg_color="gray40",
                hover_color="gray30"
            )
            logging.info("Point placement mode: INACTIVE - Zoom buttons safe to use")
        
    def _on_map_click(self, coordinates):
        """Handle map click to add waypoints"""
        # Only add points if placement mode is active
        if not self.point_placement_active:
            return
        
        lat, lon = coordinates
        
        # Add waypoint
        self.waypoints.append((lat, lon))
        waypoint_num = len(self.waypoints)
        
        # Determine marker color (start=green, end=red, middle=blue)
        if waypoint_num == 1:
            color_circle = "#27AE60"
            color_outside = "#229954"
            text = "Start"
        else:
            color_circle = "#42A5F5"
            color_outside = "#2196F3"
            text = f"Point {waypoint_num}"
        
        # Create marker
        marker = self.map_widget.set_marker(
            lat, lon,
            text=text,
            marker_color_circle=color_circle,
            marker_color_outside=color_outside
        )
        self.waypoint_markers.append(marker)
        
        # Update path
        if waypoint_num > 1:
            if self.path_line:
                self.path_line.delete()
            self.path_line = self.map_widget.set_path(
                self.waypoints,
                color="#42A5F5",
                width=3
            )
            
            # Calculate total distance
            total_distance = 0
            for i in range(len(self.waypoints) - 1):
                dist_km = self._calculate_distance_km(self.waypoints[i], self.waypoints[i+1])
                total_distance += dist_km
            
            total_distance_ft = total_distance * 3280.84
            logging.info(f"Waypoint {waypoint_num} added: ({lat:.6f}, {lon:.6f})")
            logging.info(f"Total path distance: {total_distance_ft:,.0f} feet ({total_distance:.2f} km)")
        else:
            logging.info(f"Start point set: ({lat:.6f}, {lon:.6f})")
        
        # Update UI
        self._update_waypoint_display()
            
    def _clear_points(self):
        """Clear all waypoints and reset"""
        # Delete all markers
        for marker in self.waypoint_markers:
            marker.delete()
        self.waypoint_markers.clear()
        
        # Delete path
        if self.path_line:
            self.path_line.delete()
            self.path_line = None
        
        # Clear waypoints and selection
        self.waypoints.clear()
        self.waypoint_selection = None
        
        # Clear depth zones
        self.depth_zones = []
        self.pending_zone_start = None
        self._update_overrides_display()
        
        # Update UI
        self._update_waypoint_display()
        logging.info("All waypoints cleared. Click points on the map to create a new path.")
        
    def _undo_last_waypoint(self):
        """Remove the last waypoint"""
        if not self.waypoints:
            return
        
        # Remove last waypoint
        self.waypoints.pop()
        
        # Remove last marker
        if self.waypoint_markers:
            marker = self.waypoint_markers.pop()
            marker.delete()
        
        # Update path
        if self.path_line:
            self.path_line.delete()
            self.path_line = None
        
        if len(self.waypoints) > 1:
            self.path_line = self.map_widget.set_path(
                self.waypoints,
                color="#42A5F5",
                width=3
            )
        
        # Update UI
        self._update_waypoint_display()
        logging.info(f"Last waypoint removed. {len(self.waypoints)} waypoints remaining.")
        
    def _open_waypoint_selector(self):
        """Open dialog to select waypoint range for profile generation"""
        if len(self.waypoints) < 2:
            messagebox.showinfo("Not Enough Waypoints",
                "You need at least 2 waypoints to select a range.\n\n"
                "Add more waypoints by clicking on the map or importing a file.")
            return
        
        # Auto-select all if only 2 waypoints
        if len(self.waypoints) == 2:
            self.waypoint_selection = [0, 1]
            self._refresh_map_markers()
            self._update_waypoint_display()
            messagebox.showinfo("Range Auto-Selected",
                "Both waypoints selected automatically.\n\n"
                "Click 'Generate Profile' to create the elevation profile.")
            return
        
        # Create dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title("Select Waypoint Range")
        dialog.geometry("600x700")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (700 // 2)
        dialog.geometry(f"600x700+{x}+{y}")
        
        # Title
        title = ctk.CTkLabel(
            dialog,
            text="✂️ Select Waypoint Range",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title.pack(pady=20, padx=20)
        
        desc = ctk.CTkLabel(
            dialog,
            text="Select start and end waypoints to generate profile for a specific segment",
            font=ctk.CTkFont(size=12),
            text_color="gray70"
        )
        desc.pack(pady=(0, 20), padx=20)
        
        # Frame for listbox
        list_frame = ctk.CTkFrame(dialog)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Create listbox with scrollbar
        list_container = ctk.CTkFrame(list_frame)
        list_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = ctk.CTkScrollbar(list_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        waypoint_listbox = tk.Listbox(
            list_container,
            selectmode=tk.EXTENDED,
            font=("Consolas", 10),
            bg="#2b2b2b",
            fg="white",
            selectbackground="#1f6aa5",
            yscrollcommand=scrollbar.set,
            height=20
        )
        waypoint_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.configure(command=waypoint_listbox.yview)
        
        # Populate listbox
        total_dist = 0
        for i, (lat, lon) in enumerate(self.waypoints):
            if i > 0:
                dist = self._calculate_distance(
                    self.waypoints[i-1][0], self.waypoints[i-1][1],
                    lat, lon
                )
                total_dist += dist
            
            if i == 0:
                label = f"[START]  Point {i+1:3d}  |  {lat:9.6f}, {lon:9.6f}  |  0.0 ft"
            else:
                label = f"         Point {i+1:3d}  |  {lat:9.6f}, {lon:9.6f}  |  {total_dist:,.0f} ft"
            
            waypoint_listbox.insert(tk.END, label)
        
        # Instructions
        instr_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        instr_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        instr = ctk.CTkLabel(
            instr_frame,
            text="💡 Click start point, then hold Ctrl and click end point\n"
                 "Or select multiple points to include intermediate vertices",
            font=ctk.CTkFont(size=11),
            text_color="gray60",
            justify="left"
        )
        instr.pack(anchor="w")
        
        # Quick select buttons
        quick_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        quick_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        ctk.CTkLabel(
            quick_frame,
            text="Quick Select:",
            font=ctk.CTkFont(size=11, weight="bold")
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ctk.CTkButton(
            quick_frame,
            text="All Points",
            width=90,
            height=28,
            command=lambda: waypoint_listbox.selection_set(0, tk.END),
            fg_color="gray40",
            hover_color="gray30"
        ).pack(side=tk.LEFT, padx=2)
        
        ctk.CTkButton(
            quick_frame,
            text="Clear",
            width=70,
            height=28,
            command=lambda: waypoint_listbox.selection_clear(0, tk.END),
            fg_color="gray40",
            hover_color="gray30"
        ).pack(side=tk.LEFT, padx=2)
        
        def remove_selected_points():
            selected = list(waypoint_listbox.curselection())
            if not selected:
                messagebox.showwarning("No Selection", "Please select waypoints to remove.")
                return
            
            if messagebox.askyesno("Confirm Delete", 
                f"Remove {len(selected)} selected waypoint(s)?"):
                # Remove in reverse order to maintain indices
                for idx in reversed(selected):
                    del self.waypoints[idx]
                
                # Refresh the listbox
                waypoint_listbox.delete(0, tk.END)
                total_dist = 0
                for i, (lat, lon) in enumerate(self.waypoints):
                    if i > 0:
                        dist = self._calculate_distance(
                            self.waypoints[i-1][0], self.waypoints[i-1][1],
                            lat, lon
                        )
                        total_dist += dist
                    
                    if i == 0:
                        label = f"[START]  Point {i+1:3d}  |  {lat:9.6f}, {lon:9.6f}  |  0.0 ft"
                    else:
                        label = f"         Point {i+1:3d}  |  {lat:9.6f}, {lon:9.6f}  |  {total_dist:,.0f} ft"
                    
                    waypoint_listbox.insert(tk.END, label)
                
                # Update map markers
                self._refresh_map_markers()
                
                messagebox.showinfo("Success", f"Removed {len(selected)} waypoint(s)")
                update_selection_info()
        
        ctk.CTkButton(
            quick_frame,
            text="🗑️ Remove Selected",
            width=130,
            height=28,
            command=remove_selected_points,
            fg_color="#cc0000",
            hover_color="#990000"
        ).pack(side=tk.LEFT, padx=2)
        
        # Selection info
        self.selection_info = ctk.CTkLabel(
            dialog,
            text="No selection",
            font=ctk.CTkFont(size=11),
            text_color="orange"
        )
        self.selection_info.pack(pady=(0, 15))
        
        def update_selection_info():
            selected = waypoint_listbox.curselection()
            if len(selected) < 2:
                self.selection_info.configure(
                    text="⚠️ Select at least 2 waypoints",
                    text_color="orange"
                )
            else:
                start_idx = selected[0]
                end_idx = selected[-1]
                num_points = end_idx - start_idx + 1
                self.selection_info.configure(
                    text=f"✓ Selected: Points {start_idx+1} to {end_idx+1} ({num_points} waypoints)",
                    text_color="#27AE60"
                )
        
        waypoint_listbox.bind('<<ListboxSelect>>', lambda e: update_selection_info())
        
        # Buttons
        button_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        button_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        def apply_selection():
            selected = waypoint_listbox.curselection()
            if len(selected) < 2:
                messagebox.showwarning("Invalid Selection",
                    "Please select at least 2 waypoints.\n\n"
                    "Click the start waypoint, then hold Ctrl and click the end waypoint.")
                return
            
            # Get continuous range from first to last selected
            start_idx = selected[0]
            end_idx = selected[-1]
            
            # Store selection range (don't modify waypoints)
            self.waypoint_selection = [start_idx, end_idx]
            
            # Update map markers to show selection
            self._refresh_map_markers()
            
            # Update display
            self._update_waypoint_display()
            
            dialog.destroy()
            
            selected_count = end_idx - start_idx + 1
            messagebox.showinfo("Range Selected",
                f"Selected waypoints {start_idx+1} to {end_idx+1}\n\n"
                f"Total waypoints in selection: {selected_count}\n\n"
                f"Click 'Generate Profile' to create elevation profile for this segment.\n\n"
                f"To change selection, click '✂️ Select Range' again.")
        
        ctk.CTkButton(
            button_frame,
            text="✓ Apply Selection",
            width=150,
            height=36,
            command=apply_selection,
            fg_color="#27AE60",
            hover_color="#229954",
            font=ctk.CTkFont(size=13, weight="bold")
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        def done_without_selection():
            # Update map markers and display
            self._refresh_map_markers()
            self._update_waypoint_display()
            dialog.destroy()
        
        ctk.CTkButton(
            button_frame,
            text="Done",
            width=100,
            height=36,
            command=done_without_selection,
            fg_color="#3498DB",
            hover_color="#2980B9"
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ctk.CTkButton(
            button_frame,
            text="Cancel",
            width=100,
            height=36,
            command=dialog.destroy,
            fg_color="gray40",
            hover_color="gray30"
        ).pack(side=tk.LEFT)
        
        # Auto-select all by default
        waypoint_listbox.selection_set(0, tk.END)
        update_selection_info()
    
    def _refresh_map_markers(self):
        """Refresh map markers after waypoint selection"""
        # Clear existing markers
        for marker in self.waypoint_markers:
            marker.delete()
        self.waypoint_markers.clear()
        
        # Clear path
        if self.path_line:
            self.path_line.delete()
            self.path_line = None
        
        # Determine which waypoints are selected
        if self.waypoint_selection:
            start_idx, end_idx = self.waypoint_selection
        else:
            start_idx, end_idx = 0, len(self.waypoints) - 1
        
        # Add new markers
        for i, (lat, lon) in enumerate(self.waypoints):
            # Check if this waypoint is in the selected range
            is_selected = start_idx <= i <= end_idx
            
            if i == start_idx and is_selected:
                color_circle = "#27AE60"
                color_outside = "#229954"
                text = "Start"
            elif i == end_idx and is_selected:
                color_circle = "#E74C3C"
                color_outside = "#C0392B"
                text = "End"
            elif is_selected:
                color_circle = "#42A5F5"
                color_outside = "#2196F3"
                text = f"Pt {i+1}"
            else:
                # Unselected waypoints shown in gray
                color_circle = "#95A5A6"
                color_outside = "#7F8C8D"
                text = f"Pt {i+1}"
            
            marker = self.map_widget.set_marker(
                lat, lon,
                text=text,
                marker_color_circle=color_circle,
                marker_color_outside=color_outside
            )
            self.waypoint_markers.append(marker)
        
        # Draw path connecting selected waypoints only
        if len(self.waypoints) > 1:
            if self.waypoint_selection:
                start_idx, end_idx = self.waypoint_selection
                path_coords = self.waypoints[start_idx:end_idx+1]
            else:
                path_coords = self.waypoints
            
            self.path_line = self.map_widget.set_path(
                path_coords,
                color="#42A5F5",
                width=3
            )
    
    def _update_waypoint_display(self):
        """Update the waypoint count display"""
        count = len(self.waypoints)
        if count == 0:
            text = "Waypoints: 0 points"
            color = "gray60"
        elif count == 1:
            text = f"Waypoints: 1 point (need 2+ for profile)"
            color = "orange"
        else:
            # Calculate total distance for all waypoints
            total_distance = 0
            for i in range(len(self.waypoints) - 1):
                dist_km = self._calculate_distance_km(self.waypoints[i], self.waypoints[i+1])
                total_distance += dist_km
            total_distance_ft = total_distance * 3280.84
            
            # Add selection info if a range is selected
            if self.waypoint_selection:
                start_idx, end_idx = self.waypoint_selection
                selected_count = end_idx - start_idx + 1
                
                # Calculate distance for selected waypoints only
                selected_distance = 0
                for i in range(start_idx, end_idx):
                    dist_km = self._calculate_distance_km(self.waypoints[i], self.waypoints[i+1])
                    selected_distance += dist_km
                selected_distance_ft = selected_distance * 3280.84
                
                text = f"Waypoints: {count} points ({total_distance_ft:,.0f} ft) | Selected: {selected_count} points ({selected_distance_ft:,.0f} ft)"
            else:
                text = f"Waypoints: {count} points ({total_distance_ft:,.0f} ft)"
            
            color = "#27AE60"
        
        self.waypoint_label.configure(text=text, text_color=color)
        
    def _search_location(self):
        """Search for location using Nominatim geocoding"""
        query = self.search_entry.get().strip()
        if not query:
            return
            
        logging.info(f"Searching for: {query}")
        self.progress_bar.set(0.3)
        
        # Check if it's coordinates (lat, lon)
        if ',' in query:
            try:
                parts = query.split(',')
                lat = float(parts[0].strip())
                lon = float(parts[1].strip())
                self.map_widget.set_position(lat, lon)
                self.map_widget.set_zoom(15)
                logging.info(f"Moved to coordinates: {lat:.6f}, {lon:.6f}")
                self.progress_bar.set(0)
                return
            except:
                pass
        
        # Otherwise, geocode the address
        def geocode():
            try:
                url = "https://nominatim.openstreetmap.org/search"
                params = {
                    'q': query,
                    'format': 'json',
                    'limit': 1
                }
                headers = {'User-Agent': 'ProfileAutomationTool/1.0'}
                
                response = requests.get(url, params=params, headers=headers, timeout=10)
                response.raise_for_status()
                results = response.json()
                
                if results:
                    lat = float(results[0]['lat'])
                    lon = float(results[0]['lon'])
                    self.after(0, lambda: self._on_geocode_success(lat, lon, results[0]['display_name']))
                else:
                    self.after(0, lambda: self._on_geocode_error("No results found"))
                    
            except Exception as e:
                self.after(0, lambda: self._on_geocode_error(str(e)))
                
        threading.Thread(target=geocode, daemon=True).start()
        
    def _on_geocode_success(self, lat, lon, display_name):
        """Handle successful geocoding"""
        self.map_widget.set_position(lat, lon)
        self.map_widget.set_zoom(15)
        logging.info(f"Found: {display_name}")
        logging.info(f"Coordinates: {lat:.6f}, {lon:.6f}")
        self.progress_bar.set(0)
        
    def _on_geocode_error(self, error):
        """Handle geocoding error"""
        logging.error(f"Search failed: {error}")
        self.progress_bar.set(0)
        messagebox.showerror("Search Error", f"Could not find location:\n{error}")
        
    def _cancel_generation(self):
        """Cancel the ongoing profile generation"""
        self.cancel_generation = True
        logging.info("⛔ Cancelling profile generation...")
        # Hide cancel button, show generate button
        self.cancel_generation_btn.pack_forget()
        self.generate_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.progress_bar.set(0)
        
    def _calculate_distance_km(self, point1, point2):
        """Calculate distance between two lat/lon points in kilometers"""
        lat1, lon1 = point1
        lat2, lon2 = point2
        
        # Haversine formula
        R = 6371  # Earth's radius in km
        
        lat1_rad = math.radians(lat1)
        lat2_rad = math.radians(lat2)
        delta_lat = math.radians(lat2 - lat1)
        delta_lon = math.radians(lon2 - lon1)
        
        a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        
        return R * c
    
    def _cancel_generation(self):
        """Cancel the ongoing profile generation"""
        self.cancel_generation = True
        logging.info("⛔ Cancelling profile generation...")
        # Hide cancel button, show generate button
        self.cancel_generation_btn.pack_forget()
        self.generate_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.progress_bar.set(0)
    
    def _get_active_waypoints(self):
        """Get the currently active waypoints based on selection"""
        if self.waypoint_selection:
            start_idx, end_idx = self.waypoint_selection
            return self.waypoints[start_idx:end_idx+1]
        else:
            return self.waypoints
        
    def _cancel_generation(self):
        """Cancel the ongoing profile generation"""
        self.cancel_generation = True
        logging.info("⛔ Cancelling profile generation...")
        # Hide cancel button, show generate button
        self.cancel_generation_btn.pack_forget()
        self.generate_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.progress_bar.set(0)
    
    def _generate_profile(self):
        """Generate elevation profile along waypoints"""
        if len(self.waypoints) < 2:
            messagebox.showwarning("Missing Waypoints", "Please add at least 2 waypoints on the map to create a profile.")
            return
        
        # Get active waypoints (based on selection)
        active_waypoints = self._get_active_waypoints()
        if len(active_waypoints) < 2:
            messagebox.showwarning("Invalid Selection", "Please select at least 2 waypoints for profile generation.")
            return
            
        # Update configuration from UI
        self.project_name = self.project_entry.get().strip() or "Elevation Profile"
        try:
            self.interval_ft = float(self.interval_entry.get().strip())
        except:
            self.interval_ft = 10
        
        # Reset cancel flag
        self.cancel_generation = False
        
        # Hide generate button, show cancel button
        self.generate_btn.pack_forget()
        self.cancel_generation_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.progress_bar.set(0.1)
        
        logging.info("="*60)
        logging.info("Starting elevation profile generation...")
        logging.info(f"Start: {active_waypoints[0][0]:.6f}, {active_waypoints[0][1]:.6f}")
        logging.info(f"End: {active_waypoints[-1][0]:.6f}, {active_waypoints[-1][1]:.6f}")
        if len(active_waypoints) > 2:
            logging.info(f"Waypoints: {len(active_waypoints)} points total")
        if self.waypoint_selection:
            logging.info(f"Selection: waypoints {self.waypoint_selection[0]+1} to {self.waypoint_selection[1]+1}")
        logging.info(f"Interval: {self.interval_ft} feet")
        logging.info(f"Data Source: {self.api_source}")
        
        # Run in background thread
        def generate():
            try:
                if self.cancel_generation:
                    return
                    
                if self.api_source == "Mock":
                    profile_data = self._generate_mock_profile_waypoints()
                elif self.api_source == "USGS":
                    profile_data = self._fetch_usgs_profile_waypoints()
                elif self.api_source == "ArcGIS":
                    profile_data = self._fetch_arcgis_profile_waypoints()
                else:
                    raise Exception(f"Unknown API source: {self.api_source}")
                
                # Check if cancelled before processing results
                if self.cancel_generation:
                    logging.info("Profile generation cancelled by user")
                    self.after(0, lambda: self.progress_bar.set(0))
                    return
                    
                self.after(0, lambda: self._on_profile_success(profile_data))
                
            except Exception as e:
                if not self.cancel_generation:
                    self.after(0, lambda: self._on_profile_error(str(e)))
                
        threading.Thread(target=generate, daemon=True).start()
        
    def _generate_mock_profile_waypoints(self):
        """Generate mock elevation data for waypoints"""
        waypoints = self._get_active_waypoints()
        logging.info(f"Generating mock elevation data for {len(waypoints)} waypoints...")
        self.after(0, lambda: self.progress_bar.set(0.3))
        
        profile_data = []
        cumulative_distance = 0
        
        # Process each segment between waypoints
        for segment_idx in range(len(waypoints) - 1):
            # Check if generation was cancelled
            if self.cancel_generation:
                logging.info("Mock profile generation cancelled by user")
                return []
            
            start_wp = waypoints[segment_idx]
            end_wp = waypoints[segment_idx + 1]
            
            # Calculate segment distance
            segment_dist_km = self._calculate_distance_km(start_wp, end_wp)
            segment_dist_ft = segment_dist_km * 3280.84
            
            # Generate points at exact interval spacing
            current_distance_in_segment = 0.0
            
            while current_distance_in_segment <= segment_dist_ft:
                # Calculate position along segment (0.0 to 1.0)
                progress = current_distance_in_segment / segment_dist_ft if segment_dist_ft > 0 else 0
                
                # Linear interpolation for lat/lon
                lat = start_wp[0] + (end_wp[0] - start_wp[0]) * progress
                lon = start_wp[1] + (end_wp[1] - start_wp[1]) * progress
                
                # Mock elevation (sine wave + trend + segment variation)
                total_dist = cumulative_distance + current_distance_in_segment
                elevation = 1000 + 200 * math.sin(total_dist / 500 * math.pi) + total_dist / 50 + segment_idx * 10
                
                profile_data.append({
                    'distance': total_dist,
                    'elevation': elevation,
                    'x': lon,
                    'y': lat,
                    'segment': segment_idx
                })
                
                # Move to next interval point
                current_distance_in_segment += self.interval_ft
                
                # If we've passed the end, add the endpoint if not already added
                if current_distance_in_segment > segment_dist_ft and progress < 1.0:
                    profile_data.append({
                        'distance': cumulative_distance + segment_dist_ft,
                        'elevation': 1000 + 200 * math.sin((cumulative_distance + segment_dist_ft) / 500 * math.pi) + (cumulative_distance + segment_dist_ft) / 50 + segment_idx * 10,
                        'x': end_wp[1],
                        'y': end_wp[0],
                        'segment': segment_idx
                    })
                    break
            
            cumulative_distance += segment_dist_ft
            self.after(0, lambda p=0.3 + (segment_idx + 1) / len(self.waypoints) * 0.5: self.progress_bar.set(p))
        
        # Add slopes
        profile_data = self._calculate_slopes(profile_data)
        
        logging.info(f"Generated {len(profile_data)} mock data points across {len(self.waypoints)-1} segments")
        return profile_data
    
    def _generate_mock_profile(self):
        """Generate mock elevation data for testing (legacy 2-point)"""
        logging.info("Generating mock elevation data...")
        self.after(0, lambda: self.progress_bar.set(0.3))
        
        # Calculate total distance
        if len(self.waypoints) >= 2:
            distance_km = self._calculate_distance_km(self.waypoints[0], self.waypoints[1])
        else:
            distance_km = 0
        distance_ft = distance_km * 3280.84
        
        # Generate points
        num_points = max(10, int(distance_ft / self.interval_ft))
        
        profile_data = []
        for i in range(num_points):
            progress = i / num_points
            
            # Linear interpolation for lat/lon
            if len(self.waypoints) >= 2:
                lat = self.waypoints[0][0] + (self.waypoints[1][0] - self.waypoints[0][0]) * progress
                lon = self.waypoints[0][1] + (self.waypoints[1][1] - self.waypoints[0][1]) * progress
            else:
                lat = 0
                lon = 0
            
            # Mock elevation (sine wave + trend)
            dist = progress * distance_ft
            elevation = 1000 + 200 * math.sin(progress * math.pi * 2) + progress * 100
            
            profile_data.append({
                'distance': dist,
                'elevation': elevation,
                'x': lon,
                'y': lat
            })
            
            if i % 20 == 0:
                self.after(0, lambda p=0.3 + progress * 0.6: self.progress_bar.set(p))
        
        logging.info(f"Generated {len(profile_data)} mock data points")
        return profile_data
        
    def _calculate_distance(self, lat1, lon1, lat2, lon2):
        """Calculate distance between two points using Haversine formula (returns feet)"""
        from math import radians, sin, cos, sqrt, atan2
        
        # Earth radius in feet
        R = 20902231  # feet
        
        # Convert to radians
        lat1_rad = radians(lat1)
        lat2_rad = radians(lat2)
        dLat = radians(lat2 - lat1)
        dLon = radians(lon2 - lon1)
        
        # Haversine formula
        a = sin(dLat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(dLon / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))
        distance = R * c
        
        return distance
    
    def _calculate_slopes(self, profile_data):
        """Calculate slope/grade between points"""
        for i in range(len(profile_data)):
            if i == 0:
                profile_data[i]['slope'] = 0.0
            else:
                prev = profile_data[i-1]
                curr = profile_data[i]
                
                # Calculate slope as percentage
                horizontal_dist = curr['distance'] - prev['distance']
                vertical_change = curr['elevation'] - prev['elevation']
                
                if horizontal_dist > 0:
                    slope_percent = (vertical_change / horizontal_dist) * 100
                    profile_data[i]['slope'] = slope_percent
                else:
                    profile_data[i]['slope'] = 0.0
        
        return profile_data
    
    def _fetch_usgs_profile_waypoints(self):
        """Fetch elevation profile from USGS Elevation Point Query Service for waypoints"""
        waypoints = self._get_active_waypoints()
        logging.info(f"Fetching elevation data for {len(waypoints)} waypoints from USGS EPQS...")
        self.after(0, lambda: self.progress_bar.set(0.2))
        
        # USGS EPQS v1 API endpoint (correct URL with /v1/json)
        usgs_url = "https://epqs.nationalmap.gov/v1/json"
        
        # Generate interpolated points along all segments at exact intervals
        all_points = []
        total_distance = 0.0
        
        for i in range(len(waypoints) - 1):
            start_lat, start_lon = waypoints[i]
            end_lat, end_lon = waypoints[i + 1]
            
            # Calculate segment distance using Haversine formula
            segment_distance_ft = self._calculate_distance(start_lat, start_lon, end_lat, end_lon)
            
            # Generate points at exact interval spacing
            # Start at 0 (or total_distance if not first segment)
            current_distance_in_segment = 0.0
            
            while current_distance_in_segment <= segment_distance_ft:
                # Calculate position along segment (0.0 to 1.0)
                t = current_distance_in_segment / segment_distance_ft if segment_distance_ft > 0 else 0
                
                # Interpolate lat/lon
                lat = start_lat + t * (end_lat - start_lat)
                lon = start_lon + t * (end_lon - start_lon)
                
                # Calculate absolute distance from start
                distance = total_distance + current_distance_in_segment
                
                all_points.append({
                    'lat': lat,
                    'lon': lon,
                    'distance': distance
                })
                
                # Move to next interval point
                current_distance_in_segment += self.interval_ft
                
                # If we've passed the end, add the endpoint if not already added
                if current_distance_in_segment > segment_distance_ft and t < 1.0:
                    all_points.append({
                        'lat': end_lat,
                        'lon': end_lon,
                        'distance': total_distance + segment_distance_ft
                    })
                    break
            
            total_distance += segment_distance_ft
        
        logging.info(f"Querying USGS EPQS for {len(all_points)} points...")
        self.after(0, lambda: self.progress_bar.set(0.4))
        
        # Query USGS EPQS for each point (batch requests)
        profile_data = []
        batch_size = 10  # Process in batches for better progress reporting
        
        # Retry configuration
        max_retries = 5  # 6 total attempts (initial + 5 retries)
        
        for idx, point in enumerate(all_points):
            # Check if generation was cancelled
            if self.cancel_generation:
                logging.info("USGS profile fetch cancelled by user")
                return []
            
            # Add small delay to avoid rate limiting (USGS recommendation)
            if idx > 0:
                time.sleep(0.05)  # 50ms delay between requests
            
            try:
                # Retry logic for intermittent API failures
                elevation = None
                last_error = None
                
                for retry in range(max_retries + 1):
                    try:
                        params = {
                            'x': point['lon'],
                            'y': point['lat'],
                            'units': 'Feet'
                        }
                        
                        response = requests.get(usgs_url, params=params, timeout=15)
                        
                        # Check status code before parsing
                        if response.status_code != 200:
                            raise Exception(f"HTTP {response.status_code}")
                        
                        # Parse JSON response
                        try:
                            result = response.json()
                        except ValueError as ve:
                            # API sometimes returns plain text errors instead of JSON
                            error_msg = response.text[:100] if response.text else "Empty response"
                            raise Exception(f"Non-JSON response: {error_msg}")
                        
                        # Extract elevation from USGS v1 API response
                        elevation_value = result.get('value', None)
                        resolution = result.get('resolution', None)
                        
                        if elevation_value is None:
                            raise Exception("No 'value' field in response")
                        
                        # Convert to float (API may return string)
                        elevation = float(elevation_value)
                        
                        # USGS returns -1000000 or similar for no data
                        if elevation < -1000:
                            raise Exception("No data available (-1000000)")
                        
                        # Store resolution for logging (only log first point)
                        if idx == 0 and resolution is not None:
                            resolution_m = float(resolution)
                            resolution_ft = resolution_m * 3.28084
                            logging.info(f"Data resolution: {resolution_m:.2f}m ({resolution_ft:.2f} feet) - USGS 3DEP")
                        
                        # Success! Break out of retry loop
                        break
                        
                    except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                        last_error = e
                        if retry < max_retries:
                            # Network error - retry with backoff
                            wait_time = 1.0 * (2 ** retry)  # Longer wait for network issues
                            logging.warning(f"Point {idx + 1}: Network error (attempt {retry + 1}/{max_retries + 1}), retrying in {wait_time:.1f}s...")
                            time.sleep(wait_time)
                            continue
                        else:
                            # All retries exhausted
                            raise Exception(f"Network timeout after {max_retries + 1} attempts: {str(e)}")
                    
                    except Exception as e:
                        last_error = e
                        if retry < max_retries:
                            # API error - retry with shorter backoff
                            wait_time = 0.5 * (2 ** retry)
                            logging.warning(f"Point {idx + 1}: API error (attempt {retry + 1}/{max_retries + 1}), retrying in {wait_time:.1f}s...")
                            time.sleep(wait_time)
                            continue
                        else:
                            # All retries exhausted
                            raise e
                
                # Store the point (successful elevation retrieved)
                profile_data.append({
                    'distance': point['distance'],
                    'elevation': float(elevation),
                    'x': point['lon'],
                    'y': point['lat']
                })
                
            except Exception as e:
                # NEVER interpolate - fail the entire operation if API doesn't respond
                error_type = "Network timeout" if "timeout" in str(e).lower() else "API error"
                error_msg = f"Failed to retrieve elevation data from USGS API for point {idx + 1}/{len(all_points)}. {error_type} after {max_retries + 1} attempts. Cannot continue without valid API data."
                logging.error(error_msg)
                raise Exception(error_msg)
            
            # Update progress
            if (idx + 1) % batch_size == 0 or idx == len(all_points) - 1:
                progress = 0.4 + (0.5 * (idx + 1) / len(all_points))
                self.after(0, lambda p=progress: self.progress_bar.set(p))
                logging.info(f"Progress: {idx + 1}/{len(all_points)} points retrieved")
        
        if not profile_data:
            raise Exception("Failed to retrieve elevation data from USGS EPQS v1 API")
        
        # Add slopes
        profile_data = self._calculate_slopes(profile_data)
        
        logging.info(f"Successfully retrieved {len(profile_data)} elevation points from USGS EPQS v1 API")
        return profile_data
    
    def _fetch_arcgis_profile_waypoints(self):
        """Fetch elevation profile from ArcGIS service for waypoints"""
        waypoints = self._get_active_waypoints()
        logging.info(f"Fetching elevation data for {len(waypoints)} waypoints from ArcGIS service...")
        self.after(0, lambda: self.progress_bar.set(0.3))
        
        # Build multi-segment geometry
        path_coords = [[wp[1], wp[0]] for wp in waypoints]  # [lon, lat] format
        
        geometry = {
            'paths': [path_coords],
            'spatialReference': {'wkid': 4326}
        }
        
        # Build request parameters
        params = {
            'f': 'json',
            'InputLineFeatures': json.dumps(geometry),
            'returnZ': True,
            'returnM': True,
            'ProfileIDField': '',
            'DEMResolution': 'FINEST',
            'MaximumSampleDistance': self.interval_ft,
            'MaximumSampleDistanceUnits': 'Feet'
        }
        
        if self.api_token:
            params['token'] = self.api_token
            
        logging.info("Sending request to ArcGIS server...")
        
        try:
            response = requests.post(self.service_url, data=params, timeout=60)
            response.raise_for_status()
            result = response.json()
            
            self.after(0, lambda: self.progress_bar.set(0.7))
            
            if 'error' in result:
                raise Exception(f"ArcGIS Error: {result['error'].get('message', 'Unknown error')}")
                
            if 'results' not in result or not result['results']:
                raise Exception("No results returned from ArcGIS service")
                
            # Parse profile data
            profile_result = result['results'][0]
            if 'value' not in profile_result or 'features' not in profile_result['value']:
                raise Exception("Invalid response format from ArcGIS")
                
            features = profile_result['value']['features']
            
            profile_data = []
            for feature in features:
                geom = feature['geometry']
                
                if 'paths' in geom and geom['paths']:
                    for point in geom['paths'][0]:
                        lon, lat = point[0], point[1]
                        elevation = point[2] if len(point) > 2 else 0
                        distance = point[3] if len(point) > 3 else 0
                        
                        profile_data.append({
                            'distance': distance,
                            'elevation': elevation,
                            'x': lon,
                            'y': lat
                        })
            
            if not profile_data:
                raise Exception("No elevation points found in response")
            
            # Add slopes
            profile_data = self._calculate_slopes(profile_data)
            
            logging.info(f"Successfully retrieved {len(profile_data)} elevation points")
            return profile_data
            
        except requests.exceptions.Timeout:
            raise Exception("Request timed out. The service may be unavailable.")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Network error: {str(e)}")
    
    def _fetch_arcgis_profile(self):
        """Fetch elevation profile from ArcGIS service (legacy 2-point)"""
        logging.info("Fetching elevation data from ArcGIS service...")
        self.after(0, lambda: self.progress_bar.set(0.3))
        
        # Build geometry
        if len(self.waypoints) >= 2:
            geometry = {
                'paths': [[
                    [self.waypoints[0][1], self.waypoints[0][0]],
                    [self.waypoints[1][1], self.waypoints[1][0]]
                ]],
                'spatialReference': {'wkid': 4326}
            }
        else:
            raise Exception("Need at least 2 waypoints")
        geometry = {
            'paths': [[
                [self.waypoints[0][1], self.waypoints[0][0]],
                [self.waypoints[1][1], self.waypoints[1][0]]
            ]],
            'spatialReference': {'wkid': 4326}
        }
        
        # Build request parameters
        params = {
            'f': 'json',
            'InputLineFeatures': json.dumps(geometry),
            'returnZ': True,
            'returnM': True,
            'ProfileIDField': '',
            'DEMResolution': 'FINEST',
            'MaximumSampleDistance': self.interval_ft,
            'MaximumSampleDistanceUnits': 'Feet'
        }
        
        if self.api_token:
            params['token'] = self.api_token
            
        logging.info("Sending request to ArcGIS server...")
        
        try:
            response = requests.post(self.service_url, data=params, timeout=60)
            response.raise_for_status()
            result = response.json()
            
            self.after(0, lambda: self.progress_bar.set(0.7))
            
            if 'error' in result:
                raise Exception(f"ArcGIS Error: {result['error'].get('message', 'Unknown error')}")
                
            if 'results' not in result or not result['results']:
                raise Exception("No results returned from ArcGIS service")
                
            # Parse profile data
            profile_result = result['results'][0]
            if 'value' not in profile_result or 'features' not in profile_result['value']:
                raise Exception("Invalid response format from ArcGIS")
                
            features = profile_result['value']['features']
            
            profile_data = []
            for feature in features:
                geom = feature['geometry']
                attrs = feature.get('attributes', {})
                
                # Extract coordinates and elevation
                if 'paths' in geom and geom['paths']:
                    for point in geom['paths'][0]:
                        lon, lat = point[0], point[1]
                        elevation = point[2] if len(point) > 2 else 0
                        distance = point[3] if len(point) > 3 else 0
                        
                        profile_data.append({
                            'distance': distance,
                            'elevation': elevation,
                            'x': lon,
                            'y': lat
                        })
            
            if not profile_data:
                raise Exception("No elevation points found in response")
                
            logging.info(f"Successfully retrieved {len(profile_data)} elevation points")
            return profile_data
            
        except requests.exceptions.Timeout:
            raise Exception("Request timed out. The service may be unavailable.")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Network error: {str(e)}")
            
    def _on_profile_success(self, profile_data):
        """Handle successful profile generation"""
        try:
            self.profile_data = profile_data
            self.progress_bar.set(0.9)
            
            logging.info(f"Profile generated successfully with {len(profile_data)} points")
            
            # Use current running line depth (set in config window or default)
            # No need to read from UI - depth is already stored in self.running_line_depth_inches
                
            # Calculate running line with current settings
            if self.enable_running_line:
                # Simple calculation: running line = ground elevation - depth
                depth_ft = self.running_line_depth_inches / 12.0
                for point in profile_data:
                    point['running_line'] = point['elevation'] - depth_ft
                    point['cut_fill'] = depth_ft
                    point['cover'] = depth_ft
                    point['depth_inches'] = self.running_line_depth_inches
                    point['segment_id'] = 0
                logging.info(f"Running line calculated with equal depth: {self.running_line_depth_inches:.0f} inches")
            
            # Calculate statistics
            elevations = [p['elevation'] for p in profile_data]
            min_elev = min(elevations)
            max_elev = max(elevations)
            total_distance = profile_data[-1]['distance']
            
            logging.info(f"Distance: {total_distance:,.0f} feet ({total_distance/5280:.2f} miles)")
            logging.info(f"Ground Elevation: {min_elev:,.1f} - {max_elev:,.1f} feet")
            
            if self.enable_running_line:
                running_elevations = [p['running_line'] for p in profile_data]
                min_rl = min(running_elevations)
                max_rl = max(running_elevations)
                logging.info(f"Running Line: {min_rl:,.1f} - {max_rl:,.1f} feet")
                if self.depth_zones:
                    logging.info(f"Variable Depth: {len(self.depth_zones)} override zone(s)")
                else:
                    logging.info(f"Installation Depth: {self.running_line_depth_inches:.0f} inches ({self.running_line_depth_inches/12:.2f} feet)")
            
            # Log data quality information
            if self.api_source == "USGS":
                logging.info(f"Data Quality: 1m resolution USGS 3DEP (±0.3-1m vertical accuracy)")
            
            # Draw chart
            self._draw_profile_chart()
            
            # Hide cancel button, show generate button
            self.cancel_generation_btn.pack_forget()
            self.generate_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
            self.progress_bar.set(1.0)
            
            logging.info("="*60)
            
        except Exception as e:
            logging.error(f"Error in _on_profile_success: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            self._on_profile_error(f"Post-processing error: {str(e)}")
        
    def _on_profile_error(self, error):
        """Handle profile generation error"""
        logging.error(f"Failed to generate profile: {error}")
        self.progress_bar.set(0)
        # Hide cancel button, show generate button
        self.cancel_generation_btn.pack_forget()
        self.generate_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        messagebox.showerror("Profile Generation Error", f"Failed to generate elevation profile:\n\n{error}")
        
    def _catmull_rom_spline(self, points, num_samples=50):
        """Generate smooth curve through points using Catmull-Rom spline
        
        Args:
            points: List of (x, y) tuples
            num_samples: Number of samples per segment
            
        Returns:
            Two lists: smoothed_x, smoothed_y
        """
        if len(points) < 2:
            return [], []
        
        if len(points) == 2:
            # Just linear for 2 points
            return [p[0] for p in points], [p[1] for p in points]
        
        smoothed_x = []
        smoothed_y = []
        
        # Add first point
        smoothed_x.append(points[0][0])
        smoothed_y.append(points[0][1])
        
        # Process each segment
        for i in range(len(points) - 1):
            # Get control points (with virtual endpoints)
            p0 = points[max(0, i - 1)]
            p1 = points[i]
            p2 = points[i + 1]
            p3 = points[min(len(points) - 1, i + 2)]
            
            # If at boundaries, use end points twice for better behavior
            if i == 0:
                p0 = p1
            if i == len(points) - 2:
                p3 = p2
            
            # Generate samples for this segment
            for t_step in range(1, num_samples + 1):
                t = t_step / num_samples
                t2 = t * t
                t3 = t2 * t
                
                # Catmull-Rom basis functions
                x = 0.5 * (
                    (2 * p1[0]) +
                    (-p0[0] + p2[0]) * t +
                    (2*p0[0] - 5*p1[0] + 4*p2[0] - p3[0]) * t2 +
                    (-p0[0] + 3*p1[0] - 3*p2[0] + p3[0]) * t3
                )
                
                y = 0.5 * (
                    (2 * p1[1]) +
                    (-p0[1] + p2[1]) * t +
                    (2*p0[1] - 5*p1[1] + 4*p2[1] - p3[1]) * t2 +
                    (-p0[1] + 3*p1[1] - 3*p2[1] + p3[1]) * t3
                )
                
                smoothed_x.append(x)
                smoothed_y.append(y)
        
        return smoothed_x, smoothed_y
    
    def _draw_profile_chart(self):
        """Draw elevation profile chart (2D or 3D)"""
        if not self.profile_data:
            return
            
        # Clear existing chart
        for widget in self.chart_container.winfo_children():
            widget.destroy()
        
        if self.view_3d:
            self._draw_3d_view()
        else:
            self._draw_2d_profile()
            
    def _draw_2d_profile(self):
        """Draw 2D elevation profile with slope annotations"""
        # Create figure
        self.chart_figure = Figure(figsize=(10, 5), dpi=100, facecolor='white')
        self.chart_ax = self.chart_figure.add_subplot(111)
        
        # Extract data
        distances = [p['distance'] for p in self.profile_data]
        elevations = [p['elevation'] for p in self.profile_data]
        slopes = [p.get('slope', 0) for p in self.profile_data]
        
        # Plot Ground Elevation (Green)
        self.chart_ax.plot(distances, elevations, color='#2d7a2d', linewidth=2, 
                          label='Ground Elevation (feet)', zorder=2)
        
        # Plot Running Line if enabled (Red)
        if self.enable_running_line and 'running_line' in self.profile_data[0]:
            running_line = [p['running_line'] for p in self.profile_data]
            
            # Apply curve smoothing if enabled and break points exist
            if self.curve_enabled and self.break_points:
                # Collect key points (start, break points, end)
                key_points = [(distances[0], running_line[0])]
                
                for bp_dist, bp_elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
                    # Find running line elevation at break point
                    for p in self.profile_data:
                        if abs(p['distance'] - bp_dist) < 1.0:
                            key_points.append((bp_dist, p['running_line']))
                            break
                
                key_points.append((distances[-1], running_line[-1]))
                
                # Generate smooth curve through key points
                smooth_x, smooth_y = self._catmull_rom_spline(key_points, num_samples=30)
                
                # Plot the smooth curve
                self.chart_ax.plot(smooth_x, smooth_y, color='#cc0000', linewidth=2,
                                 label=f'Running Line (Curved, {self.running_line_depth_inches}")', zorder=2)
            else:
                # Draw running line as single red line (configured design)
                self.chart_ax.plot(distances, running_line, color='#cc0000', linewidth=2,
                                 label=f'Running Line ({self.running_line_depth_inches}")', zorder=2)
            
            # Draw break point markers
            if self.break_points:
                for bp_dist, bp_elev, grade, locked, note in self.break_points:
                    # Find running line elevation at break point
                    bp_running = None
                    for p in self.profile_data:
                        if abs(p['distance'] - bp_dist) < 1.0:
                            bp_running = p['running_line']
                            break
                    
                    if bp_running is not None:
                        # Draw vertical line at break point
                        self.chart_ax.axvline(x=bp_dist, color='#9900cc', linestyle='--', 
                                            linewidth=1.5, alpha=0.7, zorder=5)
                        
                        # Draw marker on running line
                        self.chart_ax.scatter([bp_dist], [bp_running], color='#9900cc', 
                                            s=150, zorder=11, marker='D', edgecolors='white', 
                                            linewidths=2, label='Break Point' if bp_dist == self.break_points[0][0] else None)
                        
                        # Add label
                        label_text = f"BP\n{bp_dist:.0f}'"
                        if note:
                            label_text += f"\n{note}"
                        self.chart_ax.annotate(
                            label_text,
                            xy=(bp_dist, bp_running),
                            xytext=(0, 15), textcoords='offset points',
                            fontsize=8, color='#9900cc', ha='center',
                            bbox=dict(boxstyle='round,pad=0.4', facecolor='white', 
                                    edgecolor='#9900cc', alpha=0.9, linewidth=1.5),
                            zorder=12
                        )
            
            # Show pending zone start marker (only during editing)
            if self.depth_edit_mode and self.pending_zone_start is not None:
                # Find elevation at pending start
                for p in self.profile_data:
                    if p['distance'] >= self.pending_zone_start:
                        start_elev = p['elevation']
                        self.chart_ax.axvline(x=self.pending_zone_start, color='#E67E22', 
                                            linestyle='-', linewidth=2, zorder=10)
                        self.chart_ax.scatter([self.pending_zone_start], [start_elev],
                                            color='#E67E22', s=100, zorder=10,
                                            marker='o', edgecolors='white', linewidths=2)
                        self.chart_ax.annotate(
                            'ZONE START',
                            xy=(self.pending_zone_start, start_elev),
                            xytext=(5, 10), textcoords='offset points',
                            fontsize=9, color='#E67E22', weight='bold'
                        )
                        break
        
        # Add slope annotations at waypoint segments
        if len(self.waypoints) > 2:
            # Calculate average slope for each segment
            segment_slopes = {}
            for point in self.profile_data:
                seg = point.get('segment', 0)
                if seg not in segment_slopes:
                    segment_slopes[seg] = []
                segment_slopes[seg].append(point.get('slope', 0))
            
            # Annotate significant slopes
            for seg_idx, slopes_list in segment_slopes.items():
                if slopes_list:
                    avg_slope = sum(slopes_list) / len(slopes_list)
                    # Find mid-point of segment for annotation
                    seg_points = [p for p in self.profile_data if p.get('segment') == seg_idx]
                    if seg_points:
                        mid_idx = len(seg_points) // 2
                        mid_point = seg_points[mid_idx]
                        
                        # Color based on slope severity
                        if abs(avg_slope) < 2:
                            color = 'green'
                        elif abs(avg_slope) < 5:
                            color = 'orange'
                        else:
                            color = 'red'
                        
                        # Add annotation
                        self.chart_ax.annotate(
                            f'{avg_slope:+.1f}%',
                            xy=(mid_point['distance'], mid_point['elevation']),
                            xytext=(0, 20),
                            textcoords='offset points',
                            fontsize=8,
                            color=color,
                            ha='center',
                            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor=color, alpha=0.8)
                        )
        
        # Styling
        self.chart_ax.set_facecolor('white')
        self.chart_ax.set_xlabel('Distance (ft)', color='black', fontsize=11, fontweight='normal')
        self.chart_ax.set_ylabel('Elevation (ft)', color='black', fontsize=11, fontweight='normal')
        
        # Add title
        title_text = 'Elevation Profile'
        if hasattr(self, 'project_name') and self.project_name and self.project_name != 'Elevation Profile':
            title_text = f'Elevation Profile - {self.project_name}'
        self.chart_ax.set_title(title_text, color='black', fontsize=12, fontweight='normal', pad=10)
        
        # Add station markers if enabled
        if self.enable_stations:
            try:
                station_interval = float(self.station_entry.get())
                self.station_interval_ft = station_interval
            except:
                station_interval = 100
                
            if station_interval > 0:
                max_distance = max(distances)
                num_stations = int(max_distance / station_interval) + 1
                
                for i in range(num_stations):
                    station_dist = i * station_interval
                    if station_dist <= max_distance:
                        # Draw vertical line only (no labels)
                        self.chart_ax.axvline(x=station_dist, color='#999999', linestyle=':', 
                                             linewidth=1, alpha=0.6, zorder=1)
        
        # Grid
        self.chart_ax.grid(True, linestyle='-', alpha=0.3, color='#cccccc', linewidth=0.5)
        self.chart_ax.tick_params(colors='black', labelsize=9)
        
        # Format axes
        self.chart_ax.ticklabel_format(style='plain', axis='both')
        
        # Add legend
        self.chart_ax.legend(loc='best', frameon=True, fancybox=False, shadow=False,
                           fontsize=9, edgecolor='black')
        
        # Set spine colors
        for spine in self.chart_ax.spines.values():
            spine.set_edgecolor('black')
            spine.set_linewidth(1)
            
        self.chart_figure.tight_layout()
        
        # Embed in tkinter
        self.chart_canvas = FigureCanvasTkAgg(self.chart_figure, self.chart_container)
        self.chart_canvas.draw()
        self.chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Connect click event for depth editing
        self.chart_canvas.mpl_connect('button_press_event', self._on_chart_click)
        
    def _draw_3d_view(self):
        """Draw 3D visualization of terrain"""
        # Create figure with 3D axis
        self.chart_figure = Figure(figsize=(10, 6), dpi=100, facecolor='white')
        self.chart_ax = self.chart_figure.add_subplot(111, projection='3d')
        
        # Extract data
        distances = [p['distance'] for p in self.profile_data]
        elevations = [p['elevation'] for p in self.profile_data]
        lats = [p['y'] for p in self.profile_data]
        lons = [p['x'] for p in self.profile_data]
        
        # Create 3D line plot for ground
        self.chart_ax.plot(lons, lats, elevations, color='#2d7a2d', linewidth=2, 
                          label='Ground Elevation', zorder=2)
        
        # Plot Running Line if enabled
        if self.enable_running_line and 'running_line' in self.profile_data[0]:
            running_line = [p['running_line'] for p in self.profile_data]
            
            # Apply curve smoothing if enabled and break points exist
            if self.curve_enabled and self.break_points:
                logging.info(f"3D View: Applying curve smoothing with {len(self.break_points)} break points")
                
                # Collect key points for elevation curve
                key_points_elev = [(distances[0], running_line[0])]
                key_points_lon = [(distances[0], lons[0])]
                key_points_lat = [(distances[0], lats[0])]
                
                for bp_dist, bp_elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
                    # Find data at break point
                    for p in self.profile_data:
                        if abs(p['distance'] - bp_dist) < 1.0:
                            key_points_elev.append((bp_dist, p['running_line']))
                            key_points_lon.append((bp_dist, p['x']))
                            key_points_lat.append((bp_dist, p['y']))
                            break
                
                key_points_elev.append((distances[-1], running_line[-1]))
                key_points_lon.append((distances[-1], lons[-1]))
                key_points_lat.append((distances[-1], lats[-1]))
                
                # Generate smooth curves for all three dimensions
                smooth_distances, smooth_elevations = self._catmull_rom_spline(key_points_elev, num_samples=50)
                _, smooth_lons = self._catmull_rom_spline(key_points_lon, num_samples=50)
                _, smooth_lats = self._catmull_rom_spline(key_points_lat, num_samples=50)
                
                # Plot smooth 3D curve
                self.chart_ax.plot(smooth_lons, smooth_lats, smooth_elevations, color='#cc0000', linewidth=2,
                                 label='Running Line (Curved)', zorder=2)
            else:
                # Plot linear running line
                self.chart_ax.plot(lons, lats, running_line, color='#cc0000', linewidth=2,
                                 label='Running Line', zorder=2)
        
        # Plot waypoint markers (only active/selected waypoints)
        active_waypoints = self._get_active_waypoints()
        waypoint_lons = [wp[1] for wp in active_waypoints]
        waypoint_lats = [wp[0] for wp in active_waypoints]
        waypoint_elevs = []
        for wp in active_waypoints:
            # Find closest elevation point
            closest_point = min(self.profile_data, 
                              key=lambda p: abs(p['y'] - wp[0]) + abs(p['x'] - wp[1]))
            waypoint_elevs.append(closest_point['elevation'])
        
        self.chart_ax.scatter(waypoint_lons, waypoint_lats, waypoint_elevs,
                            c='blue', marker='o', s=100, label='Waypoints', zorder=10)
        
        # Labels
        self.chart_ax.set_xlabel('Longitude', fontsize=10)
        self.chart_ax.set_ylabel('Latitude', fontsize=10)
        self.chart_ax.set_zlabel('Elevation (ft)', fontsize=10)
        
        # Format lat/lon tick labels to show actual coordinates
        from matplotlib.ticker import FuncFormatter
        
        def lon_formatter(x, pos):
            return f'{x:.5f}'
        
        def lat_formatter(y, pos):
            return f'{y:.5f}'
        
        self.chart_ax.xaxis.set_major_formatter(FuncFormatter(lon_formatter))
        self.chart_ax.yaxis.set_major_formatter(FuncFormatter(lat_formatter))
        
        # Reduce number of ticks for cleaner display
        self.chart_ax.locator_params(axis='x', nbins=5)
        self.chart_ax.locator_params(axis='y', nbins=5)
        
        # Title
        title_text = '3D Elevation View'
        if hasattr(self, 'project_name') and self.project_name:
            title_text = f'3D View - {self.project_name}'
        self.chart_ax.set_title(title_text, fontsize=12, fontweight='normal', pad=10)
        
        # Legend
        self.chart_ax.legend(loc='upper left', fontsize=9)
        
        # Set viewing angle
        self.chart_ax.view_init(elev=25, azim=45)
        
        self.chart_figure.tight_layout()
        
        # Embed in tkinter
        self.chart_canvas = FigureCanvasTkAgg(self.chart_figure, self.chart_container)
        self.chart_canvas.draw()
        self.chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
    def _export_chart(self, format_type):
        """Export chart to file with custom dimensions"""
        if not self.profile_data or not self.chart_figure:
            messagebox.showwarning("No Data", "Please generate an elevation profile first.")
            return
        
        # Show export settings dialog
        settings = self._show_export_settings_dialog(format_type)
        if not settings:
            return  # User cancelled
            
        # File dialog
        extensions = {
            'png': [('PNG Image', '*.png')],
            'pdf': [('PDF Document', '*.pdf')],
            'svg': [('SVG Vector', '*.svg')]
        }
        
        filename = filedialog.asksaveasfilename(
            defaultextension=f'.{format_type}',
            filetypes=extensions[format_type],
            initialfile=f'elevation_profile_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format_type}'
        )
        
        if filename:
            try:
                # Create a new figure with custom dimensions
                custom_fig = Figure(figsize=(settings['width'], settings['height']), dpi=100, facecolor='white')
                custom_ax = custom_fig.add_subplot(111)
                
                # Extract data
                distances = [p['distance'] for p in self.profile_data]
                elevations = [p['elevation'] for p in self.profile_data]
                
                # Plot Ground Elevation (Green)
                custom_ax.plot(distances, elevations, color='#2d7a2d', linewidth=2, 
                              label='Ground Elevation (feet)', zorder=2)
                
                # Plot Running Line if enabled (Red)
                if self.enable_running_line and 'running_line' in self.profile_data[0]:
                    running_line = [p['running_line'] for p in self.profile_data]
                    
                    # Apply curve smoothing if enabled and break points exist
                    if self.curve_enabled and self.break_points:
                        logging.info(f"Export: Applying curve smoothing with {len(self.break_points)} break points")
                        # Collect key points (start, break points, end)
                        key_points = [(distances[0], running_line[0])]
                        
                        for bp_dist, bp_elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
                            # Find running line elevation at break point
                            for p in self.profile_data:
                                if abs(p['distance'] - bp_dist) < 1.0:
                                    key_points.append((bp_dist, p['running_line']))
                                    break
                        
                        key_points.append((distances[-1], running_line[-1]))
                        logging.info(f"Export: Generated {len(key_points)} key points for curve")
                        
                        # Generate smooth curve through key points
                        smooth_x, smooth_y = self._catmull_rom_spline(key_points, num_samples=30)
                        logging.info(f"Export: Curve generated with {len(smooth_x)} smooth points")
                        
                        # Plot the smooth curve
                        custom_ax.plot(smooth_x, smooth_y, color='#cc0000', linewidth=2,
                                     label=f'Running Line (Curved, {self.running_line_depth_inches}")', zorder=2)
                    else:
                        logging.info(f"Export: Drawing linear running line (curve_enabled={self.curve_enabled}, break_points={len(self.break_points) if self.break_points else 0})")
                        # Draw running line in red (same color regardless of depth zones)
                        custom_ax.plot(distances, running_line, color='#cc0000', linewidth=2,
                                     label=f'Running Line ({self.running_line_depth_inches}")', zorder=2)
                
                # Add Station Markers if enabled
                if hasattr(self, 'enable_station_markers') and self.enable_station_markers:
                    try:
                        station_interval = float(self.station_entry.get())
                        max_distance = distances[-1]
                        
                        # Draw vertical lines at each station
                        station = 0
                        while station <= max_distance:
                            custom_ax.axvline(x=station, color='gray', linestyle=':', linewidth=1, 
                                            alpha=0.6, zorder=1)
                            
                            # Add station label (engineering notation: 0+00, 1+00, etc.)
                            hundreds = int(station / 100)
                            tens_ones = int(station % 100)
                            label = f"{hundreds}+{tens_ones:02d}"
                            
                            # Position label at bottom of chart
                            y_min, y_max = custom_ax.get_ylim()
                            custom_ax.text(station, y_min, label, 
                                         rotation=90, verticalalignment='bottom', 
                                         horizontalalignment='right',
                                         fontsize=8, color='gray', alpha=0.7)
                            
                            station += station_interval
                    except (ValueError, AttributeError):
                        pass  # Skip if invalid interval or not enabled
                
                # Slope annotations removed per user request
                # # Add Slope Annotations if slopes exist (only in 2D view)
                # if 'slope' in self.profile_data[0] and not getattr(self, 'view_3d', False):
                #     # Add slope annotations at key points
                #     num_annotations = min(10, len(distances) // 10)  # Limit to ~10 annotations
                #     step = max(1, len(distances) // num_annotations)
                #     
                #     for i in range(0, len(distances) - 1, step):
                #         slope = self.profile_data[i].get('slope', 0)
                #         if abs(slope) > 0.1:  # Only show significant slopes
                #             mid_x = distances[i]
                #             mid_y = elevations[i]
                #             
                #             # Color code by slope severity
                #             if abs(slope) < 3:
                #                 color = '#2ecc71'  # Green (gentle)
                #             elif abs(slope) < 8:
                #                 color = '#f39c12'  # Orange (moderate)
                #             else:
                #                 color = '#e74c3c'  # Red (steep)
                #             
                #             # Add small text annotation
                #             custom_ax.annotate(f'{slope:.1f}%', 
                #                              xy=(mid_x, mid_y),
                #                              xytext=(0, 10), textcoords='offset points',
                #                              fontsize=7, color=color, weight='bold',
                #                              ha='center', alpha=0.8)
                
                # Styling
                custom_ax.set_facecolor('white')
                custom_ax.set_xlabel('Distance (ft)', color='black', fontsize=11, fontweight='normal')
                custom_ax.set_ylabel('Elevation (ft)', color='black', fontsize=11, fontweight='normal')
                
                # Title
                title_text = 'Elevation Profile'
                if hasattr(self, 'project_name') and self.project_name and self.project_name != 'Elevation Profile':
                    title_text = f'Elevation Profile - {self.project_name}'
                custom_ax.set_title(title_text, color='black', fontsize=12, fontweight='normal', pad=10)
                
                # Grid
                custom_ax.grid(True, linestyle='-', alpha=0.3, color='#cccccc', linewidth=0.5)
                custom_ax.tick_params(colors='black', labelsize=9)
                
                # Format axes
                custom_ax.ticklabel_format(style='plain', axis='both')
                
                # Custom X-axis ticks based on user setting
                if 'tick_interval' in settings and settings['tick_interval'] > 0:
                    max_distance = distances[-1]
                    tick_positions = []
                    tick_val = 0
                    while tick_val <= max_distance:
                        tick_positions.append(tick_val)
                        tick_val += settings['tick_interval']
                    
                    # Add final point if not already there
                    if tick_positions[-1] < max_distance:
                        tick_positions.append(max_distance)
                    
                    custom_ax.set_xticks(tick_positions)
                    custom_ax.set_xticklabels([f'{int(t)}' for t in tick_positions], rotation=0)
                
                # Legend
                custom_ax.legend(loc='best', frameon=True, fancybox=False, shadow=False,
                               fontsize=9, edgecolor='black')
                
                # Spines
                for spine in custom_ax.spines.values():
                    spine.set_edgecolor('black')
                    spine.set_linewidth(1)
                
                custom_fig.tight_layout()
                
                # Save with custom DPI
                custom_fig.savefig(filename, format=format_type, dpi=settings['dpi'], 
                                  bbox_inches='tight', facecolor='white', edgecolor='none')
                
                # Clean up
                plt.close(custom_fig)
                
                # Log details
                size_info = f"{settings['width']}\" x {settings['height']}\""
                if format_type == 'png':
                    size_info += f" @ {settings['dpi']} DPI"
                logging.info(f"Chart exported successfully: {filename} ({size_info})")
                messagebox.showinfo("Export Success", 
                                  f"Chart exported to:\n{filename}\n\nSize: {size_info}")
            except Exception as e:
                logging.error(f"Export failed: {str(e)}")
                messagebox.showerror("Export Error", f"Failed to export chart:\n{str(e)}")
                
    def _export_data(self, format_type):
        """Export profile data to CSV or Excel"""
        if not self.profile_data:
            messagebox.showwarning("No Data", "Please generate an elevation profile first.")
            return
            
        # Create dataframe
        df = pd.DataFrame(self.profile_data)
        
        # Format numeric columns to 2 decimal places for cleaner output
        # BUT preserve full precision for coordinates (Latitude/Longitude)
        numeric_columns = ['Distance_ft', 'Elevation_ft', 
                          'Running_Line_ft', 'Cut_Fill_ft', 'Cover_inches', 'Slope_%']
        
        for col in numeric_columns:
            if col in df.columns:
                df[col] = df[col].round(2)
        
        # File dialog
        if format_type == 'csv':
            filename = filedialog.asksaveasfilename(
                defaultextension='.csv',
                filetypes=[('CSV File', '*.csv')],
                initialfile=f'elevation_profile_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )
            if filename:
                try:
                    df.to_csv(filename, index=False)
                    logging.info(f"Data exported to CSV: {filename}")
                    messagebox.showinfo("Export Success", f"Data exported to:\n{filename}")
                except Exception as e:
                    logging.error(f"Export failed: {str(e)}")
                    messagebox.showerror("Export Error", f"Failed to export data:\n{str(e)}")
                    
        elif format_type == 'excel':
            filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel Workbook', '*.xlsx')],
                initialfile=f'elevation_profile_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            if filename:
                try:
                    self._create_professional_excel_report(filename)
                    logging.info(f"Professional Excel report exported: {filename}")
                    messagebox.showinfo("Export Success", f"Professional report exported to:\n{filename}")
                except Exception as e:
                    logging.error(f"Export failed: {str(e)}")
                    messagebox.showerror("Export Error", f"Failed to export data:\n{str(e)}")
    
    def _export_package(self):
        """Export complete package: Excel, PNG chart, and KML"""
        if not self.profile_data:
            messagebox.showwarning("No Data", "Please generate an elevation profile first.")
            return
        
        # Ask for folder location
        folder = filedialog.askdirectory(title="Select Folder for Export Package")
        if not folder:
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"elevation_profile_{timestamp}"
        
        success_count = 0
        errors = []
        
        try:
            # Export Excel
            excel_path = os.path.join(folder, f"{base_name}.xlsx")
            self._create_professional_excel_report(excel_path)
            logging.info(f"Package: Excel exported to {excel_path}")
            success_count += 1
        except Exception as e:
            errors.append(f"Excel: {str(e)}")
            logging.error(f"Package Excel export failed: {str(e)}")
        
        try:
            # Export PNG chart
            png_path = os.path.join(folder, f"{base_name}.png")
            if self.chart_figure:
                # Get default settings
                settings = {'width': 12, 'height': 6, 'dpi': 300, 'tight': True}
                custom_fig = Figure(figsize=(settings['width'], settings['height']), dpi=100, facecolor='white')
                custom_ax = custom_fig.add_subplot(111)
                
                # Extract data
                distances = [p['distance'] for p in self.profile_data]
                elevations = [p['elevation'] for p in self.profile_data]
                
                # Plot ground
                custom_ax.plot(distances, elevations, color='#2d7a2d', linewidth=2, 
                              label='Ground Elevation (feet)', zorder=2)
                
                # Plot running line with optional curve smoothing
                if self.enable_running_line and 'running_line' in self.profile_data[0]:
                    running_line = [p['running_line'] for p in self.profile_data]
                    
                    # Apply curve smoothing if enabled and break points exist
                    if self.curve_enabled and self.break_points:
                        logging.info(f"Package PNG: Applying curve smoothing with {len(self.break_points)} break points")
                        # Collect key points (start, break points, end)
                        key_points = [(distances[0], running_line[0])]
                        
                        for bp_dist, bp_elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
                            # Find running line elevation at break point
                            for p in self.profile_data:
                                if abs(p['distance'] - bp_dist) < 1.0:
                                    key_points.append((bp_dist, p['running_line']))
                                    break
                        
                        key_points.append((distances[-1], running_line[-1]))
                        
                        # Generate smooth curve through key points
                        smooth_x, smooth_y = self._catmull_rom_spline(key_points, num_samples=30)
                        
                        # Plot the smooth curve
                        custom_ax.plot(smooth_x, smooth_y, color='#cc0000', linewidth=2,
                                     label=f'Running Line (Curved, {self.running_line_depth_inches}")', zorder=2)
                    else:
                        custom_ax.plot(distances, running_line, color='#cc0000', linewidth=2,
                                     label=f'Running Line ({self.running_line_depth_inches}")', zorder=2)
                
                custom_ax.set_xlabel('Distance (feet)', fontsize=12, fontweight='bold')
                custom_ax.set_ylabel('Elevation (feet)', fontsize=12, fontweight='bold')
                custom_ax.set_title('Elevation Profile', fontsize=14, fontweight='bold')
                custom_ax.legend(loc='upper right', fontsize=10, framealpha=0.9)
                custom_ax.grid(True, alpha=0.3)
                custom_ax.set_facecolor('#f0f0f0')
                
                if settings['tight']:
                    custom_fig.tight_layout()
                
                custom_fig.savefig(png_path, dpi=settings['dpi'], facecolor='white', 
                                 bbox_inches='tight' if settings['tight'] else None)
                logging.info(f"Package: PNG chart exported to {png_path}")
                success_count += 1
        except Exception as e:
            errors.append(f"PNG: {str(e)}")
            logging.error(f"Package PNG export failed: {str(e)}")
        
        try:
            # Export KML
            kml_path = os.path.join(folder, f"{base_name}.kml")
            self._save_kml_to_file(kml_path)
            logging.info(f"Package: KML exported to {kml_path}")
            success_count += 1
        except Exception as e:
            errors.append(f"KML: {str(e)}")
            logging.error(f"Package KML export failed: {str(e)}")
        
        # Show results
        if success_count == 3:
            messagebox.showinfo("Export Package Success", 
                              f"Complete package exported to:\n{folder}\n\n"
                              f"Files:\n- {base_name}.xlsx\n- {base_name}.png\n- {base_name}.kml")
        elif success_count > 0:
            error_msg = "\n".join(errors)
            messagebox.showwarning("Partial Export", 
                                  f"Exported {success_count}/3 files to:\n{folder}\n\nErrors:\n{error_msg}")
        else:
            error_msg = "\n".join(errors)
            messagebox.showerror("Export Failed", f"Failed to export package:\n{error_msg}")
                    
    def _clear_log(self):
        """Clear the log text widget"""
        self.log_text.configure(state='normal')
        self.log_text.delete('1.0', tk.END)
        self.log_text.configure(state='disabled')
        logging.info("Log cleared")
    
    def _save_project(self):
        """Save current project to file"""
        # File dialog
        filename = filedialog.asksaveasfilename(
            defaultextension='.profproj',
            filetypes=[('Profile Project', '*.profproj'), ('JSON File', '*.json')],
            initialfile=f'profile_project_{datetime.now().strftime("%Y%m%d_%H%M%S")}.profproj'
        )
        
        if not filename:
            return
        
        try:
            # Gather all project data
            project_data = {
                'version': '1.5.0',
                'saved_date': datetime.now().isoformat(),
                'waypoints': self.waypoints,
                'project_name': self.project_entry.get() if hasattr(self, 'project_entry') else '',
                'api_source': self.api_source,
                'interval_ft': self.interval_ft,
                'running_line': {
                    'enabled': self.enable_running_line,
                    'depth_inches': self.running_line_depth_inches,
                    'depth_zones': self.depth_zones  # Variable depth zones
                },
                'station_markers': {
                    'enabled': self.enable_stations,
                    'interval_ft': float(self.station_entry.get()) if hasattr(self, 'station_entry') else 100
                },
                'profile_data': self.profile_data if self.profile_data else [],
                'service_url': self.service_url,
                'view_3d': getattr(self, 'view_3d', False)
            }
            
            # Save to JSON
            with open(filename, 'w') as f:
                json.dump(project_data, f, indent=2)
            
            logging.info(f"Project saved: {filename}")
            messagebox.showinfo("Project Saved", 
                              f"Project saved successfully!\n\n{filename}")
        
        except Exception as e:
            logging.error(f"Failed to save project: {str(e)}")
            messagebox.showerror("Save Error", f"Failed to save project:\n\n{str(e)}")
    
    def _load_project(self):
        """Load project from file"""
        # File dialog
        filename = filedialog.askopenfilename(
            filetypes=[('Profile Project', '*.profproj'), ('JSON File', '*.json'), ('All Files', '*.*')],
            title="Load Profile Project"
        )
        
        if not filename:
            return
        
        try:
            # Load JSON
            with open(filename, 'r') as f:
                project_data = json.load(f)
            
            # Clear current state
            self._clear_points()
            
            # Restore waypoints
            self.waypoints = [tuple(wp) for wp in project_data.get('waypoints', [])]
            
            # Restore settings
            if 'project_name' in project_data and hasattr(self, 'project_entry'):
                self.project_entry.delete(0, tk.END)
                self.project_entry.insert(0, project_data['project_name'])
            
            if 'interval_ft' in project_data:
                self.interval_ft = project_data['interval_ft']
                self.interval_entry.delete(0, tk.END)
                self.interval_entry.insert(0, str(self.interval_ft))
            
            if 'running_line' in project_data:
                self.enable_running_line = project_data['running_line'].get('enabled', True)
                self.running_line_depth_inches = project_data['running_line'].get('depth_inches', 42)
                
                # Restore depth zones
                self.depth_zones = project_data['running_line'].get('depth_zones', [])
                # Convert lists back to tuples (start, end, depth, note)
                self.depth_zones = [(z[0], z[1], z[2], z[3] if len(z) > 3 else '') for z in self.depth_zones]
            
            if 'station_markers' in project_data:
                self.enable_stations = project_data['station_markers'].get('enabled', True)
                self.station_interval_ft = project_data['station_markers'].get('interval_ft', 100)
            
            # Restore API source (stored internally, UI is hardcoded to USGS)
            if 'api_source' in project_data:
                self.api_source = project_data['api_source']
            
            # Restore profile data if exists
            if 'profile_data' in project_data and project_data['profile_data']:
                self.profile_data = project_data['profile_data']
                self._draw_profile_chart()
            
            # Restore waypoint markers on map
            for i, waypoint in enumerate(self.waypoints):
                lat, lon = waypoint
                if i == 0:
                    color_circle = "#27AE60"
                    color_outside = "#229954"
                    text = "Start"
                elif i == len(self.waypoints) - 1:
                    color_circle = "#E74C3C"
                    color_outside = "#C0392B"
                    text = "End"
                else:
                    color_circle = "#42A5F5"
                    color_outside = "#2196F3"
                    text = f"Pt {i + 1}"
                
                marker = self.map_widget.set_marker(
                    lat, lon,
                    text=text,
                    marker_color_circle=color_circle,
                    marker_color_outside=color_outside
                )
                self.waypoint_markers.append(marker)
            
            # Draw path
            if len(self.waypoints) > 1:
                self.path_line = self.map_widget.set_path(
                    self.waypoints,
                    color="#42A5F5",
                    width=3
                )
            
            # Update display
            self._update_waypoint_display()
            
            # Zoom to fit
            if self.waypoints:
                lats = [p[0] for p in self.waypoints]
                lons = [p[1] for p in self.waypoints]
                center_lat = (min(lats) + max(lats)) / 2
                center_lon = (min(lons) + max(lons)) / 2
                self.map_widget.set_position(center_lat, center_lon)
            
            logging.info(f"Project loaded: {filename}")
            messagebox.showinfo("Project Loaded", 
                              f"Project loaded successfully!\n\n"
                              f"Waypoints: {len(self.waypoints)}\n"
                              f"Profile data: {'Yes' if self.profile_data else 'No (generate new)'}")
        
        except Exception as e:
            logging.error(f"Failed to load project: {str(e)}")
            messagebox.showerror("Load Error", f"Failed to load project:\n\n{str(e)}")
    
    def _export_kml(self):
        """Export route and elevation data to KML for Google Earth"""
        if not self.waypoints:
            messagebox.showwarning("No Data", "Please add waypoints or generate a profile first.")
            return
        
        # File dialog
        filename = filedialog.asksaveasfilename(
            defaultextension='.kml',
            filetypes=[('KML File', '*.kml')],
            initialfile=f'elevation_profile_{datetime.now().strftime("%Y%m%d_%H%M%S")}.kml'
        )
        
        if not filename:
            return
        
        try:
            self._save_kml_to_file(filename)
            logging.info(f"KML exported: {filename}")
            messagebox.showinfo("KML Export Success", 
                              f"KML file exported successfully!\n\n"
                              f"{filename}\n\n"
                              f"Open in Google Earth to view 3D elevation profile.")
        
        except Exception as e:
            logging.error(f"KML export failed: {str(e)}")
            messagebox.showerror("KML Export Error", f"Failed to export KML:\n\n{str(e)}")
    
    def _save_kml_to_file(self, filename):
        """Save KML content to specified file (helper for _export_kml and _export_package)"""
        # Build KML content
        kml_content = ['<?xml version="1.0" encoding="UTF-8"?>']
        kml_content.append('<kml xmlns="http://www.opengis.net/kml/2.2">')
        kml_content.append('<Document>')
        
        # Document name
        project_name = self.project_entry.get() if hasattr(self, 'project_entry') and self.project_entry.get() else 'Elevation Profile'
        kml_content.append(f'<name>{project_name}</name>')
        kml_content.append(f'<description>Generated by Profile Automation Tool - {datetime.now().strftime("%Y-%m-%d %H:%M")}</description>')
        
        # Styles
        kml_content.append('<Style id="groundLine">')
        kml_content.append('<LineStyle>')
        kml_content.append('<color>ff00ff00</color>')
        kml_content.append('<width>3</width>')
        kml_content.append('</LineStyle>')
        kml_content.append('<PolyStyle>')
        kml_content.append('<color>7f00ff00</color>')
        kml_content.append('</PolyStyle>')
        kml_content.append('</Style>')
        
        kml_content.append('<Style id="runningLine">')
        kml_content.append('<LineStyle>')
        kml_content.append('<color>ff0000ff</color>')
        kml_content.append('<width>2</width>')
        kml_content.append('</LineStyle>')
        kml_content.append('<PolyStyle>')
        kml_content.append('<color>7f0000ff</color>')
        kml_content.append('</PolyStyle>')
        kml_content.append('</Style>')
        
        kml_content.append('<Style id="waypoint">')
        kml_content.append('<IconStyle>')
        kml_content.append('<color>ff00ffff</color>')
        kml_content.append('<scale>1.2</scale>')
        kml_content.append('</IconStyle>')
        kml_content.append('</Style>')
        
        # Add waypoints
        kml_content.append('<Folder>')
        kml_content.append('<name>Waypoints</name>')
        for i, waypoint in enumerate(self.waypoints):
            lat, lon = waypoint
            elev = 0
            if self.profile_data:
                closest = min(self.profile_data, key=lambda p: abs(p['y'] - lat) + abs(p['x'] - lon))
                elev = closest['elevation'] * 0.3048
            
            kml_content.append('<Placemark>')
            kml_content.append(f'<name>Point {i + 1}</name>')
            kml_content.append('<styleUrl>#waypoint</styleUrl>')
            kml_content.append('<Point>')
            kml_content.append('<altitudeMode>absolute</altitudeMode>')
            kml_content.append(f'<coordinates>{lon},{lat},{elev}</coordinates>')
            kml_content.append('</Point>')
            kml_content.append('</Placemark>')
        kml_content.append('</Folder>')
        
        # Add ground elevation line
        if self.profile_data:
            kml_content.append('<Placemark>')
            kml_content.append('<name>Ground Elevation</name>')
            kml_content.append('<styleUrl>#groundLine</styleUrl>')
            kml_content.append('<LineString>')
            kml_content.append('<extrude>1</extrude>')
            kml_content.append('<tessellate>1</tessellate>')
            kml_content.append('<altitudeMode>absolute</altitudeMode>')
            kml_content.append('<coordinates>')
            
            coords = []
            for point in self.profile_data:
                elev_m = point['elevation'] * 0.3048
                coords.append(f'{point["x"]},{point["y"]},{elev_m}')
            
            kml_content.append(' '.join(coords))
            kml_content.append('</coordinates>')
            kml_content.append('</LineString>')
            kml_content.append('</Placemark>')
            
            # Add running line
            if self.enable_running_line and 'running_line' in self.profile_data[0]:
                kml_content.append('<Placemark>')
                description = f'Utility line at {self.running_line_depth_inches}" depth'
                if self.curve_enabled and self.break_points:
                    description += ' (Curved)'
                kml_content.append('<name>Running Line</name>')
                kml_content.append(f'<description>{description}</description>')
                kml_content.append('<styleUrl>#runningLine</styleUrl>')
                kml_content.append('<LineString>')
                kml_content.append('<extrude>1</extrude>')
                kml_content.append('<tessellate>1</tessellate>')
                kml_content.append('<altitudeMode>absolute</altitudeMode>')
                kml_content.append('<coordinates>')
                
                # Generate coordinates - use curve interpolation if enabled
                if self.curve_enabled and self.break_points:
                    logging.info(f"KML: Applying curve smoothing for running line")
                    # Get distances and running line elevations
                    distances = [p['distance'] for p in self.profile_data]
                    running_line = [p['running_line'] for p in self.profile_data]
                    
                    # Collect key points for curve
                    key_points = [(distances[0], running_line[0])]
                    for bp_dist, bp_elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
                        for p in self.profile_data:
                            if abs(p['distance'] - bp_dist) < 1.0:
                                key_points.append((bp_dist, p['running_line']))
                                break
                    key_points.append((distances[-1], running_line[-1]))
                    
                    # Generate smooth curve
                    smooth_distances, smooth_elevations = self._catmull_rom_spline(key_points, num_samples=30)
                    
                    # Interpolate x,y coordinates for each smooth point
                    coords = []
                    for smooth_dist, smooth_elev in zip(smooth_distances, smooth_elevations):
                        # Find closest original points to interpolate x,y
                        idx = min(range(len(distances)), key=lambda i: abs(distances[i] - smooth_dist))
                        x = self.profile_data[idx]['x']
                        y = self.profile_data[idx]['y']
                        elev_m = smooth_elev * 0.3048
                        coords.append(f'{x},{y},{elev_m}')
                else:
                    # Use original points (linear)
                    coords = []
                    for point in self.profile_data:
                        elev_m = point['running_line'] * 0.3048
                        coords.append(f'{point["x"]},{point["y"]},{elev_m}')
                
                kml_content.append(' '.join(coords))
                kml_content.append('</coordinates>')
                kml_content.append('</LineString>')
                kml_content.append('</Placemark>')
        
        kml_content.append('</Document>')
        kml_content.append('</kml>')
        
        # Write to file
        with open(filename, 'w') as f:
            f.write('\n'.join(kml_content))
    
    def _import_points(self):
        """Import points from CSV or Shapefile"""
        # File dialog
        filename = filedialog.askopenfilename(
            title="Import Points or Line",
            filetypes=[
                ('All Supported', '*.csv;*.shp;*.kml'),
                ('CSV Files', '*.csv'),
                ('Shapefiles', '*.shp'),
                ('KML/KMZ', '*.kml;*.kmz'),
                ('All Files', '*.*')
            ]
        )
        
        if not filename:
            return
        
        try:
            file_ext = Path(filename).suffix.lower()
            
            if file_ext == '.csv':
                self._import_csv(filename)
            elif file_ext == '.shp':
                self._import_shapefile(filename)
            elif file_ext in ['.kml', '.kmz']:
                self._import_kml(filename)
            else:
                messagebox.showerror("Unsupported Format", f"File format {file_ext} is not supported.")
                
        except Exception as e:
            logging.error(f"Failed to import file: {str(e)}")
            messagebox.showerror("Import Error", f"Failed to import file:\n\n{str(e)}")
    
    def _import_csv(self, filename):
        """Import points from CSV file"""
        logging.info(f"Importing CSV: {filename}")
        
        try:
            # Try to read CSV
            df = pd.read_csv(filename)
            
            # Look for lat/lon columns (case insensitive)
            lat_col = None
            lon_col = None
            
            for col in df.columns:
                col_lower = col.lower()
                if col_lower in ['lat', 'latitude', 'y']:
                    lat_col = col
                if col_lower in ['lon', 'long', 'longitude', 'x']:
                    lon_col = col
            
            if not lat_col or not lon_col:
                messagebox.showerror("CSV Error", 
                    "Could not find latitude/longitude columns.\n\n"
                    "CSV must have columns named:\n"
                    "- lat, latitude, or y\n"
                    "- lon, long, longitude, or x")
                return
            
            # Clear existing waypoints
            self._clear_points()
            
            # Add points
            points_added = 0
            for idx, row in df.iterrows():
                try:
                    lat = float(row[lat_col])
                    lon = float(row[lon_col])
                    
                    # Validate coordinates
                    if -90 <= lat <= 90 and -180 <= lon <= 180:
                        self.waypoints.append((lat, lon))
                        
                        # Create marker
                        if points_added == 0:
                            color_circle = "#27AE60"
                            color_outside = "#229954"
                            text = "Start"
                        else:
                            color_circle = "#42A5F5"
                            color_outside = "#2196F3"
                            text = f"Point {points_added + 1}"
                        
                        marker = self.map_widget.set_marker(
                            lat, lon,
                            text=text,
                            marker_color_circle=color_circle,
                            marker_color_outside=color_outside
                        )
                        self.waypoint_markers.append(marker)
                        points_added += 1
                    else:
                        logging.warning(f"Skipping invalid coordinates at row {idx + 1}: ({lat}, {lon})")
                        
                except (ValueError, TypeError) as e:
                    logging.warning(f"Skipping row {idx + 1}: {str(e)}")
                    continue
            
            if points_added < 2:
                messagebox.showwarning("Not Enough Points", 
                    f"Only {points_added} valid point(s) imported.\n"
                    "Need at least 2 points for a profile.")
                return
            
            # Draw path
            if len(self.waypoints) > 1:
                self.path_line = self.map_widget.set_path(
                    self.waypoints,
                    color="#42A5F5",
                    width=3
                )
            
            # Zoom to fit
            if self.waypoints:
                lats = [p[0] for p in self.waypoints]
                lons = [p[1] for p in self.waypoints]
                center_lat = (min(lats) + max(lats)) / 2
                center_lon = (min(lons) + max(lons)) / 2
                self.map_widget.set_position(center_lat, center_lon)
                
                # Calculate appropriate zoom
                lat_range = max(lats) - min(lats)
                lon_range = max(lons) - min(lons)
                max_range = max(lat_range, lon_range)
                
                if max_range < 0.01:
                    zoom = 16
                elif max_range < 0.05:
                    zoom = 14
                elif max_range < 0.1:
                    zoom = 13
                else:
                    zoom = 12
                self.map_widget.set_zoom(zoom)
            
            # Update UI
            self._update_waypoint_display()
            logging.info(f"Successfully imported {points_added} points from CSV")
            messagebox.showinfo("Import Successful", 
                f"Imported {points_added} points from CSV file.\n\n"
                f"Click 'Generate Profile' to create elevation profile.")
                
        except Exception as e:
            raise Exception(f"CSV parsing error: {str(e)}")
    
    def _import_shapefile(self, filename):
        """Import points or line from Shapefile"""
        logging.info(f"Importing Shapefile: {filename}")
        
        try:
            import shapefile as shp
            
            # Read shapefile
            sf = shp.Reader(filename)
            
            # Check geometry type
            shape_type = sf.shapeTypeName
            logging.info(f"Shapefile geometry type: {shape_type}")
            
            if shape_type in ['POINT', 'POINTZ', 'POINTM']:
                self._import_point_shapefile(sf)
            elif shape_type in ['POLYLINE', 'POLYLINEZ', 'POLYLINEM', 'ARC']:
                self._import_line_shapefile(sf)
            else:
                messagebox.showerror("Unsupported Geometry", 
                    f"Shapefile geometry type '{shape_type}' is not supported.\n\n"
                    f"Supported types: POINT, POLYLINE")
                
        except ImportError:
            messagebox.showerror("Missing Library", 
                "The 'pyshp' library is required for shapefile import.\n\n"
                "Install it with: pip install pyshp")
        except Exception as e:
            raise Exception(f"Shapefile error: {str(e)}")
    
    def _import_point_shapefile(self, sf):
        """Import points from point shapefile"""
        # Clear existing waypoints
        self._clear_points()
        
        points_added = 0
        for shape in sf.shapes():
            try:
                lon, lat = shape.points[0]
                
                # Validate coordinates
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    self.waypoints.append((lat, lon))
                    
                    # Create marker
                    if points_added == 0:
                        color_circle = "#27AE60"
                        color_outside = "#229954"
                        text = "Start"
                    else:
                        color_circle = "#42A5F5"
                        color_outside = "#2196F3"
                        text = f"Point {points_added + 1}"
                    
                    marker = self.map_widget.set_marker(
                        lat, lon,
                        text=text,
                        marker_color_circle=color_circle,
                        marker_color_outside=color_outside
                    )
                    self.waypoint_markers.append(marker)
                    points_added += 1
                else:
                    logging.warning(f"Skipping invalid coordinates: ({lat}, {lon})")
                    
            except Exception as e:
                logging.warning(f"Skipping point: {str(e)}")
                continue
        
        if points_added < 2:
            messagebox.showwarning("Not Enough Points", 
                f"Only {points_added} valid point(s) imported.\n"
                "Need at least 2 points for a profile.")
            return
        
        # Draw path and zoom
        self._finalize_import(points_added, "point shapefile")
    
    def _import_line_shapefile(self, sf):
        """Import line vertices from line shapefile (densification happens during profile generation)"""
        result = messagebox.askyesno(
            "Import Line?",
            f"Line shapefile detected!\n\n"
            f"Import line vertices?\n\n"
            f"The line will be automatically densified at your specified interval\n"
            f"when you generate the elevation profile.")
        
        if not result:
            return
        
        # Clear existing waypoints
        self._clear_points()
        
        points_added = 0
        for shape in sf.shapes():
            try:
                # Get all vertices in the line
                line_points = shape.points
                
                if len(line_points) < 2:
                    continue
                
                # Store ONLY the segment vertices (no densification yet)
                for lon, lat in line_points:
                    # Validate coordinates
                    if -90 <= lat <= 90 and -180 <= lon <= 180:
                        self.waypoints.append((lat, lon))
                        points_added += 1
                
            except Exception as e:
                logging.warning(f"Skipping line feature: {str(e)}")
                continue
        
        if points_added < 2:
            messagebox.showwarning("Not Enough Points", 
                f"Only {points_added} valid vertices imported.\n"
                "Need at least 2 points for a profile.")
            return
        
        # Add markers for all vertices (since we're not densifying yet)
        for idx, (lat, lon) in enumerate(self.waypoints):
            if idx == 0:
                color_circle = "#27AE60"
                color_outside = "#229954"
                text = "Start"
            elif idx == len(self.waypoints) - 1:
                color_circle = "#E74C3C"
                color_outside = "#C0392B"
                text = "End"
            else:
                color_circle = "#42A5F5"
                color_outside = "#2196F3"
                text = f"Vertex {idx + 1}"
            
            marker = self.map_widget.set_marker(
                lat, lon,
                text=text,
                marker_color_circle=color_circle,
                marker_color_outside=color_outside
            )
            self.waypoint_markers.append(marker)
        
        # Draw path and zoom
        self._finalize_import(points_added, "line shapefile")
    
    def _import_kml(self, filename):
        """Import from KML/KMZ file"""
        messagebox.showinfo("KML Support", 
            "KML/KMZ import is not yet implemented.\n\n"
            "Please convert to Shapefile or CSV first.")
    
    def _add_markers_and_finalize(self, points_added, source_type):
        """Add markers to map and finalize import"""
        if points_added < 2:
            messagebox.showwarning("Not Enough Points", 
                f"Only {points_added} valid point(s) imported.\n"
                "Need at least 2 points for a profile.")
            return
        
        # Add markers
        marker_interval = max(1, points_added // 20)
        for idx, (lat, lon) in enumerate(self.waypoints):
            if idx % marker_interval == 0 or idx == len(self.waypoints) - 1:
                if idx == 0:
                    text = "Start"
                    color_circle = "#27AE60"
                    color_outside = "#229954"
                elif idx == len(self.waypoints) - 1:
                    text = "End"
                    color_circle = "#E74C3C"
                    color_outside = "#C0392B"
                else:
                    text = f"Pt {idx + 1}"
                    color_circle = "#42A5F5"
                    color_outside = "#2196F3"
                
                marker = self.map_widget.set_marker(
                    lat, lon,
                    text=text,
                    marker_color_circle=color_circle,
                    marker_color_outside=color_outside
                )
                self.waypoint_markers.append(marker)
        
        self._finalize_import(points_added, source_type)
    
    def _finalize_import(self, points_added, source_type):
        """Finalize import: draw path, zoom, update UI"""
        # Draw path
        if len(self.waypoints) > 1:
            self.path_line = self.map_widget.set_path(
                self.waypoints,
                color="#42A5F5",
                width=3
            )
        
        # Zoom to fit
        if self.waypoints:
            lats = [p[0] for p in self.waypoints]
            lons = [p[1] for p in self.waypoints]
            center_lat = (min(lats) + max(lats)) / 2
            center_lon = (min(lons) + max(lons)) / 2
            self.map_widget.set_position(center_lat, center_lon)
            
            # Calculate appropriate zoom
            lat_range = max(lats) - min(lats)
            lon_range = max(lons) - min(lons)
            max_range = max(lat_range, lon_range)
            
            if max_range < 0.01:
                zoom = 16
            elif max_range < 0.05:
                zoom = 14
            elif max_range < 0.1:
                zoom = 13
            else:
                zoom = 12
            self.map_widget.set_zoom(zoom)
        
        # Update UI
        self._update_waypoint_display()
        logging.info(f"Successfully imported {points_added} points from {source_type}")
        messagebox.showinfo("Import Successful", 
            f"Imported {points_added} points from {source_type}.\n\n"
            f"Click 'Generate Profile' to create elevation profile.")
        
    def _show_export_settings_dialog(self, format_type):
        """Show dialog for export dimensions and quality settings"""
        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Export {format_type.upper()} Settings")
        dialog.geometry("500x700")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (700 // 2)
        dialog.geometry(f"500x700+{x}+{y}")
        
        result = {}
        
        # Main container
        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ctk.CTkLabel(
            main_frame,
            text=f"📐 Configure {format_type.upper()} Export",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        title_label.pack(pady=(0, 20))
        
        # Preset sizes section
        preset_frame = ctk.CTkFrame(main_frame)
        preset_frame.pack(fill=tk.X, pady=(0, 15))
        
        preset_label = ctk.CTkLabel(
            preset_frame,
            text="Paper Size Presets:",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        preset_label.pack(pady=(10, 5), padx=10, anchor="w")
        
        # Preset size definitions (width x height in inches)
        presets = {
            "Custom": (30, 5),
            "Letter (8.5\" x 11\")": (11, 8.5),
            "Legal (8.5\" x 14\")": (14, 8.5),
            "Tabloid (11\" x 17\")": (17, 11),
            "A4 (8.27\" x 11.69\")": (11.69, 8.27),
            "A3 (11.69\" x 16.54\")": (16.54, 11.69),
            "Wide Screen (16:9)": (16, 9),
            "Large Format (24\" x 18\")": (24, 18),
            "Presentation (10\" x 7.5\")": (10, 7.5)
        }
        
        preset_var = tk.StringVar(value="Custom")
        
        # Width input
        width_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        width_frame.pack(fill=tk.X, pady=5)
        
        width_label = ctk.CTkLabel(
            width_frame,
            text="Width (inches):",
            font=ctk.CTkFont(size=12, weight="bold"),
            width=140,
            anchor="w"
        )
        width_label.pack(side=tk.LEFT, padx=(10, 5))
        
        width_entry = ctk.CTkEntry(width_frame, width=100)
        width_entry.pack(side=tk.LEFT, padx=(0, 5))
        width_entry.insert(0, "20")
        
        width_info = ctk.CTkLabel(
            width_frame,
            text="(chart width)",
            font=ctk.CTkFont(size=10),
            text_color="gray60"
        )
        width_info.pack(side=tk.LEFT)
        
        # Height input
        height_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        height_frame.pack(fill=tk.X, pady=5)
        
        height_label = ctk.CTkLabel(
            height_frame,
            text="Height (inches):",
            font=ctk.CTkFont(size=12, weight="bold"),
            width=140,
            anchor="w"
        )
        height_label.pack(side=tk.LEFT, padx=(10, 5))
        
        height_entry = ctk.CTkEntry(height_frame, width=100)
        height_entry.pack(side=tk.LEFT, padx=(0, 5))
        height_entry.insert(0, "5")
        
        height_info = ctk.CTkLabel(
            height_frame,
            text="(chart height)",
            font=ctk.CTkFont(size=10),
            text_color="gray60"
        )
        height_info.pack(side=tk.LEFT)
        
        # DPI input (for PNG)
        dpi_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        dpi_frame.pack(fill=tk.X, pady=5)
        
        dpi_label = ctk.CTkLabel(
            dpi_frame,
            text="Resolution (DPI):",
            font=ctk.CTkFont(size=12, weight="bold"),
            width=140,
            anchor="w"
        )
        dpi_label.pack(side=tk.LEFT, padx=(10, 5))
        
        dpi_entry = ctk.CTkEntry(dpi_frame, width=100)
        dpi_entry.pack(side=tk.LEFT, padx=(0, 5))
        dpi_entry.insert(0, "300")
        
        dpi_info = ctk.CTkLabel(
            dpi_frame,
            text=f"({'PNG only, ignored for PDF/SVG' if format_type != 'png' else 'higher = better quality'})",
            font=ctk.CTkFont(size=10),
            text_color="gray60"
        )
        dpi_info.pack(side=tk.LEFT)
        
        if format_type != 'png':
            dpi_entry.configure(state='disabled')
        
        # X-Axis Tick Interval
        tick_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        tick_frame.pack(fill=tk.X, pady=5)
        
        tick_label = ctk.CTkLabel(
            tick_frame,
            text="X-Axis Ticks (ft):",
            font=ctk.CTkFont(size=12, weight="bold"),
            width=140,
            anchor="w"
        )
        tick_label.pack(side=tk.LEFT, padx=(10, 5))
        
        tick_entry = ctk.CTkEntry(tick_frame, width=100)
        tick_entry.pack(side=tk.LEFT, padx=(0, 5))
        tick_entry.insert(0, "100")
        
        tick_info = ctk.CTkLabel(
            tick_frame,
            text="(distance tick marks: 100, 200, 300...)",
            font=ctk.CTkFont(size=10),
            text_color="gray60"
        )
        tick_info.pack(side=tk.LEFT)
        
        # Quick tick presets
        tick_preset_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        tick_preset_frame.pack(fill=tk.X, pady=5)
        
        tick_preset_label = ctk.CTkLabel(
            tick_preset_frame,
            text="Quick Ticks:",
            font=ctk.CTkFont(size=11)
        )
        tick_preset_label.pack(side=tk.LEFT, padx=(10, 5))
        
        for tick_val in [50, 100, 200, 500]:
            btn = ctk.CTkButton(
                tick_preset_frame,
                text=f"{tick_val} ft",
                width=70,
                height=28,
                command=lambda t=tick_val: (tick_entry.delete(0, tk.END), 
                                           tick_entry.insert(0, str(t)))
            )
            btn.pack(side=tk.LEFT, padx=2)
        
        # Preset selector
        def apply_preset(choice):
            if choice in presets:
                width, height = presets[choice]
                width_entry.delete(0, tk.END)
                width_entry.insert(0, str(width))
                height_entry.delete(0, tk.END)
                height_entry.insert(0, str(height))
        
        preset_selector = ctk.CTkOptionMenu(
            preset_frame,
            values=list(presets.keys()),
            command=apply_preset,
            width=300
        )
        preset_selector.set("Custom")
        preset_selector.pack(pady=5, padx=10)
        
        # Preview info
        preview_frame = ctk.CTkFrame(main_frame)
        preview_frame.pack(fill=tk.X, pady=15)
        
        preview_label = ctk.CTkLabel(
            preview_frame,
            text="📏 Size Preview",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        preview_label.pack(pady=(10, 5))
        
        preview_text = ctk.CTkTextbox(preview_frame, height=120, width=400)
        preview_text.pack(pady=5, padx=10)
        
        def update_preview():
            try:
                w = float(width_entry.get())
                h = float(height_entry.get())
                d = int(dpi_entry.get()) if format_type == 'png' else 300
                
                # Calculate pixel dimensions
                px_width = int(w * d)
                px_height = int(h * d)
                
                # Calculate file size estimate (rough)
                if format_type == 'png':
                    est_size_mb = (px_width * px_height * 3) / (1024 * 1024)  # RGB
                    size_str = f"~{est_size_mb:.1f} MB"
                else:
                    size_str = "~1-5 MB (vector)"
                
                preview_info = f"""
Output Dimensions:
• Physical Size: {w}\" × {h}\"
• Pixel Size: {px_width} × {px_height} px
• Aspect Ratio: {w/h:.2f}:1
• Resolution: {d} DPI
• Est. File Size: {size_str}

Recommended Uses:
"""
                if w >= 16 and h >= 12:
                    preview_info += "• Large format printing\n• Poster displays"
                elif w >= 11 and h >= 8:
                    preview_info += "• Standard reports\n• Plan sets"
                elif w >= 8 and h >= 6:
                    preview_info += "• Presentations\n• Documents"
                else:
                    preview_info += "• Email attachments\n• Web display"
                
                preview_text.delete("1.0", tk.END)
                preview_text.insert("1.0", preview_info.strip())
            except:
                preview_text.delete("1.0", tk.END)
                preview_text.insert("1.0", "Enter valid numbers to see preview")
        
        # Update preview on change
        width_entry.bind('<KeyRelease>', lambda e: update_preview())
        height_entry.bind('<KeyRelease>', lambda e: update_preview())
        dpi_entry.bind('<KeyRelease>', lambda e: update_preview())
        update_preview()
        
        # DPI presets for PNG
        if format_type == 'png':
            dpi_preset_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            dpi_preset_frame.pack(fill=tk.X, pady=10)
            
            dpi_preset_label = ctk.CTkLabel(
                dpi_preset_frame,
                text="Quick DPI:",
                font=ctk.CTkFont(size=11)
            )
            dpi_preset_label.pack(side=tk.LEFT, padx=(10, 5))
            
            for dpi_val, label in [(150, "Web"), (300, "Print"), (600, "High Quality")]:
                btn = ctk.CTkButton(
                    dpi_preset_frame,
                    text=f"{label}\n({dpi_val})",
                    width=80,
                    height=40,
                    command=lambda d=dpi_val: (dpi_entry.delete(0, tk.END), 
                                               dpi_entry.insert(0, str(d)),
                                               update_preview())
                )
                btn.pack(side=tk.LEFT, padx=2)
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        def on_ok():
            try:
                result['width'] = float(width_entry.get())
                result['height'] = float(height_entry.get())
                result['dpi'] = int(dpi_entry.get()) if format_type == 'png' else 300
                result['tick_interval'] = float(tick_entry.get())
                
                if result['width'] <= 0 or result['height'] <= 0:
                    messagebox.showerror("Invalid Input", "Width and height must be positive numbers")
                    return
                    
                if result['dpi'] < 72 or result['dpi'] > 1200:
                    messagebox.showerror("Invalid Input", "DPI must be between 72 and 1200")
                    return
                
                if result['tick_interval'] <= 0:
                    messagebox.showerror("Invalid Input", "Tick interval must be a positive number")
                    return
                
                dialog.destroy()
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter valid numbers")
        
        def on_cancel():
            result.clear()
            dialog.destroy()
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            width=120,
            command=on_cancel,
            fg_color="gray40",
            hover_color="gray30"
        )
        cancel_btn.pack(side=tk.LEFT, padx=(10, 5))
        
        ok_btn = ctk.CTkButton(
            button_frame,
            text=f"Export {format_type.upper()}",
            width=180,
            command=on_ok,
            fg_color="#2E7D32",
            hover_color="#1B5E20"
        )
        ok_btn.pack(side=tk.RIGHT, padx=(5, 10))
        
        # Wait for dialog to close
        self.wait_window(dialog)
        
        return result if result else None
        
    def _create_professional_excel_report(self, filename):
        """Create professional Excel report with embedded chart"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import LineChart, Reference
        from openpyxl.drawing.image import Image as XLImage
        import io
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Elevation Profile"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Station
        ws.column_dimensions['B'].width = 18  # Ground Elevation
        ws.column_dimensions['C'].width = 18  # Running Line
        ws.column_dimensions['D'].width = 15  # Cut/Fill
        ws.column_dimensions['E'].width = 15  # Cover
        ws.column_dimensions['F'].width = 12  # Slope
        ws.column_dimensions['G'].width = 10  # Segment
        ws.column_dimensions['H'].width = 15  # Latitude
        ws.column_dimensions['I'].width = 15  # Longitude
        
        # Title styles
        title_font = Font(size=16, bold=True, color="000000")
        header_font = Font(size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        row = 1
        
        # Project Title
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        project_name = getattr(self, 'project_name', 'Elevation Profile Report')
        cell.value = project_name
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Project Information
        row += 1
        ws[f'A{row}'] = "Project Information"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        if len(self.waypoints) >= 2:
            ws[f'A{row}'] = "Start Point:"
            ws[f'B{row}'] = f"{self.waypoints[0][0]:.6f}, {self.waypoints[0][1]:.6f}"
            row += 1
            
            ws[f'A{row}'] = "End Point:"
            ws[f'B{row}'] = f"{self.waypoints[-1][0]:.6f}, {self.waypoints[-1][1]:.6f}"
            row += 1
            
            if len(self.waypoints) > 2:
                ws[f'A{row}'] = "Waypoints:"
                ws[f'B{row}'] = f"{len(self.waypoints)} points total"
                row += 1
        else:
            ws[f'A{row}'] = "Points:"
            ws[f'B{row}'] = "No waypoints set"
            row += 1
        
        ws[f'A{row}'] = "Total Distance:"
        ws[f'B{row}'] = f"{self.profile_data[-1]['distance']:,.2f} feet ({self.profile_data[-1]['distance']/5280:.2f} miles)"
        row += 1
        
        if self.enable_running_line:
            ws[f'A{row}'] = "Running Line Depth:"
            ws[f'B{row}'] = f"{self.running_line_depth_inches:.0f} inches ({self.running_line_depth_inches/12:.2f} feet)"
            row += 1
            
            if self.break_points:
                ws[f'A{row}'] = "Design Type:"
                ws[f'B{row}'] = f"Multi-Segment ({len(self.break_points) + 1} segments, {len(self.break_points)} break points)"
                row += 1
            else:
                ws[f'A{row}'] = "Design Type:"
                ws[f'B{row}'] = "Single-Segment"
                row += 1
        
        ws[f'A{row}'] = "Date Generated:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        
        ws[f'A{row}'] = "Data Source:"
        if self.api_source == "USGS":
            ws[f'B{row}'] = "USGS 3DEP Elevation API (1m resolution, ±0.3-1m vertical accuracy)"
        elif self.api_source == "ArcGIS":
            ws[f'B{row}'] = "ArcGIS World Elevation Service"
        else:
            ws[f'B{row}'] = "Mock/Test Data"
        row += 1
        
        ws[f'A{row}'] = "Generated By:"
        ws[f'B{row}'] = "Profile Automation Tool - Developed by Omid Zanganeh"
        row += 2
        
        # Break Points Summary (if any)
        if self.break_points:
            ws[f'A{row}'] = "Break Points (Segment Transitions)"
            ws[f'A{row}'].font = Font(size=12, bold=True, color="9900CC")
            row += 1
            
            # Break points header
            bp_headers = ['Distance (ft)', 'Elevation (ft)', 'Grade (%)', 'Note']
            for col, header in enumerate(bp_headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(size=10, bold=True)
                cell.fill = PatternFill(start_color="E6CCE6", end_color="E6CCE6", fill_type="solid")
                cell.border = border
            row += 1
            
            # Break points data
            for bp_dist, bp_elev, grade, locked, note in sorted(self.break_points, key=lambda x: x[0]):
                ws.cell(row=row, column=1, value=bp_dist).number_format = '0.0'
                ws.cell(row=row, column=2, value=bp_elev).number_format = '0.00'
                grade_display = grade if grade is not None else "Auto"
                ws.cell(row=row, column=3, value=grade_display)
                if isinstance(grade_display, (int, float)):
                    ws.cell(row=row, column=3).number_format = '0.00'
                ws.cell(row=row, column=4, value=note if note else "")
                row += 1
            
            row += 1
        
        # Data Table Header
        header_row = row
        headers = ['Station (ft)', 'Ground Elevation (ft)', 'Running Line (ft)', 'Cut/Fill (ft)', 'Cover (in)', 'Slope (%)', 'Segment', 'Latitude', 'Longitude']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        data_start_row = row
        
        # Data rows
        for point in self.profile_data:
            ws.cell(row=row, column=1, value=round(point['distance'], 1))
            ws.cell(row=row, column=2, value=round(point['elevation'], 2))
            
            if self.enable_running_line and 'running_line' in point:
                ws.cell(row=row, column=3, value=round(point['running_line'], 2))
                ws.cell(row=row, column=4, value=round(point['cut_fill'], 2))
                ws.cell(row=row, column=5, value=round(point['cover'] * 12, 1))  # Convert to inches
            else:
                ws.cell(row=row, column=3, value="N/A")
                ws.cell(row=row, column=4, value="N/A")
                ws.cell(row=row, column=5, value="N/A")
            
            # Add slope
            ws.cell(row=row, column=6, value=round(point.get('slope', 0), 2))
            
            # Add segment ID (if multi-segment design)
            if self.break_points:
                seg_id = point.get('segment_id', 0) + 1  # Display as 1-indexed
                ws.cell(row=row, column=7, value=seg_id)
            else:
                ws.cell(row=row, column=7, value=1)
            
            # Add coordinates (latitude, longitude)
            ws.cell(row=row, column=8, value=round(point.get('y', 0), 7))  # Latitude
            ws.cell(row=row, column=9, value=round(point.get('x', 0), 7))  # Longitude
            
            # Apply borders and alignment
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                # Number formatting
                if col == 1:
                    cell.number_format = '0.0'  # Station (distance) - 1 decimal
                elif col <= 4:
                    cell.number_format = '0.00'  # Elevations - 2 decimals
                elif col == 5:
                    cell.number_format = '0.0'  # Cover - 1 decimal
                elif col == 6:
                    cell.number_format = '0.00'  # Slope - 2 decimals
                elif col == 7:
                    cell.number_format = '0'  # Segment ID - integer
                elif col in [8, 9]:
                    cell.number_format = '0.0000000'  # Coordinates - 7 decimals
            
            row += 1
        
        data_end_row = row - 1
        
        # Summary Statistics
        row += 1
        ws[f'A{row}'] = "Summary Statistics"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        elevations = [p['elevation'] for p in self.profile_data]
        ws[f'A{row}'] = "Ground Elevation Range:"
        ws[f'B{row}'] = f"{min(elevations):.2f} - {max(elevations):.2f} feet"
        row += 1
        
        ws[f'A{row}'] = "Elevation Change:"
        ws[f'B{row}'] = f"{max(elevations) - min(elevations):.2f} feet"
        row += 1
        
        if self.enable_running_line:
            running_elevations = [p['running_line'] for p in self.profile_data]
            ws[f'A{row}'] = "Running Line Range:"
            ws[f'B{row}'] = f"{min(running_elevations):.2f} - {max(running_elevations):.2f} feet"
            row += 1
        
        ws[f'A{row}'] = "Number of Points:"
        ws[f'B{row}'] = len(self.profile_data)
        row += 1
        
        # Create embedded chart
        chart = LineChart()
        chart.title = "Elevation Profile"
        chart.style = 2
        chart.y_axis.title = "Elevation (ft)"
        chart.x_axis.title = "Distance (ft)"
        chart.width = 30  # Increased from 20 to 30 (50% bigger)
        chart.height = 12
        
        # Configure chart layout to prevent overlapping and reduce left whitespace
        # Manual layout ensures titles and labels fit within chart boundaries
        from openpyxl.chart.layout import Layout, ManualLayout
        
        # Set manual layout for the chart with minimal left margin
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.02,  # Reduced left margin (was 0.05)
                y=0.05,  # Top margin
                w=0.96,  # Increased width to use more space (was 0.9)
                h=0.85   # Height (85% to leave bottom margin for x-axis title)
            )
        )
        
        # Add Ground Elevation series
        ground_data = Reference(ws, min_col=2, min_row=data_start_row-1, max_row=data_end_row)
        station_data = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        
        series = chart.series[0] if len(chart.series) > 0 else None
        chart.add_data(ground_data, titles_from_data=True)
        chart.set_categories(station_data)
        
        # Style ground elevation line (green) - Apply smoothing to match running line behavior
        if len(chart.series) > 0:
            chart.series[0].graphicalProperties.line.solidFill = "2d7a2d"
            chart.series[0].graphicalProperties.line.width = 25000
            # Match the smoothing setting to running line
            chart.series[0].smooth = True  # Ground is always smoothed for natural terrain appearance
        
        # Add Running Line series if enabled
        if self.enable_running_line:
            running_data = Reference(ws, min_col=3, min_row=data_start_row-1, max_row=data_end_row)
            chart.add_data(running_data, titles_from_data=True)
            
            # Style running line (red) - Apply same smoothing as ground elevation
            if len(chart.series) > 1:
                chart.series[1].graphicalProperties.line.solidFill = "cc0000"
                chart.series[1].graphicalProperties.line.width = 25000
                # Match ground elevation smoothing (both should be smoothed for curved appearance)
                chart.series[1].smooth = True
                logging.info("Excel: Both ground and running line smoothing ENABLED for curved display")
        
        # Enable axis numbers (tick labels) on both axes
        # Ensure axes are not deleted - this is critical for showing axis numbers
        chart.y_axis.delete = False
        chart.x_axis.delete = False
        
        # Configure axis formatting to show numbers
        # Set major tick marks to display numbers
        try:
            # Y-axis (NumericAxis) - ensure numbers are visible
            if hasattr(chart.y_axis, 'majorTickMark'):
                chart.y_axis.majorTickMark = 'out'
            if hasattr(chart.y_axis, 'minorTickMark'):
                chart.y_axis.minorTickMark = 'none'
            if hasattr(chart.y_axis, 'tickLblPos'):
                chart.y_axis.tickLblPos = 'nextTo'
            
            # X-axis - ensure numbers are visible
            if hasattr(chart.x_axis, 'majorTickMark'):
                chart.x_axis.majorTickMark = 'out'
            if hasattr(chart.x_axis, 'minorTickMark'):
                chart.x_axis.minorTickMark = 'none'
            if hasattr(chart.x_axis, 'tickLblPos'):
                chart.x_axis.tickLblPos = 'nextTo'
        except Exception as e:
            # If properties don't exist, axes should still show numbers by default
            logging.debug(f"Could not set all axis properties: {e}")
        
        # Position chart to the right of the table (starting at column K - after coordinates and segment)
        chart_position = f'K{header_row}'
        ws.add_chart(chart, chart_position)
        
        # Save workbook
        wb.save(filename)
        logging.info(f"Professional Excel report created with {len(self.profile_data)} data points")


def main():
    """Main entry point"""
    app = ProfileAutomationTool()
    app.mainloop()


if __name__ == "__main__":
    main()

